"""
AIP Step 3 -- 가중치 산출 (Classical AHP / Fuzzy AHP / BWM)
============================================================
입력:
  data/ahp_survey_data.xlsx  (전문가 설문 수집 후 사용)
  ※ 현재는 코드 내 SAMPLE_SURVEY로 동작 테스트

출력:
  data/ahp_weights_YYYYMMDD_HHMMSS.xlsx

[방법론]
  1. Classical AHP  : Saaty(1980) 고유벡터법, CR ≤ 0.1 검정
  2. Fuzzy AHP      : Chang(1996) 확장분석법, TFN 기반
  3. BWM            : Rezaei(2015) 선형 최적화, ξ* 일관성

[계층 구조]
  Level 1 (대기준 3개, 3×3 행렬, 3쌍대비교)
    C1: 경제효율성  C2: 고객서비스  C3: Risk저감

  Level 2 하위기준
    C1: NPV, BCR
    C2: SAIDI저감, ENF저감
    C3: 재무Risk, 안전Risk, 환경Risk

[전문가 집계]
  Classical AHP : 행렬 요소별 기하평균 (Saaty 1980 권장)
  Fuzzy AHP     : TFN 요소별 기하평균
  BWM           : 전문가별 가중치의 단순 평균

실행: py -3 aip_03_ahp.py
"""

import numpy as np
import pandas as pd
from scipy.optimize import linprog
from pathlib import Path
from datetime import datetime
import warnings
warnings.filterwarnings("ignore")

# -- 경로
BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"

# ======================================================================
# 0. 상수 정의
# ======================================================================

# Saaty RI (Random Consistency Index) — n=1~10
RI_TABLE = {1: 0.00, 2: 0.00, 3: 0.58, 4: 0.90, 5: 1.12,
            6: 1.24, 7: 1.32, 8: 1.41, 9: 1.45, 10: 1.49}

# Fuzzy AHP 언어척도 → TFN (Chang 1996)
LINGUISTIC_TFN = {
    'AI': (7.0, 9.0, 9.0),   # Absolutely Important
    'VI': (5.0, 7.0, 9.0),   # Very strongly Important
    'SI': (3.0, 5.0, 7.0),   # Strongly Important
    'WI': (1.0, 3.0, 5.0),   # Weakly Important
    'EI': (1.0, 1.0, 1.0),   # Equally Important
}

# 기준 이름 (Excel 출력·로그용)
CRITERIA = {
    'level1': ['C1_경제효율성', 'C2_고객서비스', 'C3_Risk저감'],
    'c1_sub': ['C1-1_NPV', 'C1-2_BCR'],
    'c2_sub': ['C2-1_SAIDI저감', 'C2-2_ENF저감'],
    'c3_sub': ['C3-1_재무Risk', 'C3-2_안전Risk', 'C3-3_환경Risk'],
}

# 계층별 기준 수
N_CRITERIA = {'level1': 3, 'c1_sub': 2, 'c2_sub': 2, 'c3_sub': 3}

# 계층 표시명
LEVEL_NAMES = {
    'level1': '대기준 (Level 1)',
    'c1_sub': 'C1 하위기준',
    'c2_sub': 'C2 하위기준',
    'c3_sub': 'C3 하위기준',
}


# ======================================================================
# 1. Classical AHP
# ======================================================================

class ClassicalAHP:
    """Saaty(1980) 계층분석법 — 고유벡터 가중치 산출"""

    @staticmethod
    def build_matrix(n: int, comparisons: list) -> np.ndarray:
        """
        n×n 쌍대비교 행렬 구성
        comparisons: [(i, j, value)]
          value > 1  : 기준 i가 기준 j보다 value배 중요
          value < 1  : 기준 j가 기준 i보다 1/value배 중요
        """
        A = np.eye(n, dtype=float)
        for i, j, v in comparisons:
            A[i, j] = float(v)
            A[j, i] = 1.0 / float(v)
        return A

    @staticmethod
    def eigenvector_weights(A: np.ndarray):
        """최대 고유벡터로 가중치 산출 → (weights, lambda_max)"""
        eigenvalues, eigenvectors = np.linalg.eig(A)
        max_idx = int(np.argmax(eigenvalues.real))
        lambda_max = float(eigenvalues[max_idx].real)
        w = eigenvectors[:, max_idx].real
        w = np.abs(w) / np.abs(w).sum()
        return w, lambda_max

    @staticmethod
    def consistency_ratio(n: int, lambda_max: float):
        """일관성 비율 CR 산출 → (CI, CR)"""
        CI = (lambda_max - n) / (n - 1) if n > 1 else 0.0
        ri = RI_TABLE.get(n, 1.49)
        CR = CI / ri if ri > 0 else 0.0
        return CI, CR

    @staticmethod
    def aggregate_matrices(matrices: list) -> np.ndarray:
        """여러 전문가 행렬의 기하평균 집계"""
        stacked = np.array(matrices, dtype=float)
        return np.exp(np.mean(np.log(np.maximum(stacked, 1e-12)), axis=0))

    def compute(self, n: int, comparisons: list) -> dict:
        A = self.build_matrix(n, comparisons)
        w, lambda_max = self.eigenvector_weights(A)
        CI, CR = self.consistency_ratio(n, lambda_max)
        return {
            'matrix': A, 'weights': w,
            'lambda_max': lambda_max, 'CI': CI, 'CR': CR,
            'consistent': CR <= 0.1,
        }


# ======================================================================
# 2. Fuzzy AHP (Chang 1996 Extent Analysis)
# ======================================================================

class FuzzyAHP:
    """Chang(1996) 확장분석법 기반 퍼지 AHP"""

    @staticmethod
    def to_tfn(term: str, reverse: bool = False):
        """언어 척도 → TFN (l, m, u)"""
        tfn = LINGUISTIC_TFN[term]
        if reverse:
            return (1.0 / tfn[2], 1.0 / tfn[1], 1.0 / tfn[0])
        return tfn

    @staticmethod
    def build_fuzzy_matrix(n: int, comparisons: list) -> np.ndarray:
        """
        n×n×3 TFN 행렬 구성
        comparisons: [(i, j, term_or_tfn)]
          term_or_tfn: 언어척도 문자열('AI','VI','SI','WI','EI') 또는
                       (l, m, u) 튜플로 직접 입력 가능
          i가 j보다 중요할 때의 척도를 입력; 역수는 자동 계산
        """
        A = np.zeros((n, n, 3), dtype=float)
        for k in range(n):
            A[k, k] = (1.0, 1.0, 1.0)

        for i, j, term in comparisons:
            if isinstance(term, str):
                tfn   = FuzzyAHP.to_tfn(term, reverse=False)
                tfn_r = FuzzyAHP.to_tfn(term, reverse=True)
            else:
                tfn   = tuple(float(x) for x in term)
                tfn_r = (1.0/tfn[2], 1.0/tfn[1], 1.0/tfn[0])
            A[i, j] = tfn
            A[j, i] = tfn_r
        return A

    @staticmethod
    def aggregate_fuzzy_matrices(matrices: list) -> np.ndarray:
        """TFN 요소별 기하평균으로 전문가 퍼지 행렬 집계"""
        stacked = np.array(matrices, dtype=float)
        return np.exp(np.mean(np.log(np.maximum(stacked, 1e-12)), axis=0))

    # Chang(1996) 확장분석법의 알려진 단점: 일부 기준의 가중치가 0으로 수렴.
    # 전문가 수가 적거나 응답 분포가 편향될 때 발생. epsilon으로 하한 보정.
    EPSILON = 1e-4

    @staticmethod
    def chang_extent_analysis(A: np.ndarray) -> np.ndarray:
        """
        Chang(1996) 확장분석법으로 퍼지 가중치 산출
        A: (n, n, 3) TFN 행렬
        0 가중치 문제: epsilon 하한 보정 후 재정규화
        """
        n = A.shape[0]

        # 행별 퍼지 합 S_i
        row_sums = A.sum(axis=1)                    # (n, 3)
        total    = row_sums.sum(axis=0)             # (3,) = 전체 합
        total_inv = np.array([1.0/total[2], 1.0/total[1], 1.0/total[0]])

        # S_i = row_sum_i ⊗ total_inv
        S = np.array([
            (row_sums[i, 0] * total_inv[0],
             row_sums[i, 1] * total_inv[1],
             row_sums[i, 2] * total_inv[2])
            for i in range(n)
        ])

        def possibility(m1, m2):
            """V(M1 ≥ M2): TFN m1=(l1,mid1,u1), m2=(l2,mid2,u2)"""
            if m1[1] >= m2[1]:
                return 1.0
            if m2[0] >= m1[2]:
                return 0.0
            denom = (m1[1] - m1[2]) - (m2[1] - m2[0])
            return (m2[0] - m1[2]) / denom if denom != 0 else 0.0

        # d'(A_i) = min V(S_i ≥ S_j), j ≠ i
        d = np.array([min(possibility(S[i], S[j]) for j in range(n) if j != i)
                      for i in range(n)])

        # 0 가중치 epsilon 하한 보정 (Chang 1996의 구조적 한계 보완)
        d = np.maximum(d, FuzzyAHP.EPSILON)
        total_d = d.sum()
        return d / total_d

    def compute(self, n: int, comparisons: list) -> dict:
        A = self.build_fuzzy_matrix(n, comparisons)
        weights = self.chang_extent_analysis(A)
        return {'fuzzy_matrix': A, 'weights': weights}


# ======================================================================
# 3. BWM (Rezaei 2015 — 선형 BWM)
# ======================================================================

class BWM:
    """
    Rezaei(2015) Best-Worst Method
    선형 정식화(L-BWM): scipy.optimize.linprog 사용
    """

    @staticmethod
    def solve(n: int, best_idx: int, worst_idx: int,
              bo_vector: list, ow_vector: list) -> dict:
        """
        변수: x = [w_0, ..., w_{n-1}, ξ]  (총 n+1개)
        목적: min ξ (= x[n])

        제약:
          ∀j: w_B - a_Bj·w_j ≤ ξ   →  w_B  - a_Bj·w_j  - ξ ≤ 0
          ∀j: a_Bj·w_j - w_B ≤ ξ   → -w_B  + a_Bj·w_j  - ξ ≤ 0
          ∀j: w_j - a_jW·w_W ≤ ξ   →  w_j  - a_jW·w_W  - ξ ≤ 0
          ∀j: a_jW·w_W - w_j ≤ ξ   → -w_j  + a_jW·w_W  - ξ ≤ 0
          Σw_j = 1,  w_j ≥ 0,  ξ ≥ 0
        """
        n_vars = n + 1
        xi_idx = n

        c = np.zeros(n_vars)
        c[xi_idx] = 1.0

        A_ub_rows, b_ub_vals = [], []

        for j in range(n):
            aBj = float(bo_vector[j])
            ajW = float(ow_vector[j])

            for sign, col_B, col_j in [(1, best_idx, j), (-1, best_idx, j)]:
                row = np.zeros(n_vars)
                row[col_B]  =  sign * 1.0
                row[col_j]  = -sign * aBj
                row[xi_idx] = -1.0
                A_ub_rows.append(row)
                b_ub_vals.append(0.0)

            for sign, col_j_v, col_W in [(1, j, worst_idx), (-1, j, worst_idx)]:
                row = np.zeros(n_vars)
                row[col_j_v] =  sign * 1.0
                row[col_W]   = -sign * ajW
                row[xi_idx]  = -1.0
                A_ub_rows.append(row)
                b_ub_vals.append(0.0)

        A_eq = np.zeros((1, n_vars))
        A_eq[0, :n] = 1.0
        b_eq = [1.0]

        bounds = [(0.0, None)] * n + [(0.0, None)]

        res = linprog(c, A_ub=A_ub_rows, b_ub=b_ub_vals,
                      A_eq=A_eq, b_eq=b_eq,
                      bounds=bounds, method='highs')

        if res.success:
            return {
                'weights': res.x[:n],
                'xi': float(res.x[xi_idx]),
                'consistent': float(res.x[xi_idx]) <= 0.1,
                'status': 'optimal',
            }
        return {
            'weights': np.ones(n) / n,
            'xi': np.nan,
            'consistent': False,
            'status': res.message,
        }

    @staticmethod
    def aggregate_weights(results: list) -> np.ndarray:
        """전문가 BWM 가중치의 단순 평균"""
        return np.mean([r['weights'] for r in results], axis=0)

    def compute(self, n: int, best_idx: int, worst_idx: int,
                bo_vector: list, ow_vector: list) -> dict:
        return self.solve(n, best_idx, worst_idx, bo_vector, ow_vector)


# ======================================================================
# 4. 전역 가중치 합성
# ======================================================================

def compute_global_weights(w1: np.ndarray, wc1: np.ndarray,
                            wc2: np.ndarray, wc3: np.ndarray) -> dict:
    """전역 가중치 = 대기준 가중치 × 하위기준 가중치"""
    return {
        'C1-1_NPV':       float(w1[0] * wc1[0]),
        'C1-2_BCR':       float(w1[0] * wc1[1]),
        'C2-1_SAIDI저감': float(w1[1] * wc2[0]),
        'C2-2_ENF저감':   float(w1[1] * wc2[1]),
        'C3-1_재무Risk':  float(w1[2] * wc3[0]),
        'C3-2_안전Risk':  float(w1[2] * wc3[1]),
        'C3-3_환경Risk':  float(w1[2] * wc3[2]),
    }


# ======================================================================
# 5. 샘플 설문 데이터 (전문가 실설문 전 테스트용)
# ======================================================================
# 가정 판단 방향: 경제효율성 > Risk저감 > 고객서비스
# 실제 전문가 응답 수집 후 이 딕셔너리를 교체하거나
# load_survey_from_excel()로 대체

SAMPLE_SURVEY = {
    'expert_1': {
        'ahp': {
            'level1': [(0, 1, 3), (0, 2, 2), (2, 1, 2)],   # C1>C3>C2
            'c1_sub': [(0, 1, 3)],                           # NPV>BCR
            'c2_sub': [(0, 1, 2)],                           # SAIDI>ENF
            'c3_sub': [(0, 1, 3), (0, 2, 5), (1, 2, 2)],   # 재무>안전>환경
        },
        'fuzzy': {
            'level1': [(0, 1, 'SI'), (0, 2, 'WI'), (2, 1, 'WI')],
            'c1_sub': [(0, 1, 'SI')],
            'c2_sub': [(0, 1, 'WI')],
            'c3_sub': [(0, 1, 'SI'), (0, 2, 'VI'), (1, 2, 'WI')],
        },
        'bwm': {
            # bo_vector: Best(인덱스) 대비 [C1, C2, C3] 중요도
            # ow_vector: [C1, C2, C3] 대비 Worst(인덱스) 중요도
            'level1': {'best': 0, 'worst': 1,
                       'bo': [1, 3, 2], 'ow': [3, 1, 2]},
            'c1_sub': {'best': 0, 'worst': 1,
                       'bo': [1, 3], 'ow': [3, 1]},
            'c2_sub': {'best': 0, 'worst': 1,
                       'bo': [1, 2], 'ow': [2, 1]},
            'c3_sub': {'best': 0, 'worst': 2,
                       'bo': [1, 3, 5], 'ow': [5, 2, 1]},
        },
    },
    # 실제 전문가 응답 추가 시 아래 형식으로 확장:
    # 'expert_2': { 'ahp': {...}, 'fuzzy': {...}, 'bwm': {...} },
}


# ======================================================================
# 6. Excel 설문 응답 로딩 (실제 설문 수집 후 사용)
# ======================================================================

def load_survey_from_excel(path: Path) -> dict:
    """
    설문 응답 Excel 파일 로딩 (미구현 — 실제 설문 수집 후 완성)
    파일 형식: 시트별 전문가, 각 시트에 AHP/Fuzzy/BWM 응답
    """
    raise NotImplementedError("실제 설문 수집 후 구현 예정")


# ======================================================================
# 7. 메인 파이프라인
# ======================================================================

def run_pipeline(survey: dict) -> dict:
    """3종 방법론 가중치 산출 통합 파이프라인"""

    ahp  = ClassicalAHP()
    fahp = FuzzyAHP()
    bwm  = BWM()

    results = {'classical_ahp': {}, 'fuzzy_ahp': {}, 'bwm': {}}

    # ── 1. Classical AHP ─────────────────────────────────────────
    print("\n[1/3] Classical AHP")

    ahp_matrices = {lvl: [] for lvl in N_CRITERIA}
    for exp in survey.values():
        for lvl, n in N_CRITERIA.items():
            ahp_matrices[lvl].append(
                ahp.build_matrix(n, exp['ahp'][lvl])
            )

    for lvl, n in N_CRITERIA.items():
        mats = ahp_matrices[lvl]
        agg  = ahp.aggregate_matrices(mats) if len(mats) > 1 else mats[0]
        w, lam = ahp.eigenvector_weights(agg)
        CI, CR  = ahp.consistency_ratio(n, lam)
        results['classical_ahp'][lvl] = {
            'matrix': agg, 'weights': w,
            'lambda_max': lam, 'CI': CI, 'CR': CR,
            'consistent': CR <= 0.1, 'n_experts': len(mats),
        }
        tag = "[OK]" if CR <= 0.1 else "[!] CR > 0.1 -- 재응답 필요"
        print(f"  {LEVEL_NAMES[lvl]}: CR={CR:.4f} {tag}")
        for nm, wt in zip(CRITERIA[lvl], w):
            print(f"    {nm}: {wt:.4f}")

    # ── 2. Fuzzy AHP ─────────────────────────────────────────────
    print("\n[2/3] Fuzzy AHP (Chang 1996)")

    fuzzy_matrices = {lvl: [] for lvl in N_CRITERIA}
    for exp in survey.values():
        for lvl, n in N_CRITERIA.items():
            fuzzy_matrices[lvl].append(
                fahp.build_fuzzy_matrix(n, exp['fuzzy'][lvl])
            )

    for lvl, n in N_CRITERIA.items():
        mats = fuzzy_matrices[lvl]
        agg  = fahp.aggregate_fuzzy_matrices(mats) if len(mats) > 1 else mats[0]
        w    = fahp.chang_extent_analysis(agg)
        results['fuzzy_ahp'][lvl] = {
            'fuzzy_matrix': agg, 'weights': w, 'n_experts': len(mats),
        }
        print(f"  {LEVEL_NAMES[lvl]}:")
        for nm, wt in zip(CRITERIA[lvl], w):
            print(f"    {nm}: {wt:.4f}")

    # ── 3. BWM ───────────────────────────────────────────────────
    print("\n[3/3] BWM (Rezaei 2015)")

    bwm_exp_results = {lvl: [] for lvl in N_CRITERIA}
    for exp in survey.values():
        for lvl, n in N_CRITERIA.items():
            b = exp['bwm'][lvl]
            r = bwm.solve(n, b['best'], b['worst'], b['bo'], b['ow'])
            bwm_exp_results[lvl].append(r)

    for lvl in N_CRITERIA:
        exp_list = bwm_exp_results[lvl]
        agg_w    = BWM.aggregate_weights(exp_list)
        xi_vals  = [r['xi'] for r in exp_list if not np.isnan(r['xi'])]
        avg_xi   = float(np.mean(xi_vals)) if xi_vals else np.nan
        results['bwm'][lvl] = {
            'weights': agg_w, 'xi': avg_xi,
            'consistent': avg_xi <= 0.1 if not np.isnan(avg_xi) else False,
            'expert_details': exp_list, 'n_experts': len(exp_list),
        }
        tag = "[OK]" if avg_xi <= 0.1 else "[!]"
        print(f"  {LEVEL_NAMES[lvl]}: ξ*={avg_xi:.4f} {tag}")
        for nm, wt in zip(CRITERIA[lvl], agg_w):
            print(f"    {nm}: {wt:.4f}")

    # ── 4. 전역 가중치 합성 ──────────────────────────────────────
    print("\n[전역 가중치]")
    for method in ('classical_ahp', 'fuzzy_ahp', 'bwm'):
        r = results[method]
        gw = compute_global_weights(
            r['level1']['weights'], r['c1_sub']['weights'],
            r['c2_sub']['weights'], r['c3_sub']['weights'],
        )
        results[method]['global_weights'] = gw
        print(f"\n  [{method.upper()}]")
        for k, v in gw.items():
            print(f"    {k}: {v:.4f}")
        print(f"    합계: {sum(gw.values()):.4f}")

    return results


# ======================================================================
# 8. Excel 보고서
# ======================================================================

def save_report(results: dict, out_path: Path):
    """3종 방법론 가중치 비교 Excel 보고서 저장"""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # -- 스타일 상수
    C_HEADER = "1F497D"   # 헤더 진파랑
    C_SUB    = "4472C4"   # 서브헤더 파랑
    C_GOOD   = "E2EFDA"   # CR 통과 연녹
    C_WARN   = "FFEEBA"   # CR 초과 연황
    C_ALT    = "F5F5F5"   # 행 교대색

    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))

    def hcell(ws, r, c, val, bg=C_HEADER, fc="FFFFFF", bold=True, wrap=True):
        cell = ws.cell(row=r, column=c, value=val)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=bold, color=fc, size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=wrap)
        cell.border = thin
        return cell

    def dcell(ws, r, c, val, fmt=None, bg=None):
        cell = ws.cell(row=r, column=c, value=val)
        if fmt:   cell.number_format = fmt
        if bg:    cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin
        return cell

    GLOBAL_ORDER = ['C1-1_NPV', 'C1-2_BCR', 'C2-1_SAIDI저감',
                    'C2-2_ENF저감', 'C3-1_재무Risk', 'C3-2_안전Risk', 'C3-3_환경Risk']
    METHODS = [('classical_ahp', 'Classical AHP'),
               ('fuzzy_ahp',     'Fuzzy AHP'),
               ('bwm',           'BWM')]

    # ── 시트 1: 전역 가중치 요약 ────────────────────────────────
    ws = wb.active
    ws.title = "요약_전역가중치"
    ws.column_dimensions['A'].width = 24
    for col in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 18
    ws.row_dimensions[1].height = 32

    hcell(ws, 1, 1, "하위기준 / Sub-Criterion")
    for ci, (_, mlabel) in enumerate(METHODS):
        hcell(ws, 1, ci + 2, mlabel)
    hcell(ws, 1, 5, "평균 (Average)")

    for ri, crit in enumerate(GLOBAL_ORDER, start=2):
        bg = C_ALT if ri % 2 == 0 else "FFFFFF"
        dcell(ws, ri, 1, crit, bg=bg)
        vals = []
        for ci, (mkey, _) in enumerate(METHODS):
            v = results[mkey]['global_weights'].get(crit, 0.0)
            dcell(ws, ri, ci + 2, v, fmt='0.0000', bg=bg)
            vals.append(v)
        dcell(ws, ri, 5, float(np.mean(vals)), fmt='0.0000', bg=bg)

    r_total = len(GLOBAL_ORDER) + 2
    hcell(ws, r_total, 1, "합계", bg=C_SUB)
    for ci, (mkey, _) in enumerate(METHODS):
        t = sum(results[mkey]['global_weights'].values())
        dcell(ws, r_total, ci + 2, t, fmt='0.0000', bg="D9E1F2")
    dcell(ws, r_total, 5, "1.0000", bg="D9E1F2")

    # ── 시트 2: Classical AHP 상세 ──────────────────────────────
    ws2 = wb.create_sheet("Classical_AHP_상세")
    ws2.column_dimensions['A'].width = 26

    row = 1
    for lvl, lname in LEVEL_NAMES.items():
        r   = results['classical_ahp'][lvl]
        nms = CRITERIA[lvl]
        n   = len(nms)

        ws2.merge_cells(start_row=row, start_column=1,
                        end_row=row, end_column=n + 2)
        hcell(ws2, row, 1,
              f"▶ {lname}  |  λ_max={r['lambda_max']:.4f}  "
              f"CI={r['CI']:.4f}  CR={r['CR']:.4f}  "
              f"전문가={r['n_experts']}명")
        row += 1

        hcell(ws2, row, 1, "기준", bg=C_SUB)
        for j, nm in enumerate(nms):
            hcell(ws2, row, j + 2, nm, bg=C_SUB)
        hcell(ws2, row, n + 2, "가중치 w", bg=C_SUB)
        row += 1

        for i, nm in enumerate(nms):
            bg_row = C_ALT if i % 2 == 0 else "FFFFFF"
            dcell(ws2, row, 1, nm, bg=bg_row)
            for j in range(n):
                dcell(ws2, row, j + 2, round(float(r['matrix'][i, j]), 4),
                      fmt='0.0000', bg=bg_row)
            bg_w = C_GOOD if r['consistent'] else C_WARN
            dcell(ws2, row, n + 2, round(float(r['weights'][i]), 4),
                  fmt='0.0000', bg=bg_w)
            row += 1

        status = ("일관성 있음 (CR ≤ 0.1)"
                  if r['consistent'] else "[!] 일관성 부족 (CR > 0.1) -- 재응답 필요")
        dcell(ws2, row, 1, status, bg=C_GOOD if r['consistent'] else C_WARN)
        row += 2

    # ── 시트 3: Fuzzy AHP 상세 ──────────────────────────────────
    ws3 = wb.create_sheet("Fuzzy_AHP_상세")
    ws3.column_dimensions['A'].width = 26

    row = 1
    for lvl, lname in LEVEL_NAMES.items():
        r   = results['fuzzy_ahp'][lvl]
        nms = CRITERIA[lvl]
        n   = len(nms)
        n_col = n * 3 + 3

        ws3.merge_cells(start_row=row, start_column=1,
                        end_row=row, end_column=n_col)
        hcell(ws3, row, 1, f"▶ {lname}  |  Chang(1996) Extent Analysis  전문가={r['n_experts']}명")
        row += 1

        # 기준 헤더 (3열씩 병합)
        hcell(ws3, row, 1, "기준", bg=C_SUB)
        col = 2
        for nm in nms:
            ws3.merge_cells(start_row=row, start_column=col,
                            end_row=row, end_column=col + 2)
            hcell(ws3, row, col, nm, bg=C_SUB)
            col += 3
        hcell(ws3, row, col, "가중치 w", bg=C_SUB)
        row += 1

        # l / m / u 서브헤더
        dcell(ws3, row, 1, "", bg=C_SUB)
        col = 2
        for _ in nms:
            for lbl in ['l', 'm', 'u']:
                hcell(ws3, row, col, lbl, bg=C_SUB, bold=False)
                col += 1
        hcell(ws3, row, col, "", bg=C_SUB, bold=False)
        row += 1

        # 행렬 값
        for i, nm in enumerate(nms):
            bg_row = C_ALT if i % 2 == 0 else "FFFFFF"
            dcell(ws3, row, 1, nm, bg=bg_row)
            col = 2
            for j in range(n):
                for k in range(3):
                    dcell(ws3, row, col,
                          round(float(r['fuzzy_matrix'][i, j, k]), 4),
                          fmt='0.0000', bg=bg_row)
                    col += 1
            dcell(ws3, row, col, round(float(r['weights'][i]), 4),
                  fmt='0.0000', bg=C_GOOD)
            row += 1

        row += 1

    # ── 시트 4: BWM 상세 ─────────────────────────────────────────
    ws4 = wb.create_sheet("BWM_상세")
    ws4.column_dimensions['A'].width = 26

    row = 1
    for lvl, lname in LEVEL_NAMES.items():
        r   = results['bwm'][lvl]
        nms = CRITERIA[lvl]
        n   = len(nms)

        ws4.merge_cells(start_row=row, start_column=1,
                        end_row=row, end_column=n + 2)
        hcell(ws4, row, 1,
              f"▶ {lname}  |  BWM (Rezaei 2015)  ξ*={r['xi']:.4f}  전문가={r['n_experts']}명")
        row += 1

        hcell(ws4, row, 1, "구분", bg=C_SUB)
        for j, nm in enumerate(nms):
            hcell(ws4, row, j + 2, nm, bg=C_SUB)
        hcell(ws4, row, n + 2, "ξ*", bg=C_SUB)
        row += 1

        for ei, exp_r in enumerate(r['expert_details']):
            bg_row = "E8F0FE" if ei % 2 == 0 else "FFFFFF"
            dcell(ws4, row, 1, f"전문가 {ei + 1}", bg=bg_row)
            for j in range(n):
                dcell(ws4, row, j + 2,
                      round(float(exp_r['weights'][j]), 4), fmt='0.0000', bg=bg_row)
            dcell(ws4, row, n + 2,
                  round(float(exp_r['xi']), 4) if not np.isnan(exp_r['xi']) else "오류",
                  fmt='0.0000', bg=bg_row)
            row += 1

        bg_agg = C_GOOD if r['consistent'] else C_WARN
        hcell(ws4, row, 1, "집계 (평균)", bg=C_SUB)
        for j in range(n):
            dcell(ws4, row, j + 2,
                  round(float(r['weights'][j]), 4), fmt='0.0000', bg=bg_agg)
        dcell(ws4, row, n + 2,
              round(r['xi'], 4) if not np.isnan(r['xi']) else "오류",
              fmt='0.0000', bg=bg_agg)
        row += 2

    wb.save(out_path)
    print(f"\n[보고서 저장] {out_path}")


# ======================================================================
# 9. 그룹별 분석 (선택 실행)
# ======================================================================

def run_group_analysis(survey: dict, groups: dict, out_path: Path):
    """그룹별 가중치 산출 후 Excel 시트로 비교"""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    GROUP_COLORS = {
        'A_경제경영':   'FFF2CC',
        'B_현장직':     'E2EFDA',
        'C_계통엔지니어': 'DDEEFF',
        'D_자산관리':   'F2E5FF',
        'E_규제정책':   'FFE5E5',
    }
    GLOBAL_ORDER = ['C1-1_NPV', 'C1-2_BCR', 'C2-1_SAIDI저감',
                    'C2-2_ENF저감', 'C3-1_재무Risk', 'C3-2_안전Risk', 'C3-3_환경Risk']

    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "그룹별_가중치_비교"
    ws.column_dimensions['A'].width = 24

    col_cursor = 2
    group_col_map = {}

    print("\n[그룹별 분석]")
    for gname, exp_ids in groups.items():
        group_survey = {eid: survey[eid] for eid in exp_ids if eid in survey}
        if not group_survey:
            continue
        print(f"  {gname} ({len(group_survey)}명) ...")
        r = run_pipeline(group_survey)
        group_col_map[gname] = (col_cursor, r)

        ws.column_dimensions[chr(64 + col_cursor)].width = 16
        col_cursor += 1

    # 헤더 행
    ws.cell(row=1, column=1, value="하위기준").font = Font(bold=True)
    ws.cell(row=1, column=1).border = thin
    for gname, (col, _) in group_col_map.items():
        c = ws.cell(row=1, column=col, value=gname)
        c.font = Font(bold=True, size=9)
        c.fill = PatternFill("solid", fgColor=GROUP_COLORS.get(gname, 'FFFFFF'))
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = thin

    # 데이터 행
    for ri, crit in enumerate(GLOBAL_ORDER, start=2):
        ws.cell(row=ri, column=1, value=crit).border = thin
        for gname, (col, r) in group_col_map.items():
            v = r['classical_ahp']['global_weights'].get(crit, 0.0)
            c = ws.cell(row=ri, column=col, value=round(v, 4))
            c.number_format = '0.0000'
            c.fill = PatternFill("solid", fgColor=GROUP_COLORS.get(gname, 'FFFFFF'))
            c.alignment = Alignment(horizontal='center')
            c.border = thin

    wb.save(out_path)
    print(f"  저장 완료: {out_path}")


# ======================================================================
# 10. 진입점
# ======================================================================

if __name__ == "__main__":
    print("=" * 70)
    print("  AIP Step 3 -- 가중치 산출 (Classical AHP / Fuzzy AHP / BWM)")
    print("=" * 70)

    # 설문 데이터 로드 (20명 가상 전문가)
    from survey_data import SURVEY_20, EXPERT_GROUPS
    survey = SURVEY_20
    groups = EXPERT_GROUPS

    # 전문가 수 확인
    n_exp = len(survey)
    print(f"\n[설문 데이터] 가상 전문가 {n_exp}명 (그룹: {len(groups)}개)")
    for gname, ids in groups.items():
        print(f"  {gname}: {len(ids)}명")

    # 파이프라인 실행 (전체 통합)
    print("\n[전체 통합 분석]")
    results = run_pipeline(survey)

    # 보고서 저장 (전체 통합)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = DATA_DIR / f"ahp_weights_{ts}.xlsx"
    save_report(results, out_path)

    # 그룹별 비교 보고서
    group_path = DATA_DIR / f"ahp_group_compare_{ts}.xlsx"
    run_group_analysis(survey, groups, group_path)

    print("\n" + "=" * 70)
    print("  완료")
    print("=" * 70)
