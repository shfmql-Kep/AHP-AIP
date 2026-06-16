"""
AIP Step 1 -- 배전설비 자산 목록 랜덤 생성 및 Excel 고정
===========================================================
생성 결과:
  data/input_assets.xlsx   ← 자산별 입력 속성 (나이, 위치, 상태 등)
  data/cnaim_params.xlsx   ← CNAIM 파라미터 (K값, CoF 기준, 비율 근거 등)

[설비 비율 근거]
  - KEPCO 전력통계정보시스템 2022 (변압기 비중 높음)
  - UK Power Networks RIIO-ED1 Business Plan 2015-2023
  - Western Power Distribution Asset Management Plan 2015
  - CIGRE Technical Brochure 309 (Life Cycle Cost of HV Assets)

[자산군 CNAIM 매핑]
  CNAIM v2.1, April 2021, Table 20(NEL) · Table 21(K값)

실행: py -3 aip_01_generate_input.py
"""

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime
import sys

import cnaim_condition_tables as cct

# -- 경로
BASE_DIR   = Path(__file__).resolve().parent.parent
DATA_DIR   = BASE_DIR / "data"
OUT_ASSETS = DATA_DIR / "input_assets.xlsx"
OUT_PARAMS = DATA_DIR / "cnaim_params.xlsx"
DATA_DIR.mkdir(exist_ok=True)

RANDOM_SEED = 42
# 주상변압기 10,000기 기준 역산: round(10_000 / 0.370) = 27,027
# 비율 구조를 유지하면서 주상변압기가 정확히 1만대가 되도록 설정
TR_P_TARGET  = 10_000
TOTAL_ASSETS = round(TR_P_TARGET / 0.370)  # = 27,027

# ==================================================================
# 1. 자산군 정의 (12종)
# ==================================================================
# 필드 순서: (한국명, CNAIM카테고리, 코드, K_pct, C, HSL, NEL_년,
#             비율, 참조고객수, 교체비용_만원, 공사시간_h, 투입인력_인일,
#             CoF_재무_천원, CoF_안전_천원, CoF_환경_천원, CoF_계통_천원)

ASSET_DEFS = {
    # -- 변압기류 ----------------------------------------------------
    "주상변압기": (
        "6.6/11kV Transformer (GM)", "TR_P",
        0.0078, 1.087, 4, 60,
        0.370,  # 37.0% -- 한국 22.9kV 핵심 자산, KEPCO 최다
        40,     # 참조고객수
        250, 8, 2,     # 교체비용/공사시간/투입인력
        14_000, 7_000, 5_500, 9_000,  # CoF F/S/E/NP (천원)
    ),
    "지상변압기": (
        "6.6/11kV Transformer (GM)", "TR_G",
        0.0078, 1.087, 4, 60,
        0.060,  # 6.0% -- 도심 지상형
        100,
        800, 16, 4,
        22_000, 7_000, 6_000, 14_000,
    ),
    "지중변압기": (
        "6.6/11kV Transformer (SfL)", "TR_V",
        0.0078, 1.087, 4, 60,
        0.030,  # 3.0% -- 도심 지하 기계실형
        120,
        1_000, 20, 5,
        24_000, 8_000, 6_500, 16_000,
    ),
    # -- 개폐기류 ----------------------------------------------------
    "가공개폐기": (
        "HV Switchgear Distribution", "SW_OH",
        0.0067, 1.087, 4, 55,
        0.130,  # 13.0% -- 자동재폐로 + 구분개폐기
        60,
        350, 8, 2,
        11_000, 16_000, 2_000, 22_000,
    ),
    "지중개폐기_RMU": (
        "HV Switchgear Distribution", "SW_UG",
        0.0067, 1.087, 4, 55,
        0.050,  # 5.0% -- Ring Main Unit
        120,
        600, 16, 4,
        13_000, 18_000, 2_500, 28_000,
    ),
    "특고압차단기": (
        "HV Switchgear Primary", "CB_HV",
        0.0052, 1.087, 4, 55,
        0.030,  # 3.0% -- 진공차단기(VCB), 1차 변전소급
        200,
        1_200, 24, 6,
        8_000, 24_000, 1_600, 42_000,
    ),
    # -- 배전선로류 --------------------------------------------------
    "가공배전선로": (
        "Poles (Wood) -- OHL Section", "OHL",
        0.0285, 1.087, 4, 55,
        0.130,  # 13.0% -- 500m 구간 단위
        30,
        450, 12, 3,
        7_000, 2_000, 1_500, 4_500,
    ),
    "지중케이블": (
        "Non-Pressurised Cable", "UGC",
        0.0050, 1.087, 4, 40,  # K: CIGRE TB 379/B1.10 기반 (UK 0.0658% → 한국 22.9kV XLPE 0.0050%)
        0.070,  # 7.0% -- XLPE 케이블, 300m 구간 단위 (PoF/km → ×길이)
        40,
        600, 16, 4,
        9_000, 1_500, 2_000, 5_000,
    ),
    # -- 지지물류 ----------------------------------------------------
    "목주": (
        "Poles (Wood)", "PW",
        0.0285, 1.087, 4, 55,
        0.050,  # 5.0% -- 농촌형 목주
        5,
        50, 4, 1,
        500, 200, 100, 300,
    ),
    "콘크리트주": (
        "Poles (Concrete)", "PC",
        0.0285, 1.087, 4, 60,
        0.060,  # 6.0% -- 가장 일반적인 지지물
        5,
        60, 4, 1,
        600, 200, 100, 300,
    ),
    "철주": (
        "Poles (Steel)", "PS",
        0.0285, 1.087, 4, 50,
        0.010,  # 1.0% -- 간선/특수구간
        5,
        80, 6, 2,
        800, 250, 120, 350,
    ),
    # -- 저압 설비류 --------------------------------------------------
    "LV배전반_UGB": (
        "LV Underground General Bond", "LV_UGB",
        0.0077, 1.087, 4, 45,
        0.010,  # 1.0% -- 저압 분기함
        30,
        200, 6, 2,
        3_000, 500, 300, 1_500,
    ),
}

# 비율 합계 검증
_total_ratio = sum(v[6] for v in ASSET_DEFS.values())
assert abs(_total_ratio - 1.0) < 0.001, f"비율 합계 오류: {_total_ratio:.4f}"

# ==================================================================
# 2. 상태 조건 정의 및 목표 HI 분포
# ==================================================================
# 컨디션 입력 캘리브레이션 값은 cnaim_condition_tables.py에 정의.

# 목표 HI 등급 분포: 전체 자산군 공통으로 동일하게 적용
#   → ILP 최적화 시 자산유형 간 진짜 교환 관계(trade-off)가 생기도록
HI_PROBS  = [0.25, 0.35, 0.20, 0.14, 0.06]   # HI1~HI5 순서
HI_LABELS = ["HI1", "HI2", "HI3", "HI4", "HI5"]

# HI 등급별 잠재 중증도(z) 범위 (자산유형별 캘리브레이션)
#
# HV_TR: FFA Collar 공식 EQ.26 역산
#   FFA Collar = 2.39 × S^0.66  (S = FFA ppm, range (0.3, 7.5))
#   z=0.15 → S≈1.4ppm → Collar≈3.0 (HI1/HI2 경계)
#   z=0.38 → S≈2.9ppm → Collar≈5.0 (HI2/HI3 경계)
#   z=0.59 → S≈4.6ppm → Collar≈6.5 (HI3/HI4 경계)
#   z=0.91 → S≈7.0ppm → Collar≈8.5 (HI4/HI5 경계)
#
# DEFAULT: 상태-Collar 기반 (sev_rank과 noise_std 관계로 경험적 설정)
Z_RANGES = {
    "HV_TR": {
        "HI1": (0.00, 0.15),
        "HI2": (0.15, 0.38),
        "HI3": (0.38, 0.59),
        "HI4": (0.59, 0.91),
        "HI5": (0.91, 1.00),
    },
    "DEFAULT": {
        "HI1": (0.00, 0.25),
        "HI2": (0.25, 0.50),
        "HI3": (0.50, 0.68),
        "HI4": (0.68, 0.84),
        "HI5": (0.84, 1.00),
    },
}

# ==================================================================
# 3. 자산군별 랜덤 속성 생성
# ==================================================================

def gen_attributes(asset_type: str, adef: tuple, n: int, rng: np.random.Generator) -> pd.DataFrame:
    (cnaim_cat, code, K, C, HSL, NEL,
     ratio, ref_cust, rep_cost, maint_h, labor_d,
     cof_F, cof_S, cof_E, cof_NP) = adef

    # -- 나이: 1 ~ NEL+15, 자산군 특성 반영
    # 목주/콘크리트주 → 설계수명 넘은 설비 많음 (p=0.35 for age>NEL)
    if asset_type in ["목주", "콘크리트주", "철주"]:
        ages = rng.integers(5, NEL + 20, size=n)
    elif asset_type in ["주상변압기", "가공배전선로"]:
        ages = rng.integers(1, NEL + 10, size=n)
    else:
        ages = rng.integers(1, NEL + 5, size=n)

    # -- 위치 속성 (Korean distribution network 반영)
    # 해안거리: 내륙 많음, 연안 일부
    coast_km = rng.choice(
        [0.5, 3.0, 8.0, 15.0, 30.0],
        size=n, p=[0.04, 0.12, 0.18, 0.36, 0.30]
    )
    # 고도: 평야·구릉 중심
    altitude_m = rng.choice(
        [50, 150, 250, 400],
        size=n, p=[0.45, 0.33, 0.14, 0.08]
    )
    # 부식등급 1(낮음)~5(높음): 한국 중부 이북 내륙 3 중심
    corr_cat = rng.choice(
        [1, 2, 3, 4, 5],
        size=n, p=[0.10, 0.22, 0.44, 0.17, 0.07]
    )

    # -- 잠재 중증도(z) 생성: 목표 HI 등급 분포 → 역방향 z 배정 + 약한 나이 상관
    #
    # [설계 원칙]
    # 1. 각 자산에 목표 HI 등급을 먼저 배정(HI_PROBS) → 자산유형과 무관하게 동일 분포.
    #    덕분에 ILP 최적화 시 변압기/개폐기/지지물 간 진짜 trade-off가 발생한다.
    # 2. 배정된 HI 등급에 맞는 z 범위(Z_RANGES)에서 균등 샘플링.
    # 3. 나이가 많을수록 z를 최대 ±0.10 미세 조정(약한 상관).
    #    → "나이 많은 자산이 나쁜 상태일 확률이 약간 높다"는 현실 반영.
    # 4. 동일 자산의 모든 컨디션 입력은 이 z를 공유 → 상관성 보장.
    asset_key    = cct.ASSET_KEY_MAP[asset_type]
    z_range_map  = Z_RANGES.get(asset_key, Z_RANGES["DEFAULT"])

    hi_indices = rng.choice(len(HI_LABELS), size=n, p=HI_PROBS)
    z = np.zeros(n)
    for k, label in enumerate(HI_LABELS):
        mask = (hi_indices == k)
        if mask.any():
            lo, hi_z = z_range_map[label]
            z[mask] = rng.uniform(lo, hi_z, size=mask.sum())

    # 약한 나이 상관: 나이/NEL이 0.5보다 크면 z 소폭 증가, 작으면 감소 (최대 ±0.10)
    age_ratio = np.clip(ages / max(NEL, 1), 0.0, 1.5)
    z = np.clip(z + 0.10 * (age_ratio - 0.5), 0.0, 1.0)

    cond_cols = {}
    for category in ("observed", "measured"):
        defs = cct.CATEGORICAL_INPUTS.get(asset_key, {}).get(category, {})
        for input_name, states in defs.items():
            col_name = cct.COLUMN_NAMES[(asset_key, category, input_name)]
            cond_cols[col_name] = cct.sample_states(states, z, rng)

    # -- HV Transformer(주상/지상/지중변압기) 전용: 오일·유중가스·FFA 원시값
    if asset_key == "HV_TR":
        for input_name, col_name in cct.HV_TR_CONTINUOUS_COLUMNS.items():
            good, bad = cct.HV_TR_CONTINUOUS_RANGES[input_name]
            cond_cols[col_name] = cct.sample_continuous(z, rng, good, bad)

    # -- 운전 조건 (자산군별)
    if asset_type in ["주상변압기", "지상변압기", "지중변압기"]:
        util_pct = rng.choice(
            [35, 55, 75, 100, 125],
            size=n, p=[0.12, 0.35, 0.32, 0.16, 0.05]
        )
        duty_val = util_pct.astype(object)
        duty_col = "부하율_pct"
    elif asset_type in ["가공개폐기", "지중개폐기_RMU", "특고압차단기"]:
        ops_cat = rng.choice(
            ["저(<30회/년)", "보통(30~100)", "고(>100/자동재폐로)"],
            size=n, p=[0.45, 0.37, 0.18]
        )
        duty_val = ops_cat.astype(object)
        duty_col = "운전횟수구분"
    elif asset_type in ["가공배전선로", "지중케이블"]:
        # 선로 구간길이 (m): 가공=200~800m, 지중=100~600m
        if asset_type == "가공배전선로":
            length = rng.integers(200, 801, size=n)
        else:
            length = rng.integers(100, 601, size=n)
        duty_val = length.astype(object)
        duty_col = "선로길이_m"
    else:  # 목주, 콘크리트주, 철주, LV배전반
        duty_val = np.full(n, "해당없음", dtype=object)
        duty_col = "비고"

    # -- CoF 관련 속성
    num_cust = np.clip(
        rng.lognormal(np.log(max(ref_cust, 1)), 0.55, size=n).astype(int),
        1, 800
    )
    near_water = rng.choice(["수변", "비수변"], size=n, p=[0.12, 0.88])
    size_cat   = rng.choice(["소", "중", "대"], size=n, p=[0.30, 0.50, 0.20])

    # -- ID 생성
    ids = [f"{code}-{i+1:05d}" for i in range(n)]

    df = pd.DataFrame({
        "자산ID":           ids,
        "자산유형":         asset_type,
        "CNAIM카테고리":    cnaim_cat,
        "교체비용_천원":    np.full(n, cof_F),   # CoF_재무 기준값 = 교체비용 대리변수
        "나이_년":          ages,
        duty_col:           duty_val,
        "해안거리_km":      coast_km,
        "고도_m":           altitude_m,
        "부식등급_1~5":     corr_cat,
        **cond_cols,
        "고객수":           num_cust,
        "수변여부":         near_water,
        "설비규모":         size_cat,
    })

    return df


# ==================================================================
# 4. 전체 자산 목록 생성
# ==================================================================

def generate_all_assets() -> pd.DataFrame:
    rng = np.random.default_rng(RANDOM_SEED)
    parts = []

    print(f"\n[자산 생성] 총 {TOTAL_ASSETS:,}기 (12개 자산군)")
    print(f"{'자산유형':<14} {'수량':>6} {'비율':>7}  근거")
    print("-" * 65)

    ref_map = {
        "주상변압기":    "KEPCO 최다 자산 (22.9→220V 주력)",
        "지상변압기":    "도심 지상형 (UK DNO ~6%)",
        "지중변압기":    "도심 지하 기계실 (UK DNO ~3%)",
        "가공개폐기":    "자동재폐로+구분개폐기 (KEPCO ~13%)",
        "지중개폐기_RMU": "RMU, 도심 지중 (KEPCO ~5%)",
        "특고압차단기":  "1차 변전소 VCB (KEPCO ~3%)",
        "가공배전선로":  "500m 구간 단위 (KEPCO ~13%)",
        "지중케이블":    "XLPE, 300m 구간 (KEPCO ~7%)",
        "목주":          "농촌 나무지지물 (UK ~5%)",
        "콘크리트주":    "일반 콘크리트주 (UK ~6%)",
        "철주":          "간선형 철주 (UK ~1%)",
        "LV배전반_UGB":  "저압 분기함 (UK ~1%)",
    }

    for atype, adef in ASSET_DEFS.items():
        ratio = adef[6]
        n = max(1, round(TOTAL_ASSETS * ratio))
        print(f"  {atype:<14} {n:>6,}기  {ratio*100:>5.1f}%  {ref_map.get(atype,'')}")
        df = gen_attributes(atype, adef, n, rng)
        parts.append(df)

    all_df = pd.concat(parts, ignore_index=True)
    # 순위 재배열: 자산ID를 전역 일련번호로 재발급
    all_df.insert(0, "NO", range(1, len(all_df) + 1))
    print(f"{'합계':<14} {len(all_df):>6,}기  100.0%")
    return all_df


# ==================================================================
# 5. CNAIM 파라미터 Excel 생성
# ==================================================================

def save_cnaim_params():
    wb = openpyxl.Workbook()
    H1 = PatternFill("solid", fgColor="1F4E79")
    H2 = PatternFill("solid", fgColor="2E75B6")
    GR = PatternFill("solid", fgColor="E2EFDA")

    def hdr(cell, txt, fill=H1):
        cell.value = txt
        cell.font  = Font(bold=True, color="FFFFFF", size=9)
        cell.fill  = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # -- Sheet 1: 분석 설정
    ws0 = wb.active
    ws0.title = "0_분석설정"
    ws0.column_dimensions["A"].width = 26
    ws0.column_dimensions["B"].width = 16
    ws0.column_dimensions["C"].width = 40
    for c, t in enumerate(["항목", "값", "비고"], 1):
        hdr(ws0.cell(1, c), t)

    cfg = [
        ("할인율_pct",              5.0,   "연간 할인율 (WACC 기준, 한전 5% 추정)"),
        ("총자산수",              5_000,   "전체 생성 설비 수"),
        ("랜덤시드",                42,    "재현성용 난수 시드"),
        ("CNAIM버전",           "v2.1",    "Ofgem, April 2021"),
        ("환율_KRW_per_GBP",    1_600,    "CNAIM(£) → 한국(천원) 환산 기준"),
    ]
    for r, (k, v, n) in enumerate(cfg, 2):
        ws0.cell(r, 1, k).alignment = Alignment(horizontal="left")
        ws0.cell(r, 2, v).alignment = Alignment(horizontal="center")
        ws0.cell(r, 3, n).alignment = Alignment(horizontal="left")
        ws0.cell(r, 2).fill = PatternFill("solid", fgColor="FFF2CC")

    # -- Sheet 2: 자산군 PoF 파라미터
    ws1 = wb.create_sheet("1_PoF파라미터")
    cols1 = ["자산유형", "CNAIM카테고리", "코드", "K값(%)", "C값", "HSL", "NEL(년)", "비율(%)", "근거"]
    for c, t in enumerate(cols1, 1):
        ws1.column_dimensions[get_column_letter(c)].width = 20
        hdr(ws1.cell(1, c), t)
    ws1.row_dimensions[1].height = 25

    for r, (atype, adef) in enumerate(ASSET_DEFS.items(), 2):
        (cnaim_cat, code, K, C, HSL, NEL, ratio, *_) = adef
        row_vals = [atype, cnaim_cat, code, K, C, HSL, NEL, round(ratio*100, 1),
                    "CNAIM Table 20-21"]
        for c, v in enumerate(row_vals, 1):
            cell = ws1.cell(r, c, v)
            cell.alignment = Alignment(horizontal="center")
            if r % 2 == 0:
                cell.fill = GR

    # -- Sheet 3: CoF 기준값 + 투자 파라미터
    ws2 = wb.create_sheet("2_CoF및투자파라미터")
    cols2 = ["자산유형", "참조고객수", "교체비용(만원)", "공사시간(h)", "투입인력(인일)",
             "CoF_재무(천원)", "CoF_안전(천원)", "CoF_환경(천원)", "CoF_계통(천원)",
             "CoF합계기준(천원)", "출처"]
    for c, t in enumerate(cols2, 1):
        ws2.column_dimensions[get_column_letter(c)].width = 17
        hdr(ws2.cell(1, c), t)

    for r, (atype, adef) in enumerate(ASSET_DEFS.items(), 2):
        (_, _, _, _, _, _, _, ref_cust, rep_cost, maint_h, labor_d,
         cof_F, cof_S, cof_E, cof_NP) = adef
        row_vals = [atype, ref_cust, rep_cost, maint_h, labor_d,
                    cof_F, cof_S, cof_E, cof_NP, cof_F+cof_S+cof_E+cof_NP,
                    "CNAIM Table 16 기반 조정"]
        for c, v in enumerate(row_vals, 1):
            cell = ws2.cell(r, c, v)
            cell.alignment = Alignment(horizontal="center")
            if r % 2 == 0:
                cell.fill = GR

    # -- Sheet 4: 위치계수 룩업 테이블 (CNAIM Table 22-24)
    ws3 = wb.create_sheet("3_위치계수테이블")
    ws3["A1"] = "CNAIM Table 22~24 위치계수 (자동 적용 -- Python 내장)"
    ws3["A1"].font = Font(bold=True, color="CC0000")

    ws3["A3"] = "해안거리 계수 (CNAIM Table 22)"
    for c, t in enumerate(["거리구분", "개폐기/변압기", "목주(Wood)", "철주/콘크리트주"], 1):
        hdr(ws3.cell(4, c), t, fill=H2)
    coast_rows = [
        ("≤1 km",       1.35, 2.00, 1.50),
        (">1~5 km",     1.10, 1.50, 1.20),
        (">5~10 km",    1.05, 1.20, 1.10),
        (">10~20 km",   1.00, 1.00, 1.00),
        (">20 km",      0.90, 1.00, 1.00),
    ]
    for r, row in enumerate(coast_rows, 5):
        for c, v in enumerate(row, 1):
            ws3.cell(r, c, v).alignment = Alignment(horizontal="center")

    ws3["A11"] = "고도 계수 (CNAIM Table 23)"
    for c, t in enumerate(["고도구분", "개폐기/변압기", "목주(Wood)", "철주/콘크리트주"], 1):
        hdr(ws3.cell(12, c), t, fill=H2)
    alt_rows = [
        ("≤100 m",      0.90, 0.95, 1.00),
        (">100~200 m",  1.00, 1.00, 1.00),
        (">200~300 m",  1.05, 1.05, 1.00),
        (">300 m",      1.10, 1.15, 1.00),
    ]
    for r, row in enumerate(alt_rows, 13):
        for c, v in enumerate(row, 1):
            ws3.cell(r, c, v).alignment = Alignment(horizontal="center")

    ws3["A19"] = "부식 계수 (CNAIM Table 24)"
    for c, t in enumerate(["부식등급", "개폐기/변압기", "목주(Wood)", "철주/콘크리트주"], 1):
        hdr(ws3.cell(20, c), t, fill=H2)
    corr_rows = [
        ("1 (낮음)",    0.90, 1.00, 0.90),
        ("2",           0.95, 1.00, 0.95),
        ("3 (중간)",    1.00, 1.00, 1.00),
        ("4",           1.10, 1.00, 1.15),
        ("5 (높음)",    1.25, 1.00, 1.35),
    ]
    for r, row in enumerate(corr_rows, 21):
        for c, v in enumerate(row, 1):
            ws3.cell(r, c, v).alignment = Alignment(horizontal="center")

    # -- Sheet 5: 의무교체 기준
    ws4 = wb.create_sheet("4_의무교체기준")
    ws4["A1"] = "의무교체 판단 기준 (나이 AND 현재건강지수 HI밴드 동시 충족)"
    ws4["A1"].font = Font(bold=True)
    for c, t in enumerate(["자산유형", "나이기준(년)", "상태기준(HI밴드)", "비고"], 1):
        ws4.column_dimensions[get_column_letter(c)].width = 18
        hdr(ws4.cell(2, c), t)
    mand_rows = [
        ("주상변압기",    30, "HI5(CHS>=8.5)", "나이 AND 상태 동시 충족"),
        ("지상변압기",    30, "HI5(CHS>=8.5)", "나이 AND 상태 동시 충족"),
        ("지중변압기",    30, "HI5(CHS>=8.5)", "나이 AND 상태 동시 충족"),
        ("가공개폐기",    25, "HI5(CHS>=8.5)", "나이 AND 상태 동시 충족"),
        ("지중개폐기_RMU",25, "HI5(CHS>=8.5)", "나이 AND 상태 동시 충족"),
        ("특고압차단기",  25, "HI5(CHS>=8.5)", "나이 AND 상태 동시 충족"),
        ("가공배전선로",  25, "HI5(CHS>=8.5)", "선로 구간 교체"),
        ("지중케이블",    20, "HI5(CHS>=8.5)", "케이블 구간 교체"),
        ("목주",          30, "HI5(CHS>=8.5)", "부후(腐朽) 위험"),
        ("콘크리트주",    35, "HI5(CHS>=8.5)", "균열/부식 심각"),
        ("철주",          25, "HI5(CHS>=8.5)", "부식 심각"),
        ("LV배전반_UGB",  20, "HI5(CHS>=8.5)", "절연 열화"),
    ]
    for r, row in enumerate(mand_rows, 3):
        for c, v in enumerate(row, 1):
            ws4.cell(r, c, v).alignment = Alignment(horizontal="center")
            if r % 2 == 0:
                ws4.cell(r, c).fill = GR

    wb.save(OUT_PARAMS)
    print(f"\n[파라미터 저장] → {OUT_PARAMS}")


# ==================================================================
# 6. 입력자산 Excel 저장 (pandas + openpyxl 헤더 서식)
# ==================================================================

def save_input_assets(df: pd.DataFrame):
    print(f"\n[Excel 저장 중] {len(df):,}행 × {len(df.columns)}열...", end="", flush=True)

    # 카테고리 시트 정의: (시트명, 자산유형 목록, duty 컬럼명 or None)
    CAT_DEFS = [
        ("1_변압기류",   ["주상변압기", "지상변압기", "지중변압기"],          "부하율_pct"),
        ("2_개폐기류",   ["가공개폐기", "지중개폐기_RMU", "특고압차단기"],    "운전횟수구분"),
        ("3_배전선로류", ["가공배전선로", "지중케이블"],                       "선로길이_m"),
        ("4_지지물류",   ["목주", "콘크리트주", "철주"],                       None),
        ("5_저압설비류", ["LV배전반_UGB"],                                     None),
    ]

    with pd.ExcelWriter(OUT_ASSETS, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="전체_자산목록", index=False)

        ratio_df = (df.groupby("자산유형")
                    .size()
                    .reset_index(name="수량")
                    .assign(비율_pct=lambda x: (x["수량"] / len(df) * 100).round(2)))
        ratio_df.to_excel(writer, sheet_name="자산군_수량비율", index=False)

        for sname, atypes, duty_col in CAT_DEFS:
            sub = df[df["자산유형"].isin(atypes)].reset_index(drop=True)

            # 자산유형별 관측/측정 컨디션 컬럼 합집합 (순서 보존, 중복 제거)
            cond_cols_union = []
            for atype in atypes:
                for c in cct.get_condition_columns(atype):
                    if c not in cond_cols_union:
                        cond_cols_union.append(c)

            cols = ["NO", "자산ID", "자산유형", "CNAIM카테고리", "교체비용_천원", "나이_년"]
            if duty_col and duty_col in sub.columns:
                cols.append(duty_col)
            cols += ["해안거리_km", "고도_m", "부식등급_1~5"] + cond_cols_union \
                    + ["고객수", "수변여부", "설비규모"]
            cols = [c for c in cols if c in sub.columns]
            sub[cols].to_excel(writer, sheet_name=sname, index=False)

    wb = openpyxl.load_workbook(OUT_ASSETS)
    H1 = PatternFill("solid", fgColor="1F4E79")
    GR = PatternFill("solid", fgColor="DEEAF1")

    ws = wb["전체_자산목록"]
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28

    for c_idx, col_name in enumerate(df.columns, 1):
        max_len = max(len(str(col_name)),
                      df.iloc[:, c_idx-1].astype(str).str.len().max() if c_idx <= len(df.columns) else 0)
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 3, 22)
        cell = ws.cell(1, c_idx)
        cell.font  = Font(bold=True, color="FFFFFF", size=9)
        cell.fill  = H1
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r_idx in range(2, len(df) + 2, 2):
        for c_idx in range(1, len(df.columns) + 1):
            ws.cell(r_idx, c_idx).fill = GR

    wb.save(OUT_ASSETS)
    print(f" 완료")
    print(f"[자산 목록 저장] -> {OUT_ASSETS}")


# ==================================================================
# 7. 메인
# ==================================================================

def main():
    print("=" * 66)
    print("  AIP Step 1 -- 배전설비 자산 목록 생성")
    print("  (KEPCO + UK DNO 비율 기반, 12종 자산군, 총 5,000기)")
    print("=" * 66)

    if OUT_ASSETS.exists() and OUT_PARAMS.exists():
        ans = input(f"\n이미 파일이 존재합니다. 덮어쓰시겠습니까? (y/N): ").strip().lower()
        if ans != "y":
            print("취소됨. 기존 파일을 유지합니다.")
            return

    # CNAIM 파라미터 Excel 먼저 저장
    save_cnaim_params()

    # 자산 생성 + Excel 저장
    df = generate_all_assets()
    save_input_assets(df)

    print("\n[완료]")
    print(f"  data/cnaim_params.xlsx  : CNAIM 파라미터 (할인율, K값, CoF 기준 등)")
    print(f"  data/input_assets.xlsx  : 자산별 입력 데이터 (편집 가능)")
    print(f"\n  다음 단계: py -3 aip_02_run_analysis.py")
    print("=" * 66)


if __name__ == "__main__":
    main()

