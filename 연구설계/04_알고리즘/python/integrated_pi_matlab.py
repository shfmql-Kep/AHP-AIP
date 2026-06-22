"""설비유형 가중치를 반영한 Integrated PI를 산출한다.

현재 Local PI는 개별 설비의 경제성, 신뢰도, 안전·환경 지표를 정규화한 값이다.
Integrated PI는 여기에 설비유형별 전문가 가중치와 비용 규모 보정계수를 결합하여
이종설비 통합 최적화의 목적함수로 사용한다.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent.parent
SURVEY_DIR = BASE_DIR / "Survey"
OUTPUT_DIR = BASE_DIR / "outputs"
LOCAL_PI_FILE = OUTPUT_DIR / "local_pi_matlab.xlsx"
OUTPUT_FILE = OUTPUT_DIR / "integrated_pi_matlab.xlsx"
YEARS = [2026, 2027, 2028, 2029, 2030]
DEFAULT_ALPHA = "α=0.5"

ASSET_TYPE_LABELS = {
    "pole_transformer": "주상변압기",
    "ground_transformer": "지상변압기",
    "overhead_switch": "가공개폐기",
    "underground_switch": "지중개폐기",
    "overhead_line": "가공배전선로",
    "underground_cable": "지중케이블",
}


def find_survey_file() -> Path:
    """최신 통합설비 설문 응답 파일을 찾는다."""
    candidates = sorted(
        SURVEY_DIR.glob("*응답완료*20*.xlsx"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError("통합설비 설문 응답 파일을 찾을 수 없습니다.")
    return candidates[0]


def load_type_weights() -> pd.DataFrame:
    """설문 파일의 통합정규화 시트에서 설비유형 가중치를 읽는다."""
    survey_file = find_survey_file()
    xl = pd.ExcelFile(survey_file)
    sheet_name = next((name for name in xl.sheet_names if "통합" in name and "정규화" in name), None)
    if sheet_name is None:
        raise ValueError("통합정규화 시트를 찾을 수 없습니다.")

    raw = pd.read_excel(survey_file, sheet_name=sheet_name, header=None)
    header_idx = raw.index[raw.iloc[:, 0].astype(str).str.contains("설비유형", na=False)][0]
    table = raw.iloc[header_idx + 1 :].copy()
    table = table.iloc[:, 0:8].copy()
    table.columns = ["설비유형", "평균비용(만원)", "W_type(전문가)", "α=0.0", "α=0.3", "α=0.5", "α=0.7", "α=1.0"]
    table = table[table["설비유형"].notna()].copy()
    table = table[table["설비유형"].astype(str).isin(ASSET_TYPE_LABELS.values())].copy()

    required = ["설비유형", "평균비용(만원)", "W_type(전문가)", "α=0.0", "α=0.3", "α=0.5", "α=0.7", "α=1.0"]
    missing = [col for col in required if col not in table.columns]
    if missing:
        raise ValueError(f"통합정규화 시트에 필요한 열이 없습니다: {missing}")

    for col in required[1:]:
        table[col] = pd.to_numeric(table[col], errors="coerce")

    code_map = {label: code for code, label in ASSET_TYPE_LABELS.items()}
    table["asset_type"] = table["설비유형"].map(code_map)
    table = table.rename(
        columns={
            "설비유형": "asset_type_label",
            "평균비용(만원)": "avg_cost_10k_krw",
            "W_type(전문가)": "w_type_expert",
            "α=0.0": "w_type_alpha_0_0",
            "α=0.3": "w_type_alpha_0_3",
            "α=0.5": "w_type_alpha_0_5",
            "α=0.7": "w_type_alpha_0_7",
            "α=1.0": "w_type_alpha_1_0",
        }
    )
    return table[
        [
            "asset_type",
            "asset_type_label",
            "avg_cost_10k_krw",
            "w_type_expert",
            "w_type_alpha_0_0",
            "w_type_alpha_0_3",
            "w_type_alpha_0_5",
            "w_type_alpha_0_7",
            "w_type_alpha_1_0",
        ]
    ].reset_index(drop=True)


def build_integrated_pi(type_weights: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """Local PI와 설비유형 가중치를 결합한다."""
    local = pd.read_excel(LOCAL_PI_FILE, sheet_name="local_pi_asset_wide")
    merged = local.merge(type_weights, on="asset_type", how="left", validate="many_to_one")
    if merged["w_type_alpha_0_5"].isna().any():
        missing = sorted(merged.loc[merged["w_type_alpha_0_5"].isna(), "asset_type"].unique())
        raise ValueError(f"설비유형 가중치가 없는 asset_type이 있습니다: {missing}")

    alpha_cols = {
        "alpha_0_0": "w_type_alpha_0_0",
        "alpha_0_3": "w_type_alpha_0_3",
        "alpha_0_5": "w_type_alpha_0_5",
        "alpha_0_7": "w_type_alpha_0_7",
        "alpha_1_0": "w_type_alpha_1_0",
    }
    for year in YEARS:
        for alpha_name, weight_col in alpha_cols.items():
            merged[f"integrated_pi_ahp_{alpha_name}_{year}"] = (
                merged[f"local_pi_ahp_{year}"] * merged[weight_col]
            )
            merged[f"integrated_pi_fuzzy_adjusted_{alpha_name}_{year}"] = (
                merged[f"local_pi_fuzzy_adjusted_{year}"] * merged[weight_col]
            )

    keep_cols = [
        "asset_id",
        "asset_type",
        "asset_code",
        "asset_label",
        "asset_group",
        "asset_type_label",
        "candidate_top30_current",
        "w_type_expert",
        "w_type_alpha_0_0",
        "w_type_alpha_0_3",
        "w_type_alpha_0_5",
        "w_type_alpha_0_7",
        "w_type_alpha_1_0",
    ]
    for year in YEARS:
        keep_cols.extend(
            [
                f"local_pi_ahp_{year}",
                f"local_pi_fuzzy_adjusted_{year}",
                f"integrated_pi_ahp_alpha_0_5_{year}",
                f"integrated_pi_fuzzy_adjusted_alpha_0_5_{year}",
            ]
        )
    wide_default = merged[keep_cols].copy()

    summary_type_year = []
    for year in YEARS:
        for asset_type, sub in merged.groupby("asset_type"):
            summary_type_year.append(
                {
                    "year": year,
                    "asset_type": asset_type,
                    "asset_type_label": sub["asset_type_label"].iloc[0],
                    "asset_count": len(sub),
                    "w_type_alpha_0_5": sub["w_type_alpha_0_5"].iloc[0],
                    "sum_local_pi_ahp": sub[f"local_pi_ahp_{year}"].sum(),
                    "sum_integrated_pi_ahp_alpha_0_5": sub[f"integrated_pi_ahp_alpha_0_5_{year}"].sum(),
                    "mean_local_pi_ahp": sub[f"local_pi_ahp_{year}"].mean(),
                    "mean_integrated_pi_ahp_alpha_0_5": sub[f"integrated_pi_ahp_alpha_0_5_{year}"].mean(),
                }
            )

    return {
        "type_weights": type_weights,
        "integrated_pi_asset_wide": wide_default,
        "integrated_pi_summary_type_year": pd.DataFrame(summary_type_year),
    }


def write_workbook(sheets: dict[str, pd.DataFrame]) -> None:
    """Integrated PI 결과 파일을 저장한다."""
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

    wb = load_workbook(OUTPUT_FILE)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_idx in range(1, ws.max_column + 1):
            max_len = 10
            for row_idx in range(1, min(ws.max_row, 80) + 1):
                value = ws.cell(row_idx, col_idx).value
                if value is not None:
                    max_len = max(max_len, len(str(value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 36)
    wb.save(OUTPUT_FILE)


def main() -> None:
    """Integrated PI 산출을 실행한다."""
    type_weights = load_type_weights()
    sheets = build_integrated_pi(type_weights)
    write_workbook(sheets)
    print(f"saved: {OUTPUT_FILE}")
    print(type_weights.to_string(index=False))


if __name__ == "__main__":
    main()
