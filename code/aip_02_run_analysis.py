"""
AIP Step 2 -- CNAIM 기반 위험도 분석 및 투자가치(NPV) 산출
===========================================================
입력:
  data/input_assets.xlsx   (Step 1에서 생성한 자산 목록)
  data/cnaim_params.xlsx   (Step 1에서 생성한 파라미터)

출력:
  data/output_results_YYYYMMDD_HHMMSS.xlsx

[계산 파이프라인 -- CNAIM v2.1 EQ.3~12]
  1. 위치계수  = max(해안거리계수, 고도계수, 부식계수)  [Table 22-24]
  2. 운전계수  = 부하율/운전횟수 → Duty Factor         [Table 31-33]
  3. 기대수명  = NEL / (Duty × Location)               [EQ.4]
  4. β1        = ln(5.5/0.5) / 기대수명                [EQ.5]
  5. 초기HS    = 0.5 × exp(β1 × 나이)  [≤5.5]         [EQ.6]
  6. 건강지수계수 = MMI(관측계수, 측정계수)            [Section 6.7.2]
  7. 현재HS    = 초기HS × 건강지수계수  [cap/collar]   [EQ.7-9]
  8. PoF       = K × [1 + CH + (CH)²/2 + (CH)³/6]    [EQ.3]
  9. CoF       = F + S + E×(위치×규모계수) + NP×고객계수  [Section 7]
  10. 위험도   = PoF × CoF
  11. β2       = ln(CHS/0.5) / 나이  [≤2β1]           [EQ.10]
  12. ARF      = 노화감소계수(CHS 기반)                [Figure 5]
  13. 교체전 NPV = Σ[PoF(t)×CoF/(1+r)^t], t=0→NEL
  14. 교체후 NPV = Σ[PoF_new(t)×CoF/(1+r)^t], t=0→NEL
  15. 투자가치 = 교체전 NPV − 교체후 NPV

실행: py -3 aip_02_run_analysis.py
"""

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime
import sys
import warnings
warnings.filterwarnings("ignore")

import cnaim_condition_tables as cct

# -- 경로
BASE_DIR   = Path(__file__).resolve().parent.parent
DATA_DIR   = BASE_DIR / "data"
IN_ASSETS  = DATA_DIR / "input_assets.xlsx"
IN_PARAMS  = DATA_DIR / "cnaim_params.xlsx"

# ==================================================================
# 1. CNAIM 상수 및 룩업 테이블 (Table 22-24, 31-33, 35+)
# ==================================================================

H_NEW = 0.5   # 신규 설비 건강지수 (CNAIM 기본값)
H_EOL = 5.5   # 설계수명 도달 건강지수

# 자산 코드 → 위치계수 컬럼 종류
# 'sw_tr'=개폐기/변압기/UGB  'wood'=목주  'sc'=철주·콘크리트주
ASSET_LOC_TYPE = {
    "TR_P": "sw_tr", "TR_G": "sw_tr", "TR_V": "sw_tr",
    "SW_OH":"sw_tr", "SW_UG":"sw_tr", "CB_HV":"sw_tr",
    "OHL":  "wood",  "UGC":  "sw_tr",
    "PW":   "wood",  "PC":   "sc",    "PS":   "sc",
    "LV_UGB":"sw_tr",
}

# Table 22: Coast Distance Factor (km, factor)
COAST_TABLE = {
    "sw_tr": [(1.0,1.35),(5.0,1.10),(10.0,1.05),(20.0,1.00),(9999,0.90)],
    "wood":  [(1.0,2.00),(5.0,1.50),(10.0,1.20),(20.0,1.00),(9999,1.00)],
    "sc":    [(1.0,1.50),(5.0,1.20),(10.0,1.10),(20.0,1.00),(9999,1.00)],
}
# Table 23: Altitude Factor (m, factor)
ALT_TABLE = {
    "sw_tr": [(100,0.90),(200,1.00),(300,1.05),(9999,1.10)],
    "wood":  [(100,0.95),(200,1.00),(300,1.05),(9999,1.15)],
    "sc":    [(100,1.00),(200,1.00),(300,1.00),(9999,1.00)],
}
# Table 24: Corrosion Factor {cat: factor}
CORR_TABLE = {
    "sw_tr": {1:0.90,2:0.95,3:1.00,4:1.10,5:1.25},
    "wood":  {1:1.00,2:1.00,3:1.00,4:1.00,5:1.00},
    "sc":    {1:0.90,2:0.95,3:1.00,4:1.15,5:1.35},
}

def _lookup(table: list, value: float) -> float:
    for limit, factor in table:
        if value <= limit:
            return factor
    return table[-1][1]

def location_factor(coast_km: float, altitude_m: float, corr_cat: int,
                    loc_type: str) -> float:
    """CNAIM EQ.13~17: 위치계수 산출"""
    cf = _lookup(COAST_TABLE.get(loc_type, COAST_TABLE["sw_tr"]), coast_km)
    af = _lookup(ALT_TABLE.get(loc_type, ALT_TABLE["sw_tr"]), altitude_m)
    rf = CORR_TABLE.get(loc_type, CORR_TABLE["sw_tr"]).get(int(corr_cat), 1.0)
    max_f = max(cf, af, rf)
    INC = 0.05 if loc_type == "sw_tr" else 0.0
    return round(max_f + (max_f - 1.0) * INC, 4) if max_f > 1.0 else max_f

# Table 33: Transformer Duty Factor (부하율 %)
def duty_factor_transformer(util_pct: float) -> float:
    if util_pct <= 50:   return 0.90
    if util_pct <= 70:   return 0.95
    if util_pct <= 100:  return 1.00
    return 1.40  # 과부하

# Table 32: Switchgear Duty Factor (운전 횟수/종류)
def duty_factor_switch(ops_cat: str) -> float:
    if "저" in str(ops_cat) or "<30" in str(ops_cat):  return 0.90
    if "고" in str(ops_cat) or "자동재폐로" in str(ops_cat): return 1.20
    return 1.00  # 보통

# 관측/측정 컨디션 입력 전체 항목, Factor/Cap/Collar 캘리브레이션,
# MMI 결합 알고리즘은 cnaim_condition_tables.py(CNAIM Appendix B 전체 발췌)에
# 정의되어 있다 (cct.CATEGORICAL_INPUTS, cct.combine_category, cct.mmi_general).

def ageing_reduction_factor(chs: np.ndarray) -> np.ndarray:
    """CNAIM Figure 5: 노화감소계수 (ARF)"""
    return np.where(chs < 2.0, 1.0,
           np.where(chs >= 5.5, 1.5,
                    1.0 + 0.5 * (chs - 2.0) / 3.5))


# ==================================================================
# 2. 파라미터 로드
# ==================================================================

def load_params() -> tuple:
    """cnaim_params.xlsx 에서 파라미터 읽기"""
    xl = pd.ExcelFile(IN_PARAMS, engine="openpyxl")

    # 분석 설정
    df_cfg = pd.read_excel(xl, "0_분석설정", header=0, usecols=[0,1],
                           names=["항목","값"])
    cfg = dict(zip(df_cfg["항목"], df_cfg["값"]))
    discount_rate = float(str(cfg.get("할인율_pct", 5.0)).replace("%","")) / 100.0

    # PoF 파라미터
    df_pof = pd.read_excel(xl, "1_PoF파라미터", header=0)
    df_pof.columns = [c.strip() for c in df_pof.columns]
    # 컬럼: 자산유형, CNAIM카테고리, 코드, K값(%), C값, HSL, NEL(년), 비율(%), 근거
    df_pof = df_pof.rename(columns={
        "K값(%)":"K_pct", "C값":"C", "HSL":"HSL", "NEL(년)":"NEL"
    }).set_index("자산유형")

    # CoF + 투자 파라미터
    df_cof = pd.read_excel(xl, "2_CoF및투자파라미터", header=0)
    df_cof.columns = [c.strip() for c in df_cof.columns]
    df_cof = df_cof.rename(columns={
        "참조고객수":"ref_cust",
        "교체비용(만원)":"rep_cost",
        "공사시간(h)":"maint_h",
        "투입인력(인일)":"labor_d",
        "CoF_재무(천원)":"cof_F",
        "CoF_안전(천원)":"cof_S",
        "CoF_환경(천원)":"cof_E",
        "CoF_계통(천원)":"cof_NP",
    }).set_index("자산유형")

    # 의무교체 기준
    df_mand = pd.read_excel(xl, "4_의무교체기준", header=1)
    df_mand.columns = [c.strip() for c in df_mand.columns]
    df_mand = df_mand.rename(columns={
        "나이기준(년)":"mand_age",
    }).set_index("자산유형")

    return discount_rate, df_pof, df_cof, df_mand


# ==================================================================
# 3. 메인 계산 (벡터화)
# ==================================================================

def run_analysis(df_in: pd.DataFrame, discount_rate: float,
                 df_pof: pd.DataFrame, df_cof: pd.DataFrame,
                 df_mand: pd.DataFrame) -> pd.DataFrame:
    n = len(df_in)
    print(f"\n[계산 시작] {n:,}기 자산")

    # -- 결과 컬럼 초기화
    results = df_in.copy()
    K_arr   = np.zeros(n)
    C_arr   = np.zeros(n)
    HSL_arr = np.zeros(n, dtype=int)
    NEL_arr = np.zeros(n, dtype=int)
    loc_f   = np.ones(n)
    duty_f  = np.ones(n)
    exp_life = np.zeros(n)
    beta1    = np.zeros(n)
    init_hs  = np.zeros(n)
    obs_f    = np.ones(n)
    meas_f   = np.ones(n)
    obs_cap  = np.full(n, 10.0)
    obs_col  = np.zeros(n)
    meas_col = np.zeros(n)
    hs_factor= np.ones(n)
    curr_hs  = np.zeros(n)
    pof_cur  = np.zeros(n)
    cof_F    = np.zeros(n)
    cof_S    = np.zeros(n)
    cof_E    = np.zeros(n)
    cof_NP   = np.zeros(n)
    cof_tot  = np.zeros(n)
    ref_cust_arr = np.ones(n)
    rep_cost_arr = np.zeros(n)
    maint_h_arr  = np.zeros(n)
    labor_d_arr  = np.zeros(n)
    mandatory    = np.zeros(n, dtype=int)
    repcost_input = np.zeros(n)   # 입력 교체비용_천원 (= CoF_재무 기준값)

    # -- 코드 → 행 인덱스 매핑
    code_col = "CNAIM카테고리"  # Step1에서 저장한 컬럼명

    print("  (위치계수·운전계수·건강지수·PoF·CoF 계산 중...)")
    for i, row in df_in.iterrows():
        atype  = str(row.get("자산유형", ""))
        age    = int(row.get("나이_년", 1))
        coast  = float(row.get("해안거리_km", 15.0))
        alt    = float(row.get("고도_m", 150.0))
        corr   = int(row.get("부식등급_1~5", 3))
        cust   = float(row.get("고객수", 30))
        water  = str(row.get("수변여부", "비수변"))
        size   = str(row.get("설비규모", "중"))

        # PoF 파라미터
        if atype not in df_pof.index:
            continue
        prow = df_pof.loc[atype]
        K  = float(prow["K_pct"]) / 100.0
        C  = float(prow["C"])
        HSL= int(prow["HSL"])
        NEL= int(prow["NEL"])
        K_arr[i]=K; C_arr[i]=C; HSL_arr[i]=HSL; NEL_arr[i]=NEL

        # 코드 찾기
        code = str(prow.get("코드", "sw_tr"))
        lt   = ASSET_LOC_TYPE.get(code, "sw_tr")

        # 위치계수 (EQ.13-17)
        lf   = location_factor(coast, alt, corr, lt)
        loc_f[i] = lf

        # 운전계수 (자산군별)
        if atype in ["주상변압기","지상변압기","지중변압기"]:
            util = float(str(row.get("부하율_pct", 75)).replace("해당없음","75"))
            df_ = duty_factor_transformer(util)
        elif atype in ["가공개폐기","지중개폐기_RMU","특고압차단기"]:
            ops = str(row.get("운전횟수구분","보통(30~100)"))
            df_ = duty_factor_switch(ops)
        else:
            df_ = 1.0
        duty_f[i] = df_

        # 기대수명 / β1 (EQ.4, 5)
        el = max(NEL / (df_ * lf), 1.0)
        b1 = np.log(H_EOL / H_NEW) / el
        exp_life[i] = el
        beta1[i]    = b1

        # 초기 HS (EQ.6)
        ihs = min(H_NEW * np.exp(b1 * age), H_EOL)
        init_hs[i] = ihs

        # 관측·측정 컨디션 입력 -> Observed CF / Measured CF (1차 MMI 결합)
        asset_key = cct.ASSET_KEY_MAP.get(atype)
        obs_defs  = cct.CATEGORICAL_INPUTS.get(asset_key, {}).get("observed", {})
        meas_defs = cct.CATEGORICAL_INPUTS.get(asset_key, {}).get("measured", {})
        obs_state_dict  = {name: str(row.get(cct.COLUMN_NAMES[(asset_key,"observed",name)], ""))
                            for name in obs_defs}
        meas_state_dict = {name: str(row.get(cct.COLUMN_NAMES[(asset_key,"measured",name)], ""))
                            for name in meas_defs}
        of, oc, oll = cct.combine_category(asset_key, "observed", obs_state_dict)
        mf, mc, mll = cct.combine_category(asset_key, "measured", meas_state_dict)
        obs_f[i]=of; meas_f[i]=mf

        top_factors = [of, mf]
        top_caps    = [oc, mc]
        top_collars = [oll, mll]

        # HV Transformer(주상/지상/지중변압기) 전용: 오일/유중가스(DGA)/FFA 시험 보정
        if asset_key == "HV_TR":
            cols = cct.HV_TR_CONTINUOUS_COLUMNS
            ranges = cct.HV_TR_CONTINUOUS_RANGES
            moisture  = float(row.get(cols["moisture"],  ranges["moisture"][0]))
            acidity   = float(row.get(cols["acidity"],   ranges["acidity"][0]))
            breakdown = float(row.get(cols["breakdown"], ranges["breakdown"][0]))
            oil_f, oil_cap, oil_col = cct.oil_test_modifier(moisture, acidity, breakdown)

            h2   = float(row.get(cols["h2"],   ranges["h2"][0]))
            ch4  = float(row.get(cols["ch4"],  ranges["ch4"][0]))
            c2h4 = float(row.get(cols["c2h4"], ranges["c2h4"][0]))
            c2h6 = float(row.get(cols["c2h6"], ranges["c2h6"][0]))
            c2h2 = float(row.get(cols["c2h2"], ranges["c2h2"][0]))
            dga_f, dga_cap, dga_col = cct.dga_test_modifier(h2, ch4, c2h4, c2h6, c2h2)

            ffa_ppm = float(row.get(cols["ffa"], ranges["ffa"][0]))
            ffa_f, ffa_cap, ffa_col = cct.ffa_test_modifier(ffa_ppm)

            top_factors += [oil_f, dga_f, ffa_f]
            top_caps    += [oil_cap, dga_cap, ffa_cap]
            top_collars += [oil_col, dga_col, ffa_col]

        max_n, d1, d2 = cct.MMI_PARAMS.get(asset_key, {}).get("top", (2, 1.5, 1.5))
        hsf = cct.mmi_general(top_factors, d1, d2, max_n)
        obs_cap[i] = min(top_caps)
        obs_col[i] = max(top_collars)
        hs_factor[i] = hsf

        # 현재 HS (EQ.7-9: apply, cap, collar)
        chs = np.clip(ihs * hsf, obs_col[i], obs_cap[i])
        chs = min(chs, 10.0)
        curr_hs[i] = chs

        # 현재 PoF (EQ.3, h=max(CHS,HSL))
        h = max(chs, HSL)
        pof = K * (1 + C*h + (C*h)**2/2 + (C*h)**3/6)
        # 케이블: PoF/km × 길이(km)
        if atype == "지중케이블":
            length_m = float(str(row.get("선로길이_m", 300)).replace("해당없음","300"))
            pof *= (length_m / 1000.0)
        pof_cur[i] = pof

        # CoF
        if atype not in df_cof.index:
            continue
        crow  = df_cof.loc[atype]
        rc    = float(crow["ref_cust"])
        ref_cust_arr[i] = rc
        rep_cost_arr[i] = float(crow["rep_cost"])
        maint_h_arr[i]  = float(crow["maint_h"])
        labor_d_arr[i]  = float(crow["labor_d"])

        cust_f  = min(cust / max(rc, 1), 5.0)
        loc_env = 1.30 if "수변" in water else 1.00
        size_f  = {"소":0.80,"중":1.00,"대":1.30}.get(size, 1.00)

        cF  = float(crow["cof_F"])
        cS  = float(crow["cof_S"])
        cE  = float(crow["cof_E"]) * loc_env * size_f
        cNP = float(crow["cof_NP"]) * cust_f
        cof_F[i]=cF; cof_S[i]=cS; cof_E[i]=round(cE,1); cof_NP[i]=round(cNP,1)
        cof_tot[i] = cF + cS + cE + cNP

        # 의무교체: 나이 AND 현재건강지수가 HI5(>=8.5, 최상위 위험등급) 동시 충족
        if atype in df_mand.index:
            mrow = df_mand.loc[atype]
            if age >= int(mrow["mand_age"]) and chs >= 8.5:
                mandatory[i] = 1

        # 입력 교체비용 읽기 (CoF_재무 기준값, 천원)
        rc_val = row.get("교체비용_천원", 0)
        repcost_input[i] = float(rc_val) if pd.notna(rc_val) else 0.0

    # -- β2 / ARF (벡터화, EQ.10 + Figure 5)
    age_arr  = df_in["나이_년"].values.astype(float)
    safe_age = np.maximum(age_arr, 1.0)
    safe_chs = np.maximum(curr_hs, H_NEW + 0.0001)
    beta2    = np.where(
        curr_hs > H_NEW,
        np.minimum(np.log(safe_chs / H_NEW) / safe_age, 2.0 * beta1),
        beta1
    )
    arf = ageing_reduction_factor(curr_hs)

    # -- NPV 투자가치 (완전 벡터화)
    print("  (NPV 투자가치 계산 중 -- 벡터화)...")
    T_max = int(NEL_arr.max()) if NEL_arr.max() > 0 else 60

    npv_pre  = np.zeros(n)
    npv_post = np.zeros(n)

    # 자산군별로 NEL이 다르므로 NEL 별로 그룹 처리
    for nel_val in np.unique(NEL_arr[NEL_arr > 0]):
        mask = (NEL_arr == nel_val)
        if not mask.any():
            continue
        T    = int(nel_val)
        t_v  = np.arange(T, dtype=float)
        disc = 1.0 / (1.0 + discount_rate) ** t_v  # (T,)

        # 교체 전: FHS(t) = CHS × exp((β2/ARF) × t)
        nm = mask.sum()
        rates = (beta2[mask] / arf[mask])[:, None]       # (nm, 1)
        fhs_pre = curr_hs[mask][:, None] * np.exp(rates * t_v[None, :])
        fhs_pre = np.minimum(fhs_pre, 15.0)
        h_pre   = np.maximum(fhs_pre, HSL_arr[mask][:, None].astype(float))
        K_m     = K_arr[mask][:, None]
        C_m     = C_arr[mask][:, None]
        pof_pre = K_m * (1 + C_m*h_pre + (C_m*h_pre)**2/2 + (C_m*h_pre)**3/6)
        # 케이블 구간 길이 보정 (UGC)
        is_ugc = (df_in.loc[mask, "자산유형"] == "지중케이블").values
        if is_ugc.any():
            len_m = pd.to_numeric(
                df_in.loc[mask, "선로길이_m"].where(
                    ~df_in.loc[mask, "선로길이_m"].isin(["해당없음","비고"]), 300
                ), errors="coerce"
            ).fillna(300).values
            len_km = len_m / 1000.0
            pof_pre[is_ugc] *= len_km[is_ugc, None]

        risk_pre = pof_pre * cof_tot[mask][:, None]
        npv_pre[mask] = (risk_pre * disc[None, :]).sum(axis=1)

        # 교체 후: 새 설비 HS(t) = 0.5 × exp(β1_new × t), t=0..T
        b1_new  = np.log(H_EOL / H_NEW) / exp_life[mask][:, None]  # 동일 위치/운전 조건
        fhs_post = H_NEW * np.exp(b1_new * t_v[None, :])
        fhs_post = np.minimum(fhs_post, H_EOL)
        h_post   = np.maximum(fhs_post, HSL_arr[mask][:, None].astype(float))
        pof_post = K_m * (1 + C_m*h_post + (C_m*h_post)**2/2 + (C_m*h_post)**3/6)
        if is_ugc.any():
            pof_post[is_ugc] *= len_km[is_ugc, None]

        risk_post = pof_post * cof_tot[mask][:, None]
        npv_post[mask] = (risk_post * disc[None, :]).sum(axis=1)

    invest_val = npv_pre - npv_post

    # -- HI 밴드 분류 (CNAIM Table 5)
    hi_bands = np.select(
        [curr_hs < 3.0, curr_hs < 5.0, curr_hs < 6.5, curr_hs < 8.5],
        ["HI1", "HI2", "HI3", "HI4"], default="HI5"
    )

    # -- 결과 컬럼 추가
    results["위치계수"]           = np.round(loc_f,  4)
    results["운전계수"]           = np.round(duty_f, 3)
    results["기대수명_년"]         = np.round(exp_life, 1)
    results["β1_초기노화율"]       = np.round(beta1,  6)
    results["초기건강지수"]        = np.round(init_hs, 4)
    results["건강지수계수_HSF"]    = np.round(hs_factor, 4)
    results["현재건강지수_CHS"]    = np.round(curr_hs, 4)
    results["HI밴드"]              = hi_bands
    results["현재PoF"]             = np.round(pof_cur,   7)
    results["CoF_재무_천원"]       = np.round(cof_F,   0)
    results["CoF_안전_천원"]       = np.round(cof_S,   0)
    results["CoF_환경_천원"]       = np.round(cof_E,   0)
    results["CoF_계통_천원"]       = np.round(cof_NP,  0)
    results["CoF합계_천원"]        = np.round(cof_tot, 0)
    results["현재위험도_천원_년"]  = np.round(pof_cur * cof_tot, 3)
    results["β2_예측노화율"]       = np.round(beta2,   6)
    results["노화감소계수_ARF"]    = np.round(arf,     3)
    results["교체전위험도NPV_천원"] = np.round(npv_pre,  2)
    results["교체후위험도NPV_천원"] = np.round(npv_post, 2)
    results["투자가치NPV_천원"]    = np.round(invest_val, 2)
    results["교체비용_천원"]       = repcost_input.astype(int)
    results["순_투자가치NPV_천원"] = np.round(invest_val - repcost_input, 2)
    results["의무교체여부"]        = mandatory
    results["교체비용_만원"]       = rep_cost_arr.astype(int)
    results["공사시간_h"]          = maint_h_arr.astype(int)
    results["투입인력_인일"]       = labor_d_arr.astype(int)

    print(f"  완료: {n:,}기 처리")
    return results


# ==================================================================
# 4. 요약 통계 생성
# ==================================================================

def build_summary(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for atype in df["자산유형"].unique():
        sub  = df[df["자산유형"] == atype]
        mand = sub["의무교체여부"].sum()
        rows.append({
            "자산유형":            atype,
            "수량(기)":             len(sub),
            "비율(%)":             round(len(sub) / len(df) * 100, 1),
            "평균나이(년)":        round(sub["나이_년"].mean(), 1),
            "평균건강지수":        round(sub["현재건강지수_CHS"].mean(), 3),
            "평균PoF":             round(sub["현재PoF"].mean(), 6),
            "평균CoF(천원)":       round(sub["CoF합계_천원"].mean(), 0),
            "평균위험도(천원/년)": round(sub["현재위험도_천원_년"].mean(), 2),
            "평균투자가치(천원)":  round(sub["투자가치NPV_천원"].mean(), 0),
            "의무교체수(기)":       mand,
            "의무교체비용합계(만원)": int(sub[sub["의무교체여부"]==1]["교체비용_만원"].sum()),
            "HI1": (sub["HI밴드"]=="HI1").sum(),
            "HI2": (sub["HI밴드"]=="HI2").sum(),
            "HI3": (sub["HI밴드"]=="HI3").sum(),
            "HI4": (sub["HI밴드"]=="HI4").sum(),
            "HI5": (sub["HI밴드"]=="HI5").sum(),
        })
    return pd.DataFrame(rows)


# ==================================================================
# 5. 출력 Excel 작성
# ==================================================================

HC = "1F4E79"   # Header color (dark blue)
AC = "DEEAF1"   # Alt row color
RC = "FFC7CE"   # Mandatory replace color (light red)
YC = "FFEB9C"   # Highlight color (yellow)

def _hdr(ws, r: int, c: int, txt: str, fill_color: str = HC, size: int = 9):
    cell = ws.cell(r, c, txt)
    cell.font  = Font(bold=True, color="FFFFFF", size=size)
    cell.fill  = PatternFill("solid", fgColor=fill_color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def write_output(df_results: pd.DataFrame, summary_df: pd.DataFrame,
                 discount_rate: float, out_path: Path):
    print(f"\n[Excel 출력 작성 중]...")
    wb = openpyxl.Workbook()

    # -- Sheet 0: 개요 --------------------------------------------
    ws0 = wb.active
    ws0.title = "0_분석개요"
    ws0.column_dimensions["A"].width = 30
    ws0.column_dimensions["B"].width = 30

    ws0["A1"] = "AIP 배전설비 위험도 분석 결과"
    ws0["A1"].font = Font(bold=True, size=14, color=HC)
    ws0["A3"] = "분석 일시";     ws0["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws0["A4"] = "총 분석 설비";  ws0["B4"] = len(df_results)
    ws0["A5"] = "자산군 수";     ws0["B5"] = df_results["자산유형"].nunique()
    ws0["A6"] = "할인율 (%)";    ws0["B6"] = discount_rate * 100
    ws0["A7"] = "방법론";        ws0["B7"] = "CNAIM v2.1 (Ofgem, April 2021)"
    ws0["A8"] = "투자가치";      ws0["B8"] = "NPV(교체전Risk) − NPV(교체후Risk)"

    ws0["A10"] = "자산군별 수량 및 비율"
    ws0["A10"].font = Font(bold=True)
    for c, h in enumerate(["자산유형", "수량(기)", "비율(%)"], 1):
        _hdr(ws0, 11, c, h)
    for r, row in enumerate(summary_df[["자산유형","수량(기)","비율(%)"]].itertuples(index=False), 12):
        for c, v in enumerate(row, 1):
            ws0.cell(r, c, v).alignment = Alignment(horizontal="center")

    # -- Sheet 1: 요약 통계 --------------------------------------
    ws1 = wb.create_sheet("1_요약통계")
    ws1.row_dimensions[1].height = 30
    for c, col in enumerate(summary_df.columns, 1):
        ws1.column_dimensions[get_column_letter(c)].width = 16
        _hdr(ws1, 1, c, col)
    for r, row in enumerate(summary_df.itertuples(index=False), 2):
        for c, v in enumerate(row, 1):
            cell = ws1.cell(r, c, v)
            cell.alignment = Alignment(horizontal="center")
            if r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=AC)

    # -- Sheet 2: 전체 자산 상세 (pandas 고속 쓰기 후 서식)
    print("  전체 결과 시트 작성 중...", end="", flush=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # 이미 wb 사용 중이므로, 여기서는 전체 데이터를 별도로 작성
        df_results.to_excel(writer, sheet_name="2_전체결과", index=False)
        summary_df.to_excel(writer, sheet_name="1_요약통계_raw", index=False)

    # 다시 로드하여 서식 추가
    wb2 = openpyxl.load_workbook(out_path)

    # 전체 결과 시트 서식
    ws_all = wb2["2_전체결과"]
    ws_all.freeze_panes = "A2"
    ws_all.row_dimensions[1].height = 28
    for c_idx, col_name in enumerate(df_results.columns, 1):
        ws_all.column_dimensions[get_column_letter(c_idx)].width = 14
        cell = ws_all.cell(1, c_idx)
        cell.font  = Font(bold=True, color="FFFFFF", size=8)
        cell.fill  = PatternFill("solid", fgColor=HC)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 의무교체 행 강조 + 짝수행 배경
    mand_col = list(df_results.columns).index("의무교체여부") + 1
    for r_idx in range(2, len(df_results) + 2):
        is_mand = ws_all.cell(r_idx, mand_col).value == 1
        for c_idx in range(1, len(df_results.columns) + 1):
            ws_all.cell(r_idx, c_idx).alignment = Alignment(horizontal="center")
            if is_mand:
                ws_all.cell(r_idx, c_idx).fill = PatternFill("solid", fgColor=RC)
            elif r_idx % 2 == 0:
                ws_all.cell(r_idx, c_idx).fill = PatternFill("solid", fgColor=AC)
    print(" 완료")

    # -- Sheet 3: 자산군별 상세 (자산군별 시트 분리)
    sheet_no = 3
    for atype in df_results["자산유형"].unique():
        sub = df_results[df_results["자산유형"] == atype].reset_index(drop=True)
        sname = f"{sheet_no}_{atype}"[:31]  # Excel 시트명 31자 제한
        print(f"  [{atype}] {len(sub):,}기 시트 작성...", end="", flush=True)

        ws_t = wb2.create_sheet(sname)
        ws_t.freeze_panes = "A2"
        ws_t.row_dimensions[1].height = 28

        for c_idx, col_name in enumerate(sub.columns, 1):
            ws_t.column_dimensions[get_column_letter(c_idx)].width = 13
            cell = ws_t.cell(1, c_idx, col_name)
            cell.font  = Font(bold=True, color="FFFFFF", size=8)
            cell.fill  = PatternFill("solid", fgColor=HC)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        rows_list = sub.values.tolist()
        mand_col_t = list(sub.columns).index("의무교체여부") + 1
        for r_idx, row_data in enumerate(rows_list, 2):
            is_mand = (row_data[mand_col_t - 1] == 1)
            for c_idx, val in enumerate(row_data, 1):
                cell = ws_t.cell(r_idx, c_idx)
                if isinstance(val, float) and (np.isnan(val) or np.isinf(val)):
                    cell.value = None
                else:
                    cell.value = val
                cell.alignment = Alignment(horizontal="center")
                if is_mand:
                    cell.fill = PatternFill("solid", fgColor=RC)
                elif r_idx % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=AC)
        sheet_no += 1
        print(" 완료")

    # -- 최종 시트: 투자가치 순위 상위 500
    ws_rank = wb2.create_sheet("99_투자가치순위TOP500")
    top500 = (df_results
              .nlargest(500, "투자가치NPV_천원")
              .reset_index(drop=True))
    top500.insert(0, "순위", range(1, 501))

    key_cols = ["순위", "자산ID", "자산유형", "나이_년",
                "현재건강지수_CHS", "HI밴드", "현재PoF",
                "CoF합계_천원", "현재위험도_천원_년",
                "교체전위험도NPV_천원", "교체후위험도NPV_천원",
                "투자가치NPV_천원", "교체비용_천원", "순_투자가치NPV_천원",
                "의무교체여부", "공사시간_h", "투입인력_인일"]
    top500_out = top500[[c for c in key_cols if c in top500.columns]]

    ws_rank.row_dimensions[1].height = 28
    for c_idx, col_name in enumerate(top500_out.columns, 1):
        ws_rank.column_dimensions[get_column_letter(c_idx)].width = 18
        _hdr(ws_rank, 1, c_idx, col_name)
    for r_idx, row in enumerate(top500_out.itertuples(index=False), 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws_rank.cell(r_idx, c_idx, val)
            cell.alignment = Alignment(horizontal="center")
            if r_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=AC)

    wb2.save(out_path)
    print(f"\n[저장 완료] → {out_path}")


# ==================================================================
# 6. 메인
# ==================================================================

def main():
    print("=" * 66)
    print("  AIP Step 2 -- CNAIM 위험도 분석 및 투자가치(NPV) 산출")
    print("=" * 66)

    if not IN_ASSETS.exists():
        print(f"\n[오류] 입력 파일 없음: {IN_ASSETS}")
        print("  먼저 py -3 aip_01_generate_input.py 를 실행하세요.")
        sys.exit(1)
    if not IN_PARAMS.exists():
        print(f"\n[오류] 파라미터 파일 없음: {IN_PARAMS}")
        sys.exit(1)

    # 파라미터 로드
    print(f"\n[파라미터 로드] {IN_PARAMS.name}")
    discount_rate, df_pof, df_cof, df_mand = load_params()
    print(f"  할인율: {discount_rate*100:.1f}%")
    print(f"  자산군 수: {len(df_pof)}종")

    # 자산 목록 로드
    print(f"\n[자산 목록 로드] {IN_ASSETS.name}")
    df_in = pd.read_excel(IN_ASSETS, sheet_name="전체_자산목록", engine="openpyxl")
    print(f"  {len(df_in):,}기 × {len(df_in.columns)}개 속성 컬럼")

    # 분석 실행
    t0 = datetime.now()
    df_results = run_analysis(df_in, discount_rate, df_pof, df_cof, df_mand)
    elapsed = (datetime.now() - t0).total_seconds()
    print(f"  계산 소요시간: {elapsed:.1f}초")

    # 요약 통계
    summary_df = build_summary(df_results)
    print("\n[요약 통계]")
    print(summary_df[["자산유형","수량(기)","평균건강지수","평균PoF",
                       "평균위험도(천원/년)","평균투자가치(천원)","의무교체수(기)"]].to_string(index=False))

    # 출력 Excel
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = DATA_DIR / f"output_results_{ts}.xlsx"
    write_output(df_results, summary_df, discount_rate, out_path)

    # 전체 총계
    total_invest = df_results["투자가치NPV_천원"].sum()
    total_mand   = df_results["의무교체여부"].sum()
    total_cost   = df_results[df_results["의무교체여부"]==1]["교체비용_천원"].sum()
    print(f"\n[전체 총계]")
    print(f"  총 투자가치(NPV) : {total_invest:,.0f} 천원")
    print(f"  의무교체 설비    : {total_mand:,}기")
    print(f"  의무교체 비용    : {total_cost:,} 천원")
    print("=" * 66)


if __name__ == "__main__":
    main()

