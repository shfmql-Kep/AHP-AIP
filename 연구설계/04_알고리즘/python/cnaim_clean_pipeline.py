"""
Clean CNAIM-AIP data pipeline
=============================

생성 파일은 data 폴더의 3개 Excel로 제한한다.

1. input_assets.xlsx
   - 자산별 PoF 산출 원천 데이터, 평가 데이터, CoF 데이터
2. cnaim_params.xlsx
   - HI 평가와 PoF 변환에 필요한 CNAIM 계수와 룩업값
3. pof_5yr_output.xlsx
   - 2026~2030년 최적화/AHP 입력용 산출값

컬럼명은 모두 영어로 유지한다.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"

INPUT_XLSX = DATA_DIR / "input_assets.xlsx"
PARAM_XLSX = DATA_DIR / "cnaim_params.xlsx"
OUTPUT_XLSX = DATA_DIR / "pof_5yr_output.xlsx"

RANDOM_SEED = 42
PLAN_YEARS = [2026, 2027, 2028, 2029, 2030]
DISCOUNT_RATE = 0.05

H_NEW = 0.5
H_EOL = 5.5
CURRENT_HS_CAP = 10.0
FUTURE_HS_CAP = 15.0


@dataclass(frozen=True)
class AssetSpec:
    asset_type: str
    asset_code: str
    cnaim_category: str
    condition_profile: str
    count: int
    normal_expected_life_years: int
    k_percent: float
    c_value: float
    health_score_limit: float
    location_type: str
    duty_type: str
    pof_length_basis: str
    max_age_years: int
    reference_customers: int
    outage_duration_min: int
    cof_financial_kkrw: int
    cof_safety_kkrw: int
    cof_environment_kkrw: int
    cof_network_kkrw: int


ASSET_SPECS: list[AssetSpec] = [
    AssetSpec(
        "pole_transformer", "TR_P", "6.6/11kV Transformer (PM)", "HV_TR",
        10_000, 60, 0.0078, 1.087, 4.0, "switchgear_transformer",
        "transformer_load", "asset", 69, 40, 95, 14_000, 7_000, 5_500, 9_000,
    ),
    AssetSpec(
        "ground_transformer", "TR_G", "6.6/11kV Transformer (GM)", "HV_TR",
        1_622, 60, 0.0078, 1.087, 4.0, "switchgear_transformer",
        "transformer_load", "asset", 64, 100, 100, 22_000, 7_000, 6_000, 14_000,
    ),
    AssetSpec(
        "overhead_switch", "SW_OH", "HV Switchgear Distribution", "SW_DIST",
        3_514, 55, 0.0067, 1.087, 4.0, "switchgear_transformer",
        "switch_operation", "asset", 59, 80, 78, 11_000, 16_000, 2_000, 22_000,
    ),
    AssetSpec(
        "underground_switch", "SW_UG", "HV Switchgear Distribution", "SW_DIST",
        1_351, 55, 0.0067, 1.087, 4.0, "switchgear_transformer",
        "switch_operation", "asset", 59, 150, 100, 13_000, 18_000, 2_500, 28_000,
    ),
    AssetSpec(
        "overhead_line", "OHL", "OHL Conductor", "OHL_COND",
        3_514, 55, 0.0080, 1.087, 4.0, "ohl_conductor",
        "none", "asset", 64, 30, 60, 7_000, 2_000, 1_500, 4_500,
    ),
    AssetSpec(
        "underground_cable", "UGC", "Non-Pressurised Cable", "CABLE",
        1_892, 40, 0.0050, 1.087, 4.0, "none",
        "none", "per_km", 44, 40, 120, 9_000, 1_500, 2_000, 5_000,
    ),
]


SPEC_BY_TYPE = {spec.asset_type: spec for spec in ASSET_SPECS}


# 상태 튜플: state, factor, cap, collar, severity_score
CONDITION_FACTORS: dict[str, dict[str, dict[str, list[tuple[str, float, float, float, float]]]]] = {
    "HV_TR": {
        "observed": {
            "external_condition": [
                ("no_deterioration", 0.9, 10.0, 0.5, 0.00),
                ("superficial", 1.0, 10.0, 0.5, 0.15),
                ("slight", 1.1, 10.0, 0.5, 0.35),
                ("some", 1.25, 10.0, 3.0, 0.60),
                ("substantial", 1.4, 10.0, 8.0, 0.90),
            ],
            "cable_box_condition": [
                ("no_deterioration", 1.0, 10.0, 0.5, 0.00),
                ("some", 1.1, 10.0, 0.5, 0.50),
                ("substantial", 1.3, 10.0, 0.5, 0.90),
            ],
        },
        "measured": {
            "partial_discharge": [
                ("low", 1.0, 10.0, 0.5, 0.00),
                ("medium", 1.1, 10.0, 0.5, 0.35),
                ("high_unconfirmed", 1.3, 10.0, 5.5, 0.65),
                ("high_confirmed", 1.5, 10.0, 8.0, 0.90),
            ],
            "temperature_reading": [
                ("normal", 1.0, 10.0, 0.5, 0.00),
                ("moderately_high", 1.2, 10.0, 0.5, 0.50),
                ("very_high", 1.4, 10.0, 5.5, 0.90),
            ],
        },
    },
    "SW_DIST": {
        "observed": {
            "external_condition": [
                ("no_deterioration", 0.9, 10.0, 0.5, 0.00),
                ("superficial", 1.0, 10.0, 0.5, 0.15),
                ("some", 1.2, 10.0, 3.0, 0.55),
                ("substantial", 1.4, 10.0, 8.0, 0.90),
            ],
            "oil_gas_leak": [
                ("no_deterioration", 0.9, 10.0, 0.5, 0.00),
                ("superficial", 1.0, 10.0, 0.5, 0.15),
                ("some", 1.1, 10.0, 3.0, 0.50),
                ("substantial", 1.3, 10.0, 8.0, 0.90),
            ],
            "thermography": [
                ("ambient_or_below", 0.9, 10.0, 0.5, 0.00),
                ("above_ambient", 1.0, 10.0, 0.5, 0.30),
                ("substantially_above", 1.1, 10.0, 0.5, 0.80),
            ],
            "internal_operation": [
                ("no_deterioration", 0.9, 10.0, 0.5, 0.00),
                ("superficial", 1.0, 10.0, 0.5, 0.15),
                ("some", 1.2, 10.0, 3.0, 0.55),
                ("substantial", 1.4, 10.0, 8.0, 0.90),
            ],
            "indoor_environment": [
                ("better_than_expected", 0.9, 10.0, 0.5, 0.00),
                ("as_expected", 1.0, 10.0, 0.5, 0.25),
                ("deteriorated", 1.3, 10.0, 0.5, 0.60),
                ("severely_deteriorated", 1.5, 10.0, 0.5, 0.90),
            ],
            "cable_box_condition": [
                ("no_deterioration", 1.0, 10.0, 0.5, 0.00),
                ("some", 1.1, 10.0, 0.5, 0.50),
                ("substantial", 1.3, 10.0, 0.5, 0.90),
            ],
        },
        "measured": {
            "partial_discharge": [
                ("low", 1.0, 10.0, 0.5, 0.00),
                ("medium", 1.1, 10.0, 0.5, 0.35),
                ("high_unconfirmed", 1.3, 10.0, 5.5, 0.65),
                ("high_confirmed", 1.5, 10.0, 8.0, 0.90),
            ],
            "ductor_test": [
                ("as_new", 1.0, 10.0, 0.5, 0.00),
                ("up_to_10pct", 1.1, 10.0, 0.5, 0.50),
                ("over_10pct", 1.3, 10.0, 0.5, 0.90),
            ],
            "oil_test": [
                ("as_new", 1.0, 10.0, 0.5, 0.00),
                ("up_to_10pct", 1.1, 10.0, 0.5, 0.50),
                ("over_10pct", 1.3, 10.0, 0.5, 0.90),
            ],
            "temperature_reading": [
                ("ambient_or_below", 0.9, 10.0, 0.5, 0.00),
                ("above_ambient", 1.0, 10.0, 0.5, 0.30),
                ("substantially_above", 1.1, 10.0, 0.5, 0.80),
            ],
            "trip_test": [
                ("pass", 1.0, 10.0, 0.5, 0.00),
                ("fail", 1.4, 10.0, 0.5, 0.90),
            ],
        },
    },
    "OHL_COND": {
        "observed": {
            "conductor_visual_condition": [
                ("no_deterioration", 0.9, 10.0, 0.5, 0.00),
                ("superficial", 1.1, 10.0, 0.5, 0.30),
                ("some", 1.3, 10.0, 4.0, 0.60),
                ("substantial", 1.4, 10.0, 8.0, 0.90),
            ],
            "midspan_joint_count": [
                ("count_0", 1.0, 10.0, 0.5, 0.00),
                ("count_1", 1.05, 10.0, 0.5, 0.30),
                ("count_2", 1.1, 10.0, 0.5, 0.60),
                ("count_gt2", 1.2, 10.0, 5.5, 0.90),
            ],
        },
        "measured": {
            "conductor_sampling": [
                ("low", 1.0, 5.4, 0.5, 0.00),
                ("medium_normal", 1.1, 10.0, 3.0, 0.45),
                ("high", 1.4, 10.0, 8.0, 0.90),
            ],
            "corrosion_monitoring": [
                ("low", 1.0, 5.4, 0.5, 0.00),
                ("medium_normal", 1.1, 10.0, 3.0, 0.45),
                ("high", 1.4, 10.0, 8.0, 0.90),
            ],
        },
    },
    "CABLE": {
        "observed": {},
        "measured": {
            "sheath_test": [
                ("pass", 1.0, 10.0, 0.5, 0.00),
                ("failed_minor", 1.3, 10.0, 0.5, 0.55),
                ("failed_major", 1.6, 10.0, 5.5, 0.90),
            ],
            "partial_discharge": [
                ("low", 1.0, 10.0, 0.5, 0.00),
                ("medium", 1.15, 10.0, 0.5, 0.45),
                ("high", 1.5, 10.0, 5.5, 0.90),
            ],
            "fault_history": [
                ("none", 1.0, 5.4, 0.5, 0.00),
                ("lt_0_01_per_km", 1.3, 10.0, 0.5, 0.40),
                ("between_0_01_0_1", 1.6, 10.0, 5.5, 0.70),
                ("ge_0_1_per_km", 1.8, 10.0, 8.0, 0.90),
            ],
        },
    },
}


MMI_PARAMS = {
    "HV_TR": {"observed": (2, 1.5, 1.5), "measured": (2, 1.5, 1.5), "top": (4, 1.5, 1.5)},
    "SW_DIST": {"observed": (3, 1.5, 1.5), "measured": (3, 1.5, 1.5), "top": (2, 1.5, 1.5)},
    "OHL_COND": {"observed": (2, 1.5, 1.5), "measured": (2, 1.5, 1.5), "top": (2, 1.5, 1.5)},
    "CABLE": {"observed": (1, 1.5, 1.5), "measured": (3, 1.5, 1.5), "top": (2, 1.5, 1.5)},
}


LOCATION_FACTOR_TABLE = {
    "switchgear_transformer": {
        "coast_distance_km": [(1.0, 2.00), (5.0, 1.50), (10.0, 1.20), (20.0, 1.00), (float("inf"), 1.00)],
        "altitude_m": [(100.0, 0.95), (200.0, 1.00), (300.0, 1.05), (float("inf"), 1.15)],
        "corrosion_index": {1: 0.95, 2: 0.95, 3: 1.00, 4: 1.05, 5: 1.20},
    },
    "ohl_conductor": {
        "coast_distance_km": [(1.0, 1.80), (5.0, 1.45), (10.0, 1.20), (20.0, 1.00), (float("inf"), 0.85)],
        "altitude_m": [(100.0, 0.90), (200.0, 1.00), (300.0, 1.15), (float("inf"), 1.30)],
        "corrosion_index": {1: 0.75, 2: 0.90, 3: 1.00, 4: 1.30, 5: 1.60},
    },
    "none": {
        "coast_distance_km": [(float("inf"), 1.00)],
        "altitude_m": [(float("inf"), 1.00)],
        "corrosion_index": {1: 1.00, 2: 1.00, 3: 1.00, 4: 1.00, 5: 1.00},
    },
}


DUTY_FACTOR_TABLE = {
    "transformer_load": [(50.0, 0.90), (70.0, 0.95), (100.0, 1.00), (float("inf"), 1.40)],
    "switch_operation": {"low": 1.00, "normal": 1.00, "high_auto_recloser": 1.20},
    "none": 1.00,
}


OIL_BREAKPOINTS = {
    "oil_moisture_ppm": [(15.0, 0), (30.0, 2), (40.0, 4), (50.0, 8), (float("inf"), 10)],
    "oil_acidity_mg_koh_g": [(0.15, 2), (0.30, 4), (0.50, 8), (float("inf"), 10)],
    "oil_bd_strength_kv": [(30.0, 10), (40.0, 4), (50.0, 2), (float("inf"), 0)],
}

OIL_FACTOR_BREAKPOINTS = [(250.0, 1.00), (500.0, 1.10), (1000.0, 1.20), (float("inf"), 1.40)]
OIL_COLLAR_BREAKPOINTS = [(1000.0, 0.5), (float("inf"), 5.5)]

DGA_GAS_BREAKPOINTS = {
    "dga_h2_ppm": [(20.0, 0), (40.0, 2), (100.0, 4), (200.0, 10), (float("inf"), 16)],
    "dga_ch4_ppm": [(10.0, 0), (20.0, 2), (50.0, 4), (150.0, 10), (float("inf"), 16)],
    "dga_c2h4_ppm": [(10.0, 0), (20.0, 2), (50.0, 4), (150.0, 10), (float("inf"), 16)],
    "dga_c2h6_ppm": [(10.0, 0), (20.0, 2), (50.0, 4), (150.0, 10), (float("inf"), 16)],
    "dga_c2h2_ppm": [(1.0, 0), (5.0, 2), (20.0, 4), (100.0, 8), (float("inf"), 10)],
}

DGA_WEIGHTS = {
    "dga_h2_ppm": 50,
    "dga_ch4_ppm": 30,
    "dga_c2h4_ppm": 30,
    "dga_c2h6_ppm": 30,
    "dga_c2h2_ppm": 100,
}

DGA_DIVIDER = 220.0
FFA_FACTOR_BREAKPOINTS = [(4.0, 1.00), (5.0, 1.10), (6.0, 1.25), (7.0, 1.40), (float("inf"), 1.60)]


ALL_INPUT_COLUMNS = [
    "asset_id", "asset_type", "asset_code", "cnaim_category",
    "base_age_years", "coast_distance_km", "altitude_m", "corrosion_index",
    "water_area_flag", "size_class", "load_pct", "operation_class", "line_length_m",
    "connected_customers", "outage_duration_min",
    "cof_financial_kkrw", "cof_safety_kkrw", "cof_environment_kkrw", "cof_network_kkrw",
    "base_replacement_cost_kkrw", "contract_factor", "replacement_cost_kkrw",
    "external_condition", "cable_box_condition", "oil_gas_leak", "thermography",
    "internal_operation", "indoor_environment", "partial_discharge", "temperature_reading",
    "ductor_test", "oil_test", "trip_test", "conductor_visual_condition",
    "midspan_joint_count", "conductor_sampling", "corrosion_monitoring",
    "sheath_test", "fault_history", "oil_moisture_ppm", "oil_acidity_mg_koh_g",
    "oil_bd_strength_kv", "dga_h2_ppm", "dga_ch4_ppm", "dga_c2h4_ppm",
    "dga_c2h6_ppm", "dga_c2h2_ppm", "ffa_ppm",
]


def lookup_upper(value: float, table: list[tuple[float, Any]]) -> Any:
    """상한 기반 룩업."""
    for upper, result in table:
        if value <= upper:
            return result
    return table[-1][1]


def mmi_general(factors: list[float], divider1: float = 1.5, divider2: float = 1.5,
                max_combined: int = 2) -> float:
    """CNAIM의 Maximum/Multiple Increment 결합."""
    clean = [f for f in factors if f is not None]
    if not clean:
        return 1.0
    if len(clean) == 1:
        return clean[0]

    gt1 = sorted([f for f in clean if f > 1.0], reverse=True)
    if gt1:
        var1 = gt1[0]
        rest = gt1[1:max_combined]
        var2 = sum(f - 1.0 for f in rest)
        return var1 + var2 / divider1

    asc = sorted(clean)
    var1 = asc[0]
    var2 = asc[1] if len(asc) > 1 else 1.0
    return var1 + (var2 - 1.0) / divider2


def choose_state(states: list[tuple[str, float, float, float, float]], severity: float,
                 rng: np.random.Generator) -> str:
    """잠재 중증도에 가장 가까운 CNAIM 상태를 선택한다."""
    target = float(np.clip(severity + rng.normal(0.0, 0.10), 0.0, 1.0))
    scores = np.array([s[4] for s in states])
    idx = int(np.argmin(np.abs(scores - target)))
    return states[idx][0]


def sample_continuous(good: float, bad: float, severity: np.ndarray,
                      rng: np.random.Generator) -> np.ndarray:
    """연속형 상태값을 중증도와 연동해 생성한다."""
    span = abs(bad - good)
    values = good + (bad - good) * severity + rng.normal(0.0, span * 0.08, len(severity))
    low = min(good, bad)
    high = max(good, bad)
    return np.clip(values, low, high)


def contract_factor(asset_type: str, coast_distance_km: float, altitude_m: float,
                    corrosion_index: int, water_area_flag: int, line_length_m: float) -> float:
    """지역·시공 난이도 기반 도급비율. 범위는 1.1~1.5."""
    score = 0.0
    if asset_type == "underground_cable":
        score += 2.0
    elif asset_type in {"ground_transformer", "underground_switch"}:
        score += 1.5
    elif asset_type in {"overhead_switch", "overhead_line"}:
        score += 0.8
    elif asset_type == "pole_transformer":
        score += 0.3
    else:
        score += 0.8

    if coast_distance_km <= 1.0:
        score += 1.0
    elif coast_distance_km <= 5.0:
        score += 0.6
    elif coast_distance_km <= 10.0:
        score += 0.3

    if altitude_m >= 300:
        score += 0.7
    elif altitude_m >= 200:
        score += 0.4

    if corrosion_index >= 5:
        score += 1.0
    elif corrosion_index >= 4:
        score += 0.5

    if water_area_flag == 1:
        score += 0.5

    if asset_type == "underground_cable":
        if line_length_m >= 500:
            score += 0.7
        elif line_length_m >= 300:
            score += 0.4
    elif asset_type == "overhead_line":
        if line_length_m >= 700:
            score += 0.5
        elif line_length_m >= 500:
            score += 0.3

    step = int(np.clip(np.floor(score), 0, 4))
    return round(1.1 + 0.1 * step, 1)


def replacement_cost(base_cost_kkrw: float, factor: float) -> int:
    """도급비율 반영 교체비용. 100천원 단위 반올림."""
    return int(round(base_cost_kkrw * factor / 100.0) * 100)


def generate_input_assets() -> pd.DataFrame:
    """입력 자산 테이블 생성."""
    rng = np.random.default_rng(RANDOM_SEED)
    rows: list[dict[str, Any]] = []

    for spec in ASSET_SPECS:
        ages = rng.integers(1, spec.max_age_years + 1, size=spec.count)
        age_ratio = np.clip(ages / max(spec.max_age_years, 1), 0.0, 1.0)
        severity = np.clip(0.70 * age_ratio + rng.normal(0.0, 0.18, spec.count), 0.0, 1.0)

        coast = rng.choice([0.5, 3.0, 8.0, 15.0, 30.0], size=spec.count, p=[0.06, 0.14, 0.18, 0.30, 0.32])
        altitude = rng.choice([50, 150, 250, 350], size=spec.count, p=[0.50, 0.30, 0.14, 0.06])
        corrosion = rng.choice([1, 2, 3, 4, 5], size=spec.count, p=[0.08, 0.18, 0.42, 0.22, 0.10])
        water_area = rng.choice([0, 1], size=spec.count, p=[0.88, 0.12])
        size_class = rng.choice(["small", "medium", "large"], size=spec.count, p=[0.30, 0.50, 0.20])

        if spec.duty_type == "transformer_load":
            load_pct = rng.choice([35, 55, 75, 100, 125], size=spec.count, p=[0.12, 0.35, 0.32, 0.16, 0.05])
            operation_class = np.full(spec.count, None, dtype=object)
            line_length = np.zeros(spec.count, dtype=float)
        elif spec.duty_type == "switch_operation":
            load_pct = np.full(spec.count, np.nan)
            operation_class = rng.choice(["low", "normal", "high_auto_recloser"], size=spec.count, p=[0.45, 0.37, 0.18])
            line_length = np.zeros(spec.count, dtype=float)
        else:
            load_pct = np.full(spec.count, np.nan)
            operation_class = np.full(spec.count, None, dtype=object)
            if spec.asset_type == "overhead_line":
                line_length = rng.integers(200, 801, size=spec.count).astype(float)
            elif spec.asset_type == "underground_cable":
                line_length = rng.integers(100, 601, size=spec.count).astype(float)
            else:
                line_length = np.zeros(spec.count, dtype=float)

        connected_customers = np.clip(
            rng.lognormal(np.log(max(spec.reference_customers, 1)), 0.55, size=spec.count).astype(int),
            1, 800,
        )

        condition_values: dict[str, list[Any]] = {col: [None] * spec.count for col in ALL_INPUT_COLUMNS}
        profile_tables = CONDITION_FACTORS[spec.condition_profile]
        for group_name, group_inputs in profile_tables.items():
            for input_name, states in group_inputs.items():
                condition_values[input_name] = [choose_state(states, float(z), rng) for z in severity]

        if spec.condition_profile == "HV_TR":
            condition_values["oil_moisture_ppm"] = np.round(sample_continuous(5.0, 65.0, severity, rng), 3)
            condition_values["oil_acidity_mg_koh_g"] = np.round(sample_continuous(0.05, 0.60, severity, rng), 3)
            condition_values["oil_bd_strength_kv"] = np.round(sample_continuous(65.0, 22.0, severity, rng), 3)
            condition_values["dga_h2_ppm"] = np.round(sample_continuous(5.0, 220.0, severity, rng), 3)
            condition_values["dga_ch4_ppm"] = np.round(sample_continuous(3.0, 160.0, severity, rng), 3)
            condition_values["dga_c2h4_ppm"] = np.round(sample_continuous(3.0, 160.0, severity, rng), 3)
            condition_values["dga_c2h6_ppm"] = np.round(sample_continuous(3.0, 160.0, severity, rng), 3)
            condition_values["dga_c2h2_ppm"] = np.round(sample_continuous(0.2, 110.0, severity, rng), 3)
            condition_values["ffa_ppm"] = np.round(sample_continuous(0.3, 7.5, severity, rng), 3)

        for idx in range(spec.count):
            cf = contract_factor(
                spec.asset_type,
                float(coast[idx]),
                float(altitude[idx]),
                int(corrosion[idx]),
                int(water_area[idx]),
                float(line_length[idx]),
            )
            row = {
                "asset_id": f"{spec.asset_code}-{idx + 1:05d}",
                "asset_type": spec.asset_type,
                "asset_code": spec.asset_code,
                "cnaim_category": spec.cnaim_category,
                "base_age_years": int(ages[idx]),
                "coast_distance_km": float(coast[idx]),
                "altitude_m": float(altitude[idx]),
                "corrosion_index": int(corrosion[idx]),
                "water_area_flag": int(water_area[idx]),
                "size_class": str(size_class[idx]),
                "load_pct": None if pd.isna(load_pct[idx]) else float(load_pct[idx]),
                "operation_class": operation_class[idx],
                "line_length_m": None if line_length[idx] == 0 else float(line_length[idx]),
                "connected_customers": int(connected_customers[idx]),
                "outage_duration_min": int(spec.outage_duration_min),
                "cof_financial_kkrw": int(spec.cof_financial_kkrw),
                "cof_safety_kkrw": int(spec.cof_safety_kkrw),
                "cof_environment_kkrw": int(spec.cof_environment_kkrw),
                "cof_network_kkrw": int(spec.cof_network_kkrw),
                "base_replacement_cost_kkrw": int(spec.cof_financial_kkrw),
                "contract_factor": cf,
                "replacement_cost_kkrw": replacement_cost(spec.cof_financial_kkrw, cf),
            }
            for col in ALL_INPUT_COLUMNS:
                if col not in row:
                    value = condition_values.get(col, [None] * spec.count)[idx]
                    row[col] = value
            rows.append(row)

    return pd.DataFrame(rows, columns=ALL_INPUT_COLUMNS)


def build_asset_params() -> pd.DataFrame:
    """자산별 CNAIM PoF 파라미터."""
    rows = []
    for spec in ASSET_SPECS:
        rows.append({
            "asset_type": spec.asset_type,
            "asset_code": spec.asset_code,
            "cnaim_category": spec.cnaim_category,
            "condition_profile": spec.condition_profile,
            "normal_expected_life_years": spec.normal_expected_life_years,
            "k_percent": spec.k_percent,
            "c_value": spec.c_value,
            "health_score_limit": spec.health_score_limit,
            "location_type": spec.location_type,
            "duty_type": spec.duty_type,
            "pof_length_basis": spec.pof_length_basis,
        })
    return pd.DataFrame(rows)


def add_current_top30_flags(output_df: pd.DataFrame) -> pd.DataFrame:
    """현재년도 Risk와 투자가치 기준 상위 30% 표시 컬럼을 추가한다.

    최적화 후보군을 줄이는 목적이므로 5개년 전체 flag를 만들지 않는다.
    기준연도는 계획기간의 첫해인 2026년이며, 원본 산출값은 변경하지 않는다.
    """
    out = output_df.copy()
    top_n = int(np.ceil(len(out) * 0.30))

    current_year = PLAN_YEARS[0]
    risk_col = f"risk_{current_year}_kkrw"
    value_col = f"investment_value_{current_year}_kkrw"

    out["risk_top30_current"] = 0
    out["investment_value_top30_current"] = 0

    risk_indices = out[risk_col].nlargest(top_n).index
    value_indices = out[value_col].nlargest(top_n).index
    out.loc[risk_indices, "risk_top30_current"] = 1
    out.loc[value_indices, "investment_value_top30_current"] = 1
    out["candidate_top30_current"] = (
        (out["risk_top30_current"] == 1) | (out["investment_value_top30_current"] == 1)
    ).astype(int)
    return out


def build_location_params() -> pd.DataFrame:
    """위치계수 룩업표."""
    rows = []
    for location_type, tables in LOCATION_FACTOR_TABLE.items():
        for factor_type in ["coast_distance_km", "altitude_m"]:
            lower = -float("inf")
            for upper, factor in tables[factor_type]:
                rows.append({
                    "location_type": location_type,
                    "factor_type": factor_type,
                    "lower_exclusive": lower,
                    "upper_inclusive": upper,
                    "class_value": None,
                    "factor": factor,
                })
                lower = upper
        for class_value, factor in tables["corrosion_index"].items():
            rows.append({
                "location_type": location_type,
                "factor_type": "corrosion_index",
                "lower_exclusive": None,
                "upper_inclusive": None,
                "class_value": class_value,
                "factor": factor,
            })
    return pd.DataFrame(rows)


def build_duty_params() -> pd.DataFrame:
    """운전계수 룩업표."""
    rows = []
    lower = -float("inf")
    for upper, factor in DUTY_FACTOR_TABLE["transformer_load"]:
        rows.append({
            "duty_type": "transformer_load",
            "input_name": "load_pct",
            "lower_exclusive": lower,
            "upper_inclusive": upper,
            "class_value": None,
            "factor": factor,
        })
        lower = upper

    for class_value, factor in DUTY_FACTOR_TABLE["switch_operation"].items():
        rows.append({
            "duty_type": "switch_operation",
            "input_name": "operation_class",
            "lower_exclusive": None,
            "upper_inclusive": None,
            "class_value": class_value,
            "factor": factor,
        })

    rows.append({
        "duty_type": "none",
        "input_name": None,
        "lower_exclusive": None,
        "upper_inclusive": None,
        "class_value": None,
        "factor": 1.00,
    })
    return pd.DataFrame(rows)


def build_condition_params() -> pd.DataFrame:
    """관측·측정 상태 보정표."""
    rows = []
    for profile, groups in CONDITION_FACTORS.items():
        for group_name, inputs in groups.items():
            for input_name, states in inputs.items():
                for state, factor, cap, collar, severity in states:
                    rows.append({
                        "condition_profile": profile,
                        "condition_group": group_name,
                        "input_name": input_name,
                        "state": state,
                        "factor": factor,
                        "cap": cap,
                        "collar": collar,
                        "severity_score": severity,
                    })
    return pd.DataFrame(rows)


def build_mmi_params() -> pd.DataFrame:
    """MMI 결합 파라미터."""
    rows = []
    for profile, groups in MMI_PARAMS.items():
        for group_name, (max_combined, divider1, divider2) in groups.items():
            rows.append({
                "condition_profile": profile,
                "condition_group": group_name,
                "max_combined_factors": max_combined,
                "divider_1": divider1,
                "divider_2": divider2,
            })
    return pd.DataFrame(rows)


def build_continuous_params() -> pd.DataFrame:
    """변압기 연속형 진단값 보정표."""
    rows = []
    for input_name, table in OIL_BREAKPOINTS.items():
        lower = -float("inf")
        for upper, score in table:
            rows.append({
                "modifier": "oil_score",
                "input_name": input_name,
                "lower_exclusive": lower,
                "upper_inclusive": upper,
                "score_or_factor": score,
            })
            lower = upper

    lower = -float("inf")
    for upper, factor in OIL_FACTOR_BREAKPOINTS:
        rows.append({
            "modifier": "oil_factor",
            "input_name": "oil_condition_score",
            "lower_exclusive": lower,
            "upper_inclusive": upper,
            "score_or_factor": factor,
        })
        lower = upper

    lower = -float("inf")
    for upper, collar in OIL_COLLAR_BREAKPOINTS:
        rows.append({
            "modifier": "oil_collar",
            "input_name": "oil_condition_score",
            "lower_exclusive": lower,
            "upper_inclusive": upper,
            "score_or_factor": collar,
        })
        lower = upper

    for input_name, table in DGA_GAS_BREAKPOINTS.items():
        lower = -float("inf")
        for upper, score in table:
            rows.append({
                "modifier": "dga_score",
                "input_name": input_name,
                "lower_exclusive": lower,
                "upper_inclusive": upper,
                "score_or_factor": score,
            })
            lower = upper

    lower = -float("inf")
    for upper, factor in FFA_FACTOR_BREAKPOINTS:
        rows.append({
            "modifier": "ffa_factor",
            "input_name": "ffa_ppm",
            "lower_exclusive": lower,
            "upper_inclusive": upper,
            "score_or_factor": factor,
        })
        lower = upper

    return pd.DataFrame(rows)


def build_model_constants() -> pd.DataFrame:
    """HI와 PoF 변환에 직접 필요한 상수."""
    return pd.DataFrame([
        {"constant_name": "h_new", "value": H_NEW, "description": "new asset health score"},
        {"constant_name": "h_eol", "value": H_EOL, "description": "health score at normal expected life"},
        {"constant_name": "current_health_score_cap", "value": CURRENT_HS_CAP, "description": "current health score cap"},
        {"constant_name": "future_health_score_cap", "value": FUTURE_HS_CAP, "description": "future health score cap"},
    ])


def write_excel(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    """간결한 서식을 적용해 Excel 저장."""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 28
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        max_col = ws.max_column
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            header = str(ws.cell(1, col_idx).value or "")
            width = min(max(len(header) + 3, 12), 28)
            ws.column_dimensions[col_letter].width = width
    wb.save(path)


def location_factor(row: pd.Series, spec: AssetSpec) -> float:
    """위치계수 산출."""
    tables = LOCATION_FACTOR_TABLE[spec.location_type]
    coast_factor = lookup_upper(float(row["coast_distance_km"]), tables["coast_distance_km"])
    altitude_factor = lookup_upper(float(row["altitude_m"]), tables["altitude_m"])
    corrosion_factor = tables["corrosion_index"].get(int(row["corrosion_index"]), 1.0)
    return float(max(coast_factor, altitude_factor, corrosion_factor))


def duty_factor(row: pd.Series, spec: AssetSpec) -> float:
    """운전계수 산출."""
    if spec.duty_type == "transformer_load":
        return float(lookup_upper(float(row["load_pct"]), DUTY_FACTOR_TABLE["transformer_load"]))
    if spec.duty_type == "switch_operation":
        return float(DUTY_FACTOR_TABLE["switch_operation"].get(str(row["operation_class"]), 1.0))
    return 1.0


def combine_condition_group(profile: str, group_name: str, row: pd.Series) -> tuple[float, float, float]:
    """한 condition group의 factor/cap/collar 결합."""
    inputs = CONDITION_FACTORS[profile].get(group_name, {})
    if not inputs:
        return 1.0, 10.0, 0.5

    factors: list[float] = []
    caps: list[float] = []
    collars: list[float] = []
    for input_name, states in inputs.items():
        state_value = row.get(input_name)
        match = next((s for s in states if s[0] == state_value), None)
        if match is None:
            match = states[0]
        factors.append(float(match[1]))
        caps.append(float(match[2]))
        collars.append(float(match[3]))

    max_combined, divider1, divider2 = MMI_PARAMS[profile][group_name]
    return mmi_general(factors, divider1, divider2, max_combined), min(caps), max(collars)


def oil_modifier(row: pd.Series) -> tuple[float, float, float]:
    """HV Transformer 오일 시험 modifier."""
    moisture_score = lookup_upper(float(row["oil_moisture_ppm"]), OIL_BREAKPOINTS["oil_moisture_ppm"])
    acidity_score = lookup_upper(float(row["oil_acidity_mg_koh_g"]), OIL_BREAKPOINTS["oil_acidity_mg_koh_g"])
    breakdown_score = lookup_upper(float(row["oil_bd_strength_kv"]), OIL_BREAKPOINTS["oil_bd_strength_kv"])
    score = 80 * moisture_score + 100 * acidity_score + 80 * breakdown_score
    factor = lookup_upper(score, OIL_FACTOR_BREAKPOINTS)
    collar = lookup_upper(score, OIL_COLLAR_BREAKPOINTS)
    return float(factor), 10.0, float(collar)


def dga_modifier(row: pd.Series) -> tuple[float, float, float]:
    """HV Transformer DGA modifier. HV Transformer의 factor는 CNAIM상 1.0으로 둔다."""
    score = 0.0
    for input_name, weight in DGA_WEIGHTS.items():
        score += weight * lookup_upper(float(row[input_name]), DGA_GAS_BREAKPOINTS[input_name])
    collar = float(np.clip(score / DGA_DIVIDER, 0.5, 10.0))
    return 1.0, 10.0, collar


def ffa_modifier(row: pd.Series) -> tuple[float, float, float]:
    """FFA modifier."""
    ffa = float(row["ffa_ppm"])
    factor = lookup_upper(ffa, FFA_FACTOR_BREAKPOINTS)
    collar = float(np.clip(2.39 * (max(ffa, 0.0) ** 0.66), 0.5, 10.0))
    return float(factor), 10.0, collar


def condition_modifier(row: pd.Series, spec: AssetSpec) -> tuple[float, float, float]:
    """최종 condition modifier factor/cap/collar."""
    profile = spec.condition_profile
    observed_factor, observed_cap, observed_collar = combine_condition_group(profile, "observed", row)
    measured_factor, measured_cap, measured_collar = combine_condition_group(profile, "measured", row)

    top_factors = [observed_factor, measured_factor]
    top_caps = [observed_cap, measured_cap]
    top_collars = [observed_collar, measured_collar]

    if profile == "HV_TR":
        oil_factor, oil_cap, oil_collar = oil_modifier(row)
        dga_factor, dga_cap, dga_collar = dga_modifier(row)
        ffa_factor, ffa_cap, ffa_collar = ffa_modifier(row)
        top_factors.extend([oil_factor, dga_factor, ffa_factor])
        top_caps.extend([oil_cap, dga_cap, ffa_cap])
        top_collars.extend([oil_collar, dga_collar, ffa_collar])

    max_combined, divider1, divider2 = MMI_PARAMS[profile]["top"]
    return mmi_general(top_factors, divider1, divider2, max_combined), min(top_caps), max(top_collars)


def pof_from_health_score(health_score: float, spec: AssetSpec, line_length_m: float | None) -> float:
    """CNAIM EQ.3 기반 PoF 변환."""
    h_value = max(float(health_score), spec.health_score_limit)
    c_value = spec.c_value
    k_value = spec.k_percent / 100.0
    pof = k_value * (1 + c_value * h_value + (c_value * h_value) ** 2 / 2 + (c_value * h_value) ** 3 / 6)
    if spec.pof_length_basis == "per_km":
        length = 300.0 if line_length_m is None or pd.isna(line_length_m) else float(line_length_m)
        pof *= length / 1000.0
    return float(pof)


def ageing_reduction_factor(current_health_score: float) -> float:
    """CNAIM Ageing Reduction Factor."""
    if current_health_score < 2.0:
        return 1.0
    if current_health_score > 5.5:
        return 1.5
    return ((current_health_score - 2.0) / 7.0) + 1.0


def current_asset_state(row: pd.Series) -> dict[str, float]:
    """기준연도 현재 건강상태와 PoF 산출에 필요한 값을 계산."""
    spec = SPEC_BY_TYPE[str(row["asset_type"])]
    loc_factor = location_factor(row, spec)
    duty = duty_factor(row, spec)
    expected_life = max(spec.normal_expected_life_years / (loc_factor * duty), 1.0)
    beta1 = np.log(H_EOL / H_NEW) / expected_life
    initial_hs = min(H_NEW * np.exp(beta1 * float(row["base_age_years"])), H_EOL)
    hsf, cap, collar = condition_modifier(row, spec)
    current_hs = float(np.clip(initial_hs * hsf, collar, cap))
    current_hs = min(current_hs, CURRENT_HS_CAP)
    beta2 = beta1 * ageing_reduction_factor(current_hs)
    return {
        "location_factor": loc_factor,
        "duty_factor": duty,
        "expected_life_years": expected_life,
        "beta1": beta1,
        "condition_factor": hsf,
        "condition_cap": cap,
        "condition_collar": collar,
        "current_health_score": current_hs,
        "beta2": beta2,
    }


def pof_series_for_existing_asset(row: pd.Series, state: dict[str, float]) -> list[float]:
    """현 설비를 미교체로 둘 때의 5개년 PoF."""
    spec = SPEC_BY_TYPE[str(row["asset_type"])]
    output = []
    for offset, _year in enumerate(PLAN_YEARS):
        hs = state["current_health_score"] * np.exp(state["beta2"] * offset)
        hs = min(float(hs), FUTURE_HS_CAP)
        output.append(pof_from_health_score(hs, spec, row.get("line_length_m")))
    return output


def pof_series_after_replacement(row: pd.Series, start_offset: int, state: dict[str, float]) -> list[float]:
    """특정 연도에 교체했을 때 잔여 기간의 신규 설비 PoF."""
    spec = SPEC_BY_TYPE[str(row["asset_type"])]
    output = []
    for future_offset in range(start_offset, len(PLAN_YEARS)):
        age_after_replacement = future_offset - start_offset
        new_hs = H_NEW * np.exp(state["beta1"] * age_after_replacement)
        new_hs = min(float(new_hs), H_EOL)
        output.append(pof_from_health_score(new_hs, spec, row.get("line_length_m")))
    return output


def design_life_risk_reduction(
    row: pd.Series,
    state: dict[str, float],
    cof_total: float,
    replacement_offset: int,
) -> float:
    """교체시점부터 설계수명 동안의 할인 Risk 저감량을 계산한다.

    정의:
    - 기존 설비 유지 시: 현재 연령에서 시작해 설계수명만큼 경과한 Risk 합산량
    - 교체 시: 0살 신규 설비에서 시작해 설계수명만큼 경과한 Risk 합산량
    - Risk 저감량 = 기존 설비 누적 Risk - 신규 설비 누적 Risk

    미래 Risk는 투자비용과 동일하게 기준연도 현재가치로 할인한다.
    """
    spec = SPEC_BY_TYPE[str(row["asset_type"])]
    design_life_years = max(int(round(spec.normal_expected_life_years)), 1)
    benefit = 0.0

    for horizon_offset in range(design_life_years):
        calendar_offset = replacement_offset + horizon_offset

        existing_hs = state["current_health_score"] * np.exp(state["beta2"] * calendar_offset)
        existing_hs = min(float(existing_hs), FUTURE_HS_CAP)
        existing_pof = pof_from_health_score(existing_hs, spec, row.get("line_length_m"))

        new_hs = H_NEW * np.exp(state["beta1"] * horizon_offset)
        new_hs = min(float(new_hs), H_EOL)
        new_pof = pof_from_health_score(new_hs, spec, row.get("line_length_m"))

        delta_risk = max(existing_pof - new_pof, 0.0) * cof_total
        discount_factor = 1.0 / ((1.0 + DISCOUNT_RATE) ** calendar_offset)
        benefit += delta_risk * discount_factor

    return float(benefit)


def build_pof_output(input_df: pd.DataFrame) -> pd.DataFrame:
    """5개년 최적화/AHP용 출력 테이블 생성."""
    total_connected_customers = float(input_df["connected_customers"].sum())
    rows: list[dict[str, Any]] = []

    for _, row in input_df.iterrows():
        spec = SPEC_BY_TYPE[str(row["asset_type"])]
        state = current_asset_state(row)
        pofs = pof_series_for_existing_asset(row, state)

        cof_environment = float(row["cof_environment_kkrw"]) * (1.30 if int(row["water_area_flag"]) == 1 else 1.00)
        size_factor = {"small": 0.80, "medium": 1.00, "large": 1.30}.get(str(row["size_class"]), 1.00)
        cof_environment *= size_factor

        customer_factor = min(float(row["connected_customers"]) / max(spec.reference_customers, 1), 5.0)
        cof_network = float(row["cof_network_kkrw"]) * customer_factor
        cof_total = float(row["cof_financial_kkrw"]) + float(row["cof_safety_kkrw"]) + cof_environment + cof_network

        risks = [pof * cof_total for pof in pofs]
        saidi = [
            pof * float(row["outage_duration_min"]) * float(row["connected_customers"]) / total_connected_customers
            for pof in pofs
        ]

        out = {
            "asset_id": row["asset_id"],
            "asset_type": row["asset_type"],
            "asset_code": row["asset_code"],
            "connected_customers": int(row["connected_customers"]),
            "outage_duration_min": int(row["outage_duration_min"]),
            "cof_safety_kkrw": round(float(row["cof_safety_kkrw"]), 3),
            "cof_environment_base_kkrw": round(float(row["cof_environment_kkrw"]), 3),
            "cof_environment_adjusted_kkrw": round(cof_environment, 3),
            "cof_total_kkrw": round(cof_total, 3),
        }

        for offset, year in enumerate(PLAN_YEARS):
            out[f"age_{year}"] = int(row["base_age_years"]) + offset
        for offset, year in enumerate(PLAN_YEARS):
            out[f"pof_{year}"] = round(pofs[offset], 10)
        for offset, year in enumerate(PLAN_YEARS):
            out[f"risk_{year}_kkrw"] = round(risks[offset], 6)
        for offset, year in enumerate(PLAN_YEARS):
            out[f"saidi_{year}_min"] = round(saidi[offset], 10)

        for offset, year in enumerate(PLAN_YEARS):
            cost_pv = float(row["replacement_cost_kkrw"]) / ((1.0 + DISCOUNT_RATE) ** offset)
            out[f"replacement_cost_{year}_kkrw"] = round(cost_pv, 3)

        for offset, year in enumerate(PLAN_YEARS):
            risk_reduction = design_life_risk_reduction(row, state, cof_total, offset)
            cost_pv = out[f"replacement_cost_{year}_kkrw"]
            investment_value = risk_reduction - cost_pv
            investment_efficiency = risk_reduction / cost_pv if cost_pv > 0 else 0.0
            out[f"risk_reduction_{year}_kkrw"] = round(risk_reduction, 6)
            out[f"investment_value_{year}_kkrw"] = round(investment_value, 6)
            out[f"investment_efficiency_{year}"] = round(investment_efficiency, 8)

        for offset, year in enumerate(PLAN_YEARS):
            out[f"bcr_{year}"] = out[f"investment_efficiency_{year}"]

        rows.append(out)

    output_df = pd.DataFrame(rows)
    return add_current_top30_flags(output_df)


def clean_root_data_excels() -> None:
    """data 루트의 중복 Excel을 정리한다. incoming 원본은 건드리지 않는다."""
    keep = {INPUT_XLSX.name, PARAM_XLSX.name, OUTPUT_XLSX.name}
    locked_files = []
    for path in DATA_DIR.glob("*.xlsx"):
        if path.name not in keep:
            try:
                path.unlink(missing_ok=True)
            except PermissionError:
                locked_files.append(path.name)
    if locked_files:
        joined = ", ".join(locked_files)
        print(f"warning: locked files were not deleted: {joined}")


def main() -> None:
    DATA_DIR.mkdir(exist_ok=True)
    clean_root_data_excels()

    input_df = generate_input_assets()
    params = {
        "asset_params": build_asset_params(),
        "location_factor": build_location_params(),
        "duty_factor": build_duty_params(),
        "condition_factor": build_condition_params(),
        "mmi_params": build_mmi_params(),
        "continuous_condition": build_continuous_params(),
        "model_constants": build_model_constants(),
    }
    output_df = build_pof_output(input_df)

    write_excel(INPUT_XLSX, {"assets": input_df})
    write_excel(PARAM_XLSX, params)
    write_excel(OUTPUT_XLSX, {"pof_5yr": output_df})

    print("clean CNAIM-AIP files created")
    print(f"- {INPUT_XLSX}")
    print(f"- {PARAM_XLSX}")
    print(f"- {OUTPUT_XLSX}")
    print(f"assets: {len(input_df):,} rows")
    print(f"output: {len(output_df):,} rows x {len(output_df.columns):,} columns")


if __name__ == "__main__":
    main()
