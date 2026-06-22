from pathlib import Path

from openpyxl import load_workbook


def main() -> None:
    # 한글 파일명과 한글 시트명이 깨지지 않는지 확인하는 점검 스크립트다.
    survey_dir = Path("Survey")
    files = sorted(survey_dir.glob("*.xlsx"), key=lambda path: path.stat().st_mtime)
    if not files:
        raise FileNotFoundError("Survey 폴더에서 xlsx 파일을 찾지 못했습니다.")

    target = files[-1]
    workbook = load_workbook(target, data_only=True)

    print(f"선택 파일: {target.name}")
    print("시트 목록:")
    for sheet_name in workbook.sheetnames:
        print(f"- {sheet_name}")

    if "응답_RAW" in workbook.sheetnames:
        worksheet = workbook["응답_RAW"]
        print(f"응답_RAW 크기: {worksheet.max_row}행 × {worksheet.max_column}열")


if __name__ == "__main__":
    main()
