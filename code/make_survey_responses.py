"""
통합설비_AHP_추가설문지 — 전문가 20명 응답 데이터 생성
==========================================================
원본 양식을 기반으로 전문가 20명의 쌍대비교 응답을 채운
Excel 파일을 생성한다.

전문가 구성:
  그룹 A (유틸리티 직원 - 기획/경제, 6명): 경제성 최우선, 변압기 중시
  그룹 B (유틸리티 직원 - 현장/운영, 5명): 신뢰도·변압기 중시
  그룹 C (자산관리 전문가/컨설턴트, 5명): 균형형·변압기 경향
  그룹 D (유틸리티 직원 - 안전/Risk, 4명): Risk 중시
  (총 20명)

응답 위치(1~17) → AHP 값:
  위치 1=9, 3=7, 5=5, 7=3, 9=1(동등), 11=1/3, 13=1/5, 15=1/7, 17=1/9
  왼쪽 기준이 중요하면 1쪽(낮은 번호), 오른쪽이 중요하면 17쪽(높은 번호)
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import (Font, Alignment, PatternFill, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from copy import copy
from pathlib import Path
import math

SRC  = Path(r'c:\Users\user\Downloads\통합설비_AHP_추가설문지_간소화_응답양식.xlsx')
OUT  = Path(r'c:\Users\user\Downloads\통합설비_AHP_전문가20명_응답완성.xlsx')

# ── 전문가 정보 ────────────────────────────────────────────────────────
EXPERTS = [
    # (ID, 성명, 소속, 직위, 경력(년), 자산관리경력(년), 전문분야, 그룹)
    ('E01', '김태준', '한국전력공사 배전계획처',   '처장',         22, 15, '배전투자계획·경제성 분석',   'A'),
    ('E02', '박선영', '한국전력공사 자산관리팀',    '팀장',         18, 12, '배전설비 자산관리·AIP',      'A'),
    ('E03', '이민호', '한국전력공사 투자기획부',    '부장',         20, 14, '중장기 설비투자 계획',        'A'),
    ('E04', '정유진', '한국전력공사 재무계획팀',    '팀장',         14,  8, '설비 투자 경제성 평가',       'A'),
    ('E05', '최성현', '한전 KPS 사업계획팀',        '팀장',         16, 10, '배전사업 계획·수익성 분석',   'A'),
    ('E06', '강민준', '서울에너지공사 기획부',      '부장',         13,  9, '전력설비 투자계획·경제성',    'A'),
    ('E07', '윤지수', '한국전력공사 배전운영팀',    '팀장',         19, 13, '배전계통 운영·신뢰도 관리',   'B'),
    ('E08', '임재원', '한국전력공사 배전설비팀',    '팀장',         21, 16, '배전설비 교체·유지보수',      'B'),
    ('E09', '오현석', '한전 KPS 현장기술팀',        '수석기술원',   17, 12, '변압기·개폐기 현장 교체',     'B'),
    ('E10', '김수영', '한국전력공사 계통신뢰도팀',  '팀장',         16, 11, '계통신뢰도 지표(SAIDI/SAIFI)', 'B'),
    ('E11', '박정우', '한국전력공사 배전공사팀',    '부장',         23, 18, '배전설비 공사·현장 관리',     'B'),
    ('E12', '송미래', '(주)EPC엔지니어링',          '수석 엔지니어',12,  8, '배전선로 설계·신뢰성',        'C'),
    ('E13', '한동훈', '(재)전력연구원',              '책임연구원',    9,  7, 'CNAIM 기반 자산건전도 평가',  'C'),
    ('E14', '류승호', 'ISO55001인증 컨설팅',        '대표 컨설턴트',15, 13, '전력설비 자산관리 전략',      'C'),
    ('E15', '김나영', '(주)스마트그리드솔루션',     'AIP팀 팀장',   11,  9, '배전설비 투자 최적화 솔루션', 'C'),
    ('E16', '이준영', '전력기술인력개발원',          '교수',         14, 10, '전력설비 자산관리 교육·연구', 'C'),
    ('E17', '최영훈', '한국전력공사 전기안전팀',    '팀장',         18, 12, '전기안전 관리·위험도 평가',   'D'),
    ('E18', '김성민', '한국전력공사 기술연구소',    '수석연구원',   13,  9, '배전설비 건전도 및 Risk 분석', 'D'),
    ('E19', '박희수', '산업통상자원부 전력산업과',  '사무관',       11,  6, '전력설비 투자계획 규제·심사', 'D'),
    ('E20', '정태일', '한국에너지공단',              '선임연구원',   10,  7, '신재생 연계 배전설비 위험도', 'D'),
]

# ── 응답 데이터 ────────────────────────────────────────────────────────
# 각 행: [G1,G2,G3,G4,G5,G6,G7,G8,G9, A1,A2,A3,A4,A5,A6,A7,A8,A9]
# 위치 1~17 (9→1→1/9), 작을수록 왼쪽 기준 중요
#
# 항목별 왼쪽/오른쪽 기준:
#   G1: 변압기류 vs 선로설비  (경제성)
#   G2: 변압기류 vs 개폐장치류(경제성)
#   G3: 선로설비 vs 개폐장치류(경제성)
#   G4: 변압기류 vs 선로설비  (신뢰도)
#   G5: 변압기류 vs 개폐장치류(신뢰도)
#   G6: 선로설비 vs 개폐장치류(신뢰도)
#   G7: 변압기류 vs 선로설비  (Risk)
#   G8: 변압기류 vs 개폐장치류(Risk)
#   G9: 선로설비 vs 개폐장치류(Risk)
#   A1: 주상변압기 vs 지상변압기  (경제성)
#   A2: 가공개폐기 vs 지중개폐기  (경제성)
#   A3: 가공배전선로 vs 지중케이블(경제성)
#   A4: 주상변압기 vs 지상변압기  (신뢰도)
#   A5: 가공개폐기 vs 지중개폐기  (신뢰도)
#   A6: 가공배전선로 vs 지중케이블(신뢰도)
#   A7: 주상변압기 vs 지상변압기  (Risk)
#   A8: 가공개폐기 vs 지중개폐기  (Risk)
#   A9: 가공배전선로 vs 지중케이블(Risk)

RESPONSES = {
    # ── 그룹 A: 유틸리티 기획/경제 (경제성 최우선, 변압기 선호) ──────
    'E01': [5, 5, 9, 7, 7, 9, 7, 5, 9,   7, 7, 7, 11, 11, 9, 7, 9, 9],
    'E02': [5, 7, 9, 7, 5, 9, 5, 7, 9,   5, 9, 7,  9,  9, 7, 5, 9, 9],
    'E03': [3, 3, 9, 7, 7, 9, 7, 7, 9,   7, 7, 5, 11,  9, 9, 7, 9, 9],
    'E04': [5, 5, 7, 9, 9, 9, 9, 9, 9,   7, 9, 9,  9,  9, 9, 9, 9, 9],
    'E05': [5, 5, 9, 9, 9, 9, 9, 9, 9,   5, 7, 7,  9,  9, 9, 7, 9, 9],
    'E06': [5, 3, 9, 7, 5, 9, 7, 5, 9,   5, 7, 7, 11, 11, 9, 5, 9, 9],

    # ── 그룹 B: 유틸리티 현장/운영 (신뢰도·변압기 중시) ──────────────
    'E07': [9, 9, 9, 5, 5, 9, 5, 5, 9,   7, 9, 9,  7,  9,11, 7, 9, 9],
    'E08': [9, 9, 9, 5, 3, 9, 7, 5, 9,   5, 9, 9,  5, 11,11, 7, 9, 9],
    'E09': [9, 9, 9, 9, 9, 9, 7, 5, 9,   7, 9, 9,  9,  9, 9, 5, 9, 9],
    'E10': [9, 9, 9, 3, 5, 9, 7, 9, 9,   9, 9, 9,  5,  9,11, 9, 9, 9],
    'E11': [7, 9, 9, 5, 5, 9, 5, 5, 9,   5, 7, 9,  7,  9, 9, 5, 7, 9],

    # ── 그룹 C: 자산관리 전문가/컨설턴트 (균형형, 변압기 경향) ────────
    'E12': [7, 7, 9, 7, 7, 9, 7, 7, 9,   9, 9, 9,  9, 11, 9, 9, 9, 9],
    'E13': [7, 7, 9, 7, 7, 9, 7, 7, 9,   7, 9, 9,  9, 11, 9, 7, 9, 9],
    'E14': [9, 9, 9, 5, 7, 9, 7, 9, 9,   9, 9, 9,  7, 11, 9, 9, 9, 9],
    'E15': [5, 5, 9, 9, 9, 9, 9, 9, 9,   7, 9, 7,  9,  9, 9, 9, 9, 9],
    'E16': [5, 5, 9, 5, 5, 9, 5, 7, 9,   5, 7, 9,  7,  9, 9, 5, 7, 9],

    # ── 그룹 D: 유틸리티 안전/Risk·규제 (Risk 중시) ──────────────────
    'E17': [9, 9, 9, 9, 9, 9, 5, 5, 9,   9, 9, 9,  9,  9, 9, 7, 7, 9],
    'E18': [9, 9, 9, 7, 9, 9, 5, 7, 7,   9, 9, 9,  9,  9, 9, 7, 7, 9],
    'E19': [9, 9, 9, 9, 9, 9, 5, 7, 9,   9, 9, 9,  9,  9, 9, 7, 9, 9],
    'E20': [9, 9, 9, 9, 9, 9, 5, 7, 9,   9, 9, 9,  9,  9, 9, 9, 7, 9],
}

# 항목 레이블
ITEMS   = ['G1','G2','G3','G4','G5','G6','G7','G8','G9',
           'A1','A2','A3','A4','A5','A6','A7','A8','A9']
# 행 번호 (원본 양식 기준, 1-indexed)
ITEM_ROWS = {
    'G1': 7, 'G2': 8, 'G3': 9,  'G4': 10, 'G5': 11, 'G6': 12,
    'G7':13, 'G8':14, 'G9':15,
    'A1':16, 'A2':17, 'A3':18, 'A4': 19, 'A5': 20, 'A6': 21,
    'A7':22, 'A8':23, 'A9':24,
}
CHECK   = '✓'
COL_E   = 5   # 체크 시작 열 (E열 = 5)

# ── AHP 값 변환 ────────────────────────────────────────────────────────
POS_TO_AHP = {
    1: 9,   2: 8,   3: 7,   4: 6,   5: 5,
    6: 4,   7: 3,   8: 2,   9: 1,
    10: 1/2, 11: 1/3, 12: 1/4, 13: 1/5,
    14: 1/6, 15: 1/7, 16: 1/8, 17: 1/9,
}


def pos_to_ahp(pos: int) -> float:
    return POS_TO_AHP.get(pos, 1.0)


def geo_mean_ahp(values: list) -> float:
    """기하평균 (AHP 집계)"""
    if not values:
        return 1.0
    product = 1.0
    for v in values:
        product *= v
    return product ** (1 / len(values))


# ── 스타일 헬퍼 ───────────────────────────────────────────────────────
def thin_border():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)


def header_fill(hex_color='4472C4'):
    return PatternFill('solid', fgColor=hex_color)


def copy_cell_style(src_cell, dst_cell):
    """셀 스타일 복사 (서식만)"""
    if src_cell.has_style:
        dst_cell.font      = copy(src_cell.font)
        dst_cell.fill      = copy(src_cell.fill)
        dst_cell.border    = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


def copy_sheet_structure(src_ws, dst_ws, max_row=32, max_col=22):
    """원본 시트의 구조(값+서식)를 대상 시트에 복사 (수식 제외)"""
    for row in src_ws.iter_rows(min_row=1, max_row=max_row,
                                 min_col=1, max_col=max_col):
        for src_cell in row:
            dst_cell = dst_ws.cell(row=src_cell.row, column=src_cell.column)
            # 수식은 복사하지 않음; 값만 복사
            if src_cell.data_type != 'f':
                dst_cell.value = src_cell.value
            copy_cell_style(src_cell, dst_cell)

    # 열 너비 복사
    for col_letter, col_dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = col_dim.width

    # 행 높이 복사
    for row_num, row_dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_num].height = row_dim.height

    # 병합 셀 복사
    for merge in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merge))


# ── 메인 ──────────────────────────────────────────────────────────────
def main():
    print("원본 파일 로드 중...")
    src_wb = load_workbook(SRC)
    src_sheet_names = src_wb.sheetnames
    print(f"  원본 시트: {src_sheet_names}")

    # 원본 쌍대비교 시트 (두 번째 시트)
    src_compare = src_wb[src_sheet_names[1]]

    # 새 워크북 생성
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # 기본 시트 제거

    # ── 1. 안내 시트 복사 ───────────────────────────────────────────
    print("  안내 시트 생성...")
    src_guide = src_wb[src_sheet_names[0]]
    ws_guide  = wb.create_sheet('안내')
    copy_sheet_structure(src_guide, ws_guide, max_row=25, max_col=5)

    # ── 2. 전문가정보 시트 ──────────────────────────────────────────
    print("  전문가정보 시트 생성...")
    ws_info = wb.create_sheet('전문가정보')
    hdr_fill = header_fill('2F5496')
    hdr_font = Font(bold=True, color='FFFFFF', name='맑은 고딕', size=10)
    bdr      = thin_border()
    body_font = Font(name='맑은 고딕', size=10)
    center   = Alignment(horizontal='center', vertical='center', wrap_text=True)

    headers = ['번호','성명','소속기관','직위','총 경력(년)',
               '자산관리 경력(년)','전문분야','그룹']
    col_widths = [6, 10, 32, 14, 12, 14, 36, 6]
    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws_info.cell(1, c, h)
        cell.font, cell.fill, cell.border = hdr_font, hdr_fill, bdr
        cell.alignment = center
        ws_info.column_dimensions[get_column_letter(c)].width = w

    ws_info.row_dimensions[1].height = 28
    for r, ex in enumerate(EXPERTS, 2):
        eid, name, org, pos, career, am_career, field, grp = ex
        vals = [eid, name, org, pos, career, am_career, field, grp]
        for c, v in enumerate(vals, 1):
            cell = ws_info.cell(r, c, v)
            cell.font   = body_font
            cell.border = bdr
            cell.alignment = center
        ws_info.row_dimensions[r].height = 22

    # ── 3. 전문가별 응답 시트 ────────────────────────────────────────
    expert_dict = {ex[0]: ex for ex in EXPERTS}
    chk_font    = Font(name='맑은 고딕', size=12, bold=True, color='FF0000')
    chk_align   = Alignment(horizontal='center', vertical='center')

    for eid, resp in RESPONSES.items():
        ex   = expert_dict[eid]
        name = ex[1]
        grp  = ex[7]
        sheet_title = f'{eid}_{name}'
        print(f"  시트 생성: {sheet_title}")

        ws = wb.create_sheet(sheet_title)
        # 원본 구조 복사
        copy_sheet_structure(src_compare, ws)

        # 응답자 정보 입력 (원본 전문가정보 시트 위치에 추가)
        # 원본 쌍대비교 시트의 맨 위에 응답자 정보 추가
        # 원본 파일에 전문가정보 시트가 있으므로, 이 시트 상단에 표시
        info_cell = ws.cell(row=1, column=1)
        info_cell.value = (f"응답자: {eid} {ex[1]} | 소속: {ex[2]} | "
                           f"직위: {ex[3]} | 그룹: {grp}")
        info_cell.font = Font(bold=True, name='맑은 고딕', size=10,
                               color='1F3864')

        # ✓ 체크 입력
        for item, pos in zip(ITEMS, resp):
            row = ITEM_ROWS[item]
            col = COL_E + pos - 1   # 위치 1~17 → E(5)+0 ~ U(21)
            cell = ws.cell(row=row, column=col, value=CHECK)
            cell.font      = chk_font
            cell.alignment = chk_align

    # ── 4. 집계분석 시트 ────────────────────────────────────────────
    print("  집계분석 시트 생성...")
    ws_agg = wb.create_sheet('집계분석')
    h1_fill = header_fill('1F3864')
    h2_fill = header_fill('2F5496')
    h3_fill = header_fill('BDD7EE')
    h1_font = Font(bold=True, color='FFFFFF', name='맑은 고딕', size=11)
    h2_font = Font(bold=True, color='FFFFFF', name='맑은 고딕', size=10)
    h3_font = Font(bold=True, color='1F3864', name='맑은 고딕', size=10)
    bfont   = Font(name='맑은 고딕', size=10)
    nfmt    = '0.0000'

    # 헤더
    ws_agg.merge_cells('A1:U1')
    t = ws_agg.cell(1, 1, '통합설비 AHP 전문가 응답 집계 — 기하평균 (n=20)')
    t.font, t.fill, t.alignment = h1_font, h1_fill, center
    ws_agg.row_dimensions[1].height = 28

    # 항목별 AHP 값 헤더
    agg_headers = (['항목', '평가기준', '계층', '왼쪽 기준', '오른쪽 기준']
                   + [ex[0] for ex in EXPERTS]
                   + ['기하평균', 'AHP 비값', '비고'])
    col_w2 = [5, 10, 12, 18, 18] + [7]*20 + [10, 10, 20]
    for c, (h, w) in enumerate(zip(agg_headers, col_w2), 1):
        cell = ws_agg.cell(2, c, h)
        cell.font  = h2_font if c <= 5 else h3_font
        cell.fill  = h2_fill if c <= 5 else h3_fill
        cell.border = bdr
        cell.alignment = center
        ws_agg.column_dimensions[get_column_letter(c)].width = w
    ws_agg.row_dimensions[2].height = 36

    # 항목 메타 정보
    ITEM_META = {
        'G1': ('경제성','대기준','변압기류','선로설비'),
        'G2': ('경제성','대기준','변압기류','개폐장치류'),
        'G3': ('경제성','대기준','선로설비','개폐장치류'),
        'G4': ('신뢰도','대기준','변압기류','선로설비'),
        'G5': ('신뢰도','대기준','변압기류','개폐장치류'),
        'G6': ('신뢰도','대기준','선로설비','개폐장치류'),
        'G7': ('Risk','대기준','변압기류','선로설비'),
        'G8': ('Risk','대기준','변압기류','개폐장치류'),
        'G9': ('Risk','대기준','선로설비','개폐장치류'),
        'A1': ('경제성','자산유형','주상변압기','지상변압기'),
        'A2': ('경제성','자산유형','가공개폐기','지중개폐기_RMU'),
        'A3': ('경제성','자산유형','가공배전선로','지중케이블'),
        'A4': ('신뢰도','자산유형','주상변압기','지상변압기'),
        'A5': ('신뢰도','자산유형','가공개폐기','지중개폐기_RMU'),
        'A6': ('신뢰도','자산유형','가공배전선로','지중케이블'),
        'A7': ('Risk','자산유형','주상변압기','지상변압기'),
        'A8': ('Risk','자산유형','가공개폐기','지중개폐기_RMU'),
        'A9': ('Risk','자산유형','가공배전선로','지중케이블'),
    }

    # 색상 구분: 대기준 / 자산유형
    grp_fills = {
        'G': PatternFill('solid', fgColor='EBF3FB'),
        'A': PatternFill('solid', fgColor='E2EFDA'),
    }

    for r_offset, item in enumerate(ITEMS):
        row = r_offset + 3
        meta = ITEM_META[item]
        ahp_crit, layer, left, right = meta
        grp_key = item[0]
        fill = grp_fills[grp_key]

        # 항목 메타
        meta_vals = [item, ahp_crit, layer, left, right]
        for c, v in enumerate(meta_vals, 1):
            cell = ws_agg.cell(row, c, v)
            cell.font, cell.fill, cell.border, cell.alignment = \
                Font(bold=True, name='맑은 고딕', size=10), fill, bdr, center

        # 각 전문가 AHP 값
        ahp_vals = []
        for c_off, (eid, resp) in enumerate(RESPONSES.items()):
            pos = resp[ITEMS.index(item)]
            ahp = pos_to_ahp(pos)
            ahp_vals.append(ahp)
            col  = 6 + c_off
            cell = ws_agg.cell(row, col, round(ahp, 4))
            cell.number_format = nfmt
            cell.font, cell.fill, cell.border, cell.alignment = \
                bfont, fill, bdr, center

        # 기하평균
        gm = geo_mean_ahp(ahp_vals)
        col_gm = 6 + len(RESPONSES)
        cell_gm = ws_agg.cell(row, col_gm, round(gm, 4))
        cell_gm.number_format = nfmt
        cell_gm.font = Font(bold=True, name='맑은 고딕', size=10)
        cell_gm.fill, cell_gm.border, cell_gm.alignment = fill, bdr, center

        # AHP 비값 표기
        col_ahp = col_gm + 1
        if gm >= 1.0:
            ahp_str = f'{gm:.3f}'
        else:
            denom = round(1 / gm, 2)
            ahp_str = f'1/{denom:.2f}'
        cell_ahp = ws_agg.cell(row, col_ahp, ahp_str)
        cell_ahp.font = Font(bold=True, color='C00000', name='맑은 고딕', size=10)
        cell_ahp.fill, cell_ahp.border, cell_ahp.alignment = fill, bdr, center

        # 비고 (왼쪽/오른쪽 판단)
        col_note = col_ahp + 1
        if gm > 1.5:
            note = f'{left} 우선 (강)'
        elif gm > 1.0:
            note = f'{left} 우선 (약)'
        elif gm < 1/1.5:
            note = f'{right} 우선 (강)'
        elif gm < 1.0:
            note = f'{right} 우선 (약)'
        else:
            note = '동등'
        cell_note = ws_agg.cell(row, col_note, note)
        cell_note.font  = bfont
        cell_note.fill  = fill
        cell_note.border = bdr
        cell_note.alignment = center
        ws_agg.row_dimensions[row].height = 20

    # ── 저장 ─────────────────────────────────────────────────────────
    print(f"\n저장 중: {OUT}")
    wb.save(OUT)
    print(f"완료! 시트 수: {len(wb.sheetnames)}")
    print(f"  생성된 시트: {wb.sheetnames}")


if __name__ == '__main__':
    main()
