"""NSGA 115% 총량제약 결과를 기존 형식의 비교표로 정리한다."""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent.parent
OUTPUT_DIR = BASE_DIR / "outputs"
YEARS = [2026, 2027, 2028, 2029, 2030]

METHOD_LABELS = {
    "investment_value_nsga_kpi": "투자가치 NSGA",
    "local_pi_ahp_nsga_kpi": "PI NSGA",
}
BASELINE_METHOD = "investment_value_nsga_kpi"
COMPARISON_METHOD = "local_pi_ahp_nsga_kpi"

DISPLAY_COLS = [
    "기법",
    "구분",
    "투자대수",
    "투자비용",
    "Risk 저감량",
    "투자가치",
    "SAIDI",
    "PI",
]


def latest_nsga_workbook() -> Path:
    """가장 최근의 타임스탬프 NSGA 결과 파일을 찾는다."""
    candidates = sorted(
        OUTPUT_DIR.glob("nsga_portfolio_optimization_matlab_20*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError("타임스탬프 NSGA 결과 파일을 찾을 수 없습니다.")
    return candidates[0]


def is_primary_representative(value: object) -> bool:
    """대표해 유형이 max_primary를 포함하는지 확인한다."""
    return "max_primary" in str(value)


def safe_divide(numerator: float, denominator: float) -> float:
    """0 나누기를 방지한다."""
    if denominator == 0 or pd.isna(denominator):
        return 0.0
    return float(numerator) / float(denominator)


def pct_change(value: float, baseline: float) -> float:
    """기준 대비 증감률을 계산한다."""
    if baseline == 0 or pd.isna(baseline):
        return np.nan
    return (float(value) - float(baseline)) / abs(float(baseline)) * 100.0


def format_value(value: float, metric: str) -> str:
    """논문 표에 넣기 쉬운 표시 문자열을 만든다."""
    if pd.isna(value):
        return "-"
    if metric == "투자대수":
        return f"{value:,.0f}"
    if metric in {"SAIDI", "PI", "투자효율"}:
        return f"{value:,.6f}"
    return f"{value:,.0f}"


def build_primary_annual(annual: pd.DataFrame) -> pd.DataFrame:
    """max_primary 대표해의 연도별 투자효과를 정리한다."""
    primary = annual[annual["representative_type"].map(is_primary_representative)].copy()
    primary["method_label"] = primary["method"].map(METHOD_LABELS).fillna(primary["method"])
    primary["investment_efficiency"] = primary.apply(
        lambda r: safe_divide(r["risk_reduction_kkrw"], r["investment_cost_kkrw"]),
        axis=1,
    )
    primary = primary[
        [
            "scenario_id",
            "method",
            "method_label",
            "solution_id",
            "representative_type",
            "year",
            "selected_count",
            "investment_cost_kkrw",
            "risk_reduction_kkrw",
            "investment_value_kkrw",
            "saidi_at_selection_min",
            "local_pi_ahp",
            "investment_efficiency",
            "baseline_saidi_min",
            "saidi_after_cumulative_min",
            "saidi_cap_min",
            "saidi_cap_ok",
            "baseline_risk_kkrw",
            "risk_after_cumulative_kkrw",
            "risk_cap_kkrw",
            "risk_cap_ok",
        ]
    ]
    return primary


def append_total_rows(primary: pd.DataFrame) -> pd.DataFrame:
    """연도별 표 아래에 합산 행을 추가한다."""
    total_rows = []
    for _, sub in primary.groupby(["scenario_id", "method", "method_label", "solution_id", "representative_type"]):
        last = sub.sort_values("year").iloc[-1]
        total_rows.append(
            {
                "scenario_id": last["scenario_id"],
                "method": last["method"],
                "method_label": last["method_label"],
                "solution_id": last["solution_id"],
                "representative_type": last["representative_type"],
                "year": "합산",
                "selected_count": sub["selected_count"].sum(),
                "investment_cost_kkrw": sub["investment_cost_kkrw"].sum(),
                "risk_reduction_kkrw": sub["risk_reduction_kkrw"].sum(),
                "investment_value_kkrw": sub["investment_value_kkrw"].sum(),
                "saidi_at_selection_min": sub["saidi_at_selection_min"].sum(),
                "local_pi_ahp": sub["local_pi_ahp"].sum(),
                "investment_efficiency": safe_divide(
                    sub["risk_reduction_kkrw"].sum(), sub["investment_cost_kkrw"].sum()
                ),
                "baseline_saidi_min": np.nan,
                "saidi_after_cumulative_min": last["saidi_after_cumulative_min"],
                "saidi_cap_min": last["saidi_cap_min"],
                "saidi_cap_ok": bool(sub["saidi_cap_ok"].all()),
                "baseline_risk_kkrw": np.nan,
                "risk_after_cumulative_kkrw": last["risk_after_cumulative_kkrw"],
                "risk_cap_kkrw": last["risk_cap_kkrw"],
                "risk_cap_ok": bool(sub["risk_cap_ok"].all()),
            }
        )
    return pd.concat([primary, pd.DataFrame(total_rows)], ignore_index=True)


def build_display_table(primary_with_total: pd.DataFrame) -> pd.DataFrame:
    """기존 형식의 표시용 연도별/합산표를 만든다."""
    rows = []
    for _, row in primary_with_total.iterrows():
        rows.append(
            {
                "기법": row["method_label"],
                "구분": str(row["year"]),
                "투자대수": row["selected_count"],
                "투자비용": row["investment_cost_kkrw"],
                "Risk 저감량": row["risk_reduction_kkrw"],
                "투자가치": row["investment_value_kkrw"],
                "SAIDI": row["saidi_at_selection_min"],
                "PI": row["local_pi_ahp"],
            }
        )
    display = pd.DataFrame(rows)
    order = {2026: 1, 2027: 2, 2028: 3, 2029: 4, 2030: 5, "합산": 6}
    display["_method_order"] = display["기법"].map({"투자가치 NSGA": 1, "PI NSGA": 2}).fillna(99)
    display["_year_order"] = display["구분"].map(lambda x: order.get(int(x), 99) if str(x).isdigit() else order.get(x, 99))
    display = display.sort_values(["_method_order", "_year_order"]).drop(columns=["_method_order", "_year_order"])
    return display[DISPLAY_COLS]


def build_total_comparison(primary_with_total: pd.DataFrame) -> pd.DataFrame:
    """투자가치 NSGA 대비 PI NSGA의 합산 성과 비교표를 만든다."""
    total = primary_with_total[primary_with_total["year"].astype(str) == "합산"].copy()
    total = total.set_index("method")
    baseline = total.loc[BASELINE_METHOD]
    comparison = total.loc[COMPARISON_METHOD]

    metric_map = {
        "투자대수": "selected_count",
        "투자비용": "investment_cost_kkrw",
        "Risk 저감량": "risk_reduction_kkrw",
        "투자가치": "investment_value_kkrw",
        "SAIDI": "saidi_at_selection_min",
        "PI": "local_pi_ahp",
        "투자효율": "investment_efficiency",
    }
    rows = []
    for label, col in metric_map.items():
        base_value = float(baseline[col])
        comp_value = float(comparison[col])
        change = pct_change(comp_value, base_value)
        rows.append(
            {
                "구분": label,
                "베이스라인(투자가치 NSGA)": format_value(base_value, label),
                "비교(PI NSGA)": f"{format_value(comp_value, label)} ({change:+.2f}%)",
                "baseline_value": base_value,
                "comparison_value": comp_value,
                "change_vs_baseline_pct": change,
            }
        )
    return pd.DataFrame(rows)


def build_constraint_check(primary_with_total: pd.DataFrame) -> pd.DataFrame:
    """SAIDI와 Risk 총량 제약 충족 여부를 정리한다."""
    rows = []
    yearly = primary_with_total[primary_with_total["year"].astype(str) != "합산"].copy()
    for _, row in yearly.iterrows():
        rows.append(
            {
                "기법": row["method_label"],
                "연도": int(row["year"]),
                "SAIDI after": row["saidi_after_cumulative_min"],
                "SAIDI cap": row["saidi_cap_min"],
                "SAIDI OK": bool(row["saidi_cap_ok"]),
                "Risk after": row["risk_after_cumulative_kkrw"],
                "Risk cap": row["risk_cap_kkrw"],
                "Risk OK": bool(row["risk_cap_ok"]),
            }
        )
    return pd.DataFrame(rows)


def write_result_tables(workbook_path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    """결과 workbook에 비교표 시트를 추가한다."""
    with pd.ExcelWriter(workbook_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)


def format_result_sheets(workbook_path: Path, sheet_names: list[str]) -> None:
    """비교표 시트의 가독성을 높인다."""
    wb = load_workbook(workbook_path)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    total_fill = PatternFill("solid", fgColor="D9EAF7")

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            if any(str(cell.value) == "합산" for cell in row):
                for cell in row:
                    cell.fill = total_fill
                    cell.font = Font(bold=True)
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if isinstance(cell.value, float):
                    cell.number_format = "#,##0.000000"
                elif isinstance(cell.value, int):
                    cell.number_format = "#,##0"
        for col_idx in range(1, ws.max_column + 1):
            max_len = 10
            for row_idx in range(1, min(ws.max_row, 80) + 1):
                value = ws.cell(row_idx, col_idx).value
                if value is not None:
                    max_len = max(max_len, len(str(value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 36)
    wb.save(workbook_path)


def main() -> None:
    """비교표를 생성한다."""
    workbook_path = latest_nsga_workbook()
    annual = pd.read_excel(workbook_path, sheet_name="annual_summary")
    representative = pd.read_excel(workbook_path, sheet_name="representative_summary")
    run_summary = pd.read_excel(workbook_path, sheet_name="run_summary")

    primary = build_primary_annual(annual)
    primary_with_total = append_total_rows(primary)
    display = build_display_table(primary_with_total)
    total_comparison = build_total_comparison(primary_with_total)
    constraint_check = build_constraint_check(primary_with_total)

    sheets = {
        "comparison_annual_raw": primary_with_total,
        "comparison_annual_display": display,
        "comparison_total_display": total_comparison,
        "constraint_check_display": constraint_check,
        "representatives_used": representative[representative["representative_type"].map(is_primary_representative)],
        "run_summary_copy": run_summary,
    }
    write_result_tables(workbook_path, sheets)
    format_result_sheets(workbook_path, list(sheets))

    print(workbook_path)
    print(display.to_string(index=False))
    print(total_comparison[["구분", "베이스라인(투자가치 NSGA)", "비교(PI NSGA)"]].to_string(index=False))


if __name__ == "__main__":
    main()
