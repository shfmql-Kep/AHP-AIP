"""
AIP 배전설비 위험도 분석 시스템 v2.0
CNAIM v2.1 기반 합성 데이터 생성 및 투자가치(NPV) 산출

[처리 순서]
  1. input_parameters.xlsx 읽기 (없으면 자동 생성)
  2. 자산군별 5,000개 합성 데이터 생성
     - 나이, 위치계수, 운전계수, 관측/측정 상태 등 랜덤 생성
  3. CNAIM EQ.3~12 기반 Health Score → PoF 산출
  4. CoF (재무/안전/환경/계통성능) 4항목 산출 → 현재 Risk
  5. 투자가치(NPV) 산출
     - 교체 전 Risk NPV : 현재 시점부터 설계수명 동안
     - 교체 후 Risk NPV : 새 설비 0세 → 설계수명까지
     - 투자가치 = 교체전 NPV - 교체후 NPV
  6. 결과를 output_results_YYYYMMDD_HHMMSS.xlsx 저장

사용법:
  python aip_risk_analysis.py
"""

import numpy as np
import pandas as pd
import os
import sys
from datetime import datetime
from pathlib import Path

# ── Excel 라이브러리 확인
try:
    import openpyxl
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                  numbers)
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    print("openpyxl 설치 필요: pip install openpyxl")
    sys.exit(1)

# ══════════════════════════════════════════════════════════════════
# 경로 설정
# ══════════════════════════════════════════════════════════════════
BASE_DIR   = Path(__file__).resolve().parent.parent
DATA_DIR   = BASE_DIR / "data"
INPUT_PATH = DATA_DIR / "input_parameters.xlsx"
DATA_DIR.mkdir(exist_ok=True)


# ══════════════════════════════════════════════════════════════════
# 1. 입력 Excel 템플릿 자동 생성
# ══════════════════════════════════════════════════════════════════

def create_input_template():
    """입력 파라미터 Excel 템플릿 생성 (CNAIM v2.1 기반)"""
    wb = openpyxl.Workbook()

    # ── 색상 정의
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    SUB_FILL    = PatternFill("solid", fgColor="2E75B6")
    YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
    GREEN_FILL  = PatternFill("solid", fgColor="E2EFDA")

    def header_style(cell, text, fill=HEADER_FILL):
        cell.value = text
        cell.font  = Font(bold=True, color="FFFFFF", size=11)
        cell.fill  = fill
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)

    def data_style(cell, value):
        cell.value = value
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Sheet 1: 분석 설정 ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = "1_분석설정"
    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 40

    header_style(ws1["A1"], "항목")
    header_style(ws1["B1"], "값")
    header_style(ws1["C1"], "비고")
    ws1.row_dimensions[1].height = 25

    settings = [
        ("할인율 (%)",            5.0,    "NPV 계산용 연간 할인율"),
        ("자산군별 생성 수량",     5000,   "각 자산군 당 합성 데이터 수"),
        ("랜덤 시드",              42,     "재현성을 위한 난수 시드"),
    ]
    for i, (item, val, note) in enumerate(settings, 2):
        ws1[f"A{i}"] = item
        ws1[f"B{i}"] = val
        ws1[f"B{i}"].fill = YELLOW_FILL
        ws1[f"B{i}"].alignment = Alignment(horizontal="center")
        ws1[f"C{i}"] = note

    # ── Sheet 2: 자산군 PoF 파라미터 ────────────────────────────
    ws2 = wb.create_sheet("2_PoF파라미터")
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 12

    for col, txt in enumerate(["자산군", "K값(%)", "C값", "건강지수한계(HSL)", "설계수명(NEL,년)"], 1):
        header_style(ws2.cell(1, col), txt)

    pof_params = [
        # (자산군, K%, C, HSL, NEL) - CNAIM Table 20, 21
        ("주상변압기", 0.0078, 1.087, 4, 60),
        ("지상변압기", 0.0078, 1.087, 4, 60),
        ("가공개폐기", 0.0067, 1.087, 4, 55),
        ("지중개폐기", 0.0052, 1.087, 4, 55),
        ("배전선로",   0.0285, 1.087, 4, 55),
    ]
    for r, row in enumerate(pof_params, 2):
        for c, val in enumerate(row, 1):
            cell = ws2.cell(r, c, val)
            cell.alignment = Alignment(horizontal="center")
            if r % 2 == 0:
                cell.fill = GREEN_FILL

    # ── Sheet 3: CoF 기준값 ──────────────────────────────────────
    ws3 = wb.create_sheet("3_CoF기준값")
    for c, txt in enumerate(["자산군", "재무CoF(천원)", "안전CoF(천원)", "환경CoF(천원)",
                              "계통성능CoF(천원)", "참조(CNAIM Table 16)"], 1):
        ws3.column_dimensions[get_column_letter(c)].width = 18
        header_style(ws3.cell(1, c), txt)

    cof_ref = [
        # CNAIM Table 16 참조값 × 환율/국내조건 보정
        # (자산군, F, S, E, NP)
        ("주상변압기", 14_000,  7_000,  5_500,  9_000,  "6.6/11kV Transformer (GM) 기반"),
        ("지상변압기", 22_000,  7_000,  6_000, 14_000,  "용량 큰 변압기 기준"),
        ("가공개폐기", 11_000, 16_000,  2_000, 22_000,  "HV Switchgear Distribution 기반"),
        ("지중개폐기", 13_000, 38_000,  2_500, 64_000,  "HV Switchgear Primary 기반"),
        ("배전선로",    7_000,  2_000,  1_500,  4_500,  "Poles 기반"),
    ]
    for r, row in enumerate(cof_ref, 2):
        for c, val in enumerate(row, 1):
            cell = ws3.cell(r, c, val)
            cell.alignment = Alignment(horizontal="center")
            if r % 2 == 0:
                cell.fill = GREEN_FILL

    # ── Sheet 4: 투자 파라미터 ───────────────────────────────────
    ws4 = wb.create_sheet("4_투자파라미터")
    for c, txt in enumerate(["자산군", "교체비용(만원)", "공사시간(h)", "투입인력(인일)",
                              "의무교체_나이기준(년)", "의무교체_상태기준"], 1):
        ws4.column_dimensions[get_column_letter(c)].width = 20
        header_style(ws4.cell(1, c), txt)

    invest_params = [
        ("주상변압기", 250,  8,  2, 30, "substantial"),
        ("지상변압기", 800, 20,  5, 30, "substantial"),
        ("가공개폐기", 350,  6,  2, 30, "substantial"),
        ("지중개폐기", 600, 14,  4, 30, "substantial"),
        ("배전선로",   450, 12,  3, 30, "substantial"),
    ]
    for r, row in enumerate(invest_params, 2):
        for c, val in enumerate(row, 1):
            cell = ws4.cell(r, c, val)
            cell.alignment = Alignment(horizontal="center")
            if r % 2 == 0:
                cell.fill = GREEN_FILL

    # ── Sheet 5: 위치계수 테이블 (CNAIM Table 22-24 요약) ────────
    ws5 = wb.create_sheet("5_위치계수테이블")
    ws5["A1"] = "CNAIM Table 22-24 위치계수 참조표 (자동 사용 - 수정 불필요)"
    ws5["A1"].font = Font(bold=True, color="FF0000")

    coast_header = ["해안거리 구분", "개폐기/변압기", "목주", "철주/콘크리트", "철탑"]
    for c, h in enumerate(coast_header, 1):
        header_style(ws5.cell(3, c), h, fill=SUB_FILL)
    coast_data = [
        ("≤1km",             1.35, 2.0, 1.5, 1.8),
        (">1~5km",           1.10, 1.5, 1.2, 1.45),
        (">5~10km",          1.05, 1.2, 1.1, 1.2),
        (">10~20km",         1.00, 1.0, 1.0, 1.0),
        (">20km",            0.90, 1.0, 1.0, 0.85),
        ("기본값(미상)",      1.00, 1.0, 1.0, 1.0),
    ]
    for r, row in enumerate(coast_data, 4):
        for c, val in enumerate(row, 1):
            ws5.cell(r, c, val).alignment = Alignment(horizontal="center")

    ws5["A11"] = "고도 구분"
    for c, h in enumerate(coast_header, 1):
        header_style(ws5.cell(12, c), h, fill=SUB_FILL)
    alt_data = [
        ("≤100m",   0.90, 0.95, 1.0, 0.90),
        (">100~200m",1.00, 1.00, 1.0, 1.00),
        (">200~300m",1.05, 1.05, 1.0, 1.15),
        (">300m",   1.10, 1.15, 1.0, 1.30),
        ("기본값",   1.00, 1.00, 1.0, 1.00),
    ]
    for r, row in enumerate(alt_data, 13):
        for c, val in enumerate(row, 1):
            ws5.cell(r, c, val).alignment = Alignment(horizontal="center")

    # ── Sheet 6: 상태평가 계수 (CNAIM Table 35~ 요약) ─────────────
    ws6 = wb.create_sheet("6_상태계수테이블")
    ws6["A1"] = "관측상태(Observed Condition) 계수 - CNAIM Table 35~ 기반"
    ws6["A1"].font = Font(bold=True)
    for c, h in enumerate(["관측상태", "계수(Factor)", "상한Cap", "하한Collar", "설명"], 1):
        ws6.column_dimensions[get_column_letter(c)].width = 18
        header_style(ws6.cell(2, c), h, fill=SUB_FILL)
    obs_data = [
        ("no_deterioration",  0.9, 10, 0.5, "이상 없음 (양호)"),
        ("superficial",       1.0, 10, 0.5, "경미한 열화"),
        ("some",              1.2, 10, 0.5, "일부 열화 (주의)"),
        ("substantial",       1.4, 10, 5.5, "심각한 열화 (불량)"),
        ("기본값(미상)",       1.0, 10, 0.5, "데이터 없을 때 기본값"),
    ]
    for r, row in enumerate(obs_data, 3):
        for c, val in enumerate(row, 1):
            ws6.cell(r, c, val).alignment = Alignment(horizontal="center")

    ws6["A10"] = "측정상태(Measured Condition) 계수"
    ws6["A10"].font = Font(bold=True)
    for c, h in enumerate(["측정상태", "계수(Factor)", "상한Cap", "하한Collar", "설명"], 1):
        header_style(ws6.cell(11, c), h, fill=SUB_FILL)
    meas_data = [
        ("good",     0.9, 10, 0.5, "양호"),
        ("normal",   1.0, 10, 0.5, "보통"),
        ("moderate", 1.2, 10, 0.5, "주의"),
        ("poor",     1.5, 10, 5.5, "불량"),
        ("기본값",   1.0, 10, 0.5, "데이터 없을 때 기본값"),
    ]
    for r, row in enumerate(meas_data, 12):
        for c, val in enumerate(row, 1):
            ws6.cell(r, c, val).alignment = Alignment(horizontal="center")

    wb.save(INPUT_PATH)
    print(f"[입력 템플릿 생성] → {INPUT_PATH}")


# ══════════════════════════════════════════════════════════════════
# 2. 입력 Excel 읽기
# ══════════════════════════════════════════════════════════════════

def load_parameters():
    """입력 Excel에서 파라미터 읽기"""
    xl = pd.ExcelFile(INPUT_PATH, engine="openpyxl")

    # 분석 설정
    df_cfg = pd.read_excel(xl, sheet_name="1_분석설정", header=0,
                           usecols=[0, 1], names=["항목", "값"])
    settings = dict(zip(df_cfg["항목"], df_cfg["값"]))

    # PoF 파라미터
    df_pof = pd.read_excel(xl, sheet_name="2_PoF파라미터", header=0)
    df_pof.columns = ["자산군", "K_pct", "C", "HSL", "NEL"]
    df_pof = df_pof.set_index("자산군")

    # CoF 기준값
    df_cof = pd.read_excel(xl, sheet_name="3_CoF기준값", header=0)
    df_cof.columns = ["자산군", "F", "S", "E", "NP", "_ref"]
    df_cof = df_cof.set_index("자산군")

    # 투자 파라미터
    df_inv = pd.read_excel(xl, sheet_name="4_투자파라미터", header=0)
    df_inv.columns = ["자산군", "replace_cost", "maint_hours", "labor_days",
                      "mand_age", "mand_cond"]
    df_inv = df_inv.set_index("자산군")

    return settings, df_pof, df_cof, df_inv


# ══════════════════════════════════════════════════════════════════
# 3. CNAIM 수식 모듈
# ══════════════════════════════════════════════════════════════════

H_NEW = 0.5      # 새 설비 건강지수 (CNAIM)
H_EOL = 5.5      # 설계수명 도달 시 건강지수 (CNAIM)


def location_factor_lookup(coast_km: float, altitude_m: float,
                            corrosion_cat: int, asset_col: str = "sw_tr") -> float:
    """
    위치계수 산출 - CNAIM Table 22, 23, 24 + EQ.13~17
    asset_col: 'sw_tr'=개폐기/변압기, 'pole_w'=목주, 'pole_sc'=철주/콘크리트
    """
    # Table 22: Distance from coast factor
    coast_map = {
        "sw_tr": [(1,1.35),(5,1.10),(10,1.05),(20,1.00),(9999,0.90)],
        "pole_w":[(1,2.00),(5,1.50),(10,1.20),(20,1.00),(9999,1.00)],
        "pole_sc":[(1,1.50),(5,1.20),(10,1.10),(20,1.00),(9999,1.00)],
    }
    cf = 1.0
    for km_lim, fac in coast_map.get(asset_col, coast_map["sw_tr"]):
        if coast_km <= km_lim:
            cf = fac
            break

    # Table 23: Altitude factor
    alt_map = {
        "sw_tr": [(100,0.90),(200,1.00),(300,1.05),(9999,1.10)],
        "pole_w":[(100,0.95),(200,1.00),(300,1.05),(9999,1.15)],
        "pole_sc":[(100,1.00),(200,1.00),(300,1.00),(9999,1.00)],
    }
    af = 1.0
    for m_lim, fac in alt_map.get(asset_col, alt_map["sw_tr"]):
        if altitude_m <= m_lim:
            af = fac
            break

    # Table 24: Corrosion factor
    corr_map = {
        "sw_tr": {1:0.90, 2:0.95, 3:1.00, 4:1.10, 5:1.25},
        "pole_w":{1:1.00, 2:1.00, 3:1.00, 4:1.00, 5:1.00},
        "pole_sc":{1:0.90,2:0.95,3:1.00,4:1.15,5:1.35},
    }
    rf = corr_map.get(asset_col, corr_map["sw_tr"]).get(corrosion_cat, 1.0)

    # EQ.13/14: outdoor location factor (INC=0.05 for sw_tr)
    INC = 0.05 if asset_col in ["sw_tr"] else 0.0
    max_f = max(cf, af, rf)
    if max_f > 1.0:
        lf = max_f + ((max_f - 1) * INC)
    else:
        lf = max_f
    return round(lf, 4)


def mmi_combine(f_obs: float, f_meas: float,
                div1: float = 1.5, div2: float = 1.5) -> float:
    """
    MMI (Maximum and Multiple Increment) 기법 - CNAIM Section 6.7.2
    두 Condition Factor 결합 (max 2 factors)
    """
    a, b = max(f_obs, f_meas), min(f_obs, f_meas)
    if a > 1.0:
        return a + ((b - 1.0) / div1) if b > 1.0 else a
    else:
        return b + ((a - 1.0) / div2)


def ageing_reduction_factor(chs: float) -> float:
    """노화감소계수 - CNAIM Table 216 (Figure 5)"""
    if chs < 2.0:
        return 1.0
    elif chs >= 5.5:
        return 1.5
    else:
        return 1.0 + (1.5 - 1.0) * (chs - 2.0) / (5.5 - 2.0)


def compute_pof(H: np.ndarray, K: float, C: float, HSL: int) -> np.ndarray:
    """
    PoF per annum - CNAIM EQ.3 (Taylor 3차)
    PoF = K × [1 + CH + (CH)²/2! + (CH)³/3!]
    H = max(Health Score, HSL)
    """
    h = np.maximum(H, HSL)
    return K * (1.0 + (C * h) + (C * h) ** 2 / 2.0 + (C * h) ** 3 / 6.0)


# ══════════════════════════════════════════════════════════════════
# 4. 합성 데이터 생성 (자산군별)
# ══════════════════════════════════════════════════════════════════

# 상태 테이블
OBS_STATES = {
    "no_deterioration": {"factor": 0.9, "cap": 10.0, "collar": 0.5, "label": "이상없음"},
    "superficial":      {"factor": 1.0, "cap": 10.0, "collar": 0.5, "label": "경미열화"},
    "some":             {"factor": 1.2, "cap": 10.0, "collar": 0.5, "label": "일부열화"},
    "substantial":      {"factor": 1.4, "cap": 10.0, "collar": 5.5, "label": "심각열화"},
}
MEAS_STATES = {
    "good":     {"factor": 0.9, "cap": 10.0, "collar": 0.5, "label": "양호"},
    "normal":   {"factor": 1.0, "cap": 10.0, "collar": 0.5, "label": "보통"},
    "moderate": {"factor": 1.2, "cap": 10.0, "collar": 0.5, "label": "주의"},
    "poor":     {"factor": 1.5, "cap": 10.0, "collar": 5.5, "label": "불량"},
}

# 자산군별 위치계수 컬럼 매핑
ASSET_LOC_COL = {
    "주상변압기": "sw_tr",
    "지상변압기": "sw_tr",
    "가공개폐기": "sw_tr",
    "지중개폐기": "sw_tr",
    "배전선로":   "pole_w",
}

# 기준 고객수 (CoF Network Performance 계수 산출용)
REF_CUSTOMERS = {
    "주상변압기": 40,
    "지상변압기": 80,
    "가공개폐기": 60,
    "지중개폐기": 120,
    "배전선로":   30,
}


def generate_single_asset_type(asset_type: str, n: int,
                                pof_row, cof_row, inv_row, rng: np.random.Generator,
                                discount_rate: float = 0.05) -> pd.DataFrame:
    """
    단일 자산군 n개 합성 데이터 생성 및 계산

    반환 컬럼:
      자산ID, 자산유형, 나이, 설계수명, 기대수명, 해안거리(km), 고도(m), 부식등급,
      위치계수, 운전계수, 부하율/운전횟수, 관측상태, 관측계수, 측정상태, 측정계수,
      건강지수계수, 초기건강지수, 현재건강지수, HI밴드, 현재PoF,
      고객수, 고객계수, 환경위치계수, 환경규모계수,
      CoF_재무, CoF_안전, CoF_환경, CoF_계통, CoF합계,
      현재위험도, β1, β2, 노화감소계수,
      교체전_위험도NPV, 교체후_위험도NPV, 투자가치_NPV,
      의무교체여부, 교체비용(만원), 공사시간(h), 투입인력(인일)
    """
    K   = pof_row["K_pct"] / 100.0
    C   = float(pof_row["C"])
    HSL = int(pof_row["HSL"])
    NEL = int(pof_row["NEL"])

    loc_col = ASSET_LOC_COL[asset_type]

    # ── 나이 (연속균등: 1~NEL+15, 가중치 부여해 실제 분포 모사)
    ages = rng.integers(1, NEL + 16, size=n)

    # ── 위치 속성
    coast_cats = rng.choice([0.5, 3, 8, 15, 30], size=n,
                             p=[0.05, 0.15, 0.20, 0.35, 0.25])
    altitude_cats = rng.choice([50, 150, 250, 400], size=n,
                                p=[0.40, 0.35, 0.15, 0.10])
    corrosion_cats = rng.choice([1, 2, 3, 4, 5], size=n,
                                 p=[0.10, 0.20, 0.45, 0.18, 0.07])
    loc_factors = np.array([
        location_factor_lookup(coast_cats[i], altitude_cats[i],
                               corrosion_cats[i], loc_col)
        for i in range(n)
    ])

    # ── 운전계수 (Duty Factor)
    if asset_type in ["주상변압기", "지상변압기"]:
        util_pct = rng.choice([35, 55, 80, 100, 120],
                               size=n, p=[0.15, 0.35, 0.30, 0.15, 0.05])
        duty_factors = np.where(util_pct <= 50,  0.90,
                       np.where(util_pct <= 70,  0.95,
                       np.where(util_pct <= 100, 1.00, 1.40)))
        duty_input = util_pct
        duty_label = "부하율(%)"
    elif asset_type == "가공개폐기":
        ops = rng.choice(["일반/저", "고(자동재폐로)"], size=n, p=[0.82, 0.18])
        duty_factors = np.where(ops == "고(자동재폐로)", 1.20, 1.00)
        duty_input = ops
        duty_label = "운전횟수구분"
    else:
        duty_factors = np.ones(n)
        duty_input = np.full(n, "해당없음", dtype=object)
        duty_label = "운전횟수구분"

    # ── 기대수명 / β1 (EQ.4, 5)
    expected_lives = NEL / (duty_factors * loc_factors)
    expected_lives = np.maximum(expected_lives, 1.0)
    beta1_arr = np.log(H_EOL / H_NEW) / expected_lives

    # ── Initial Health Score (EQ.6)
    init_hs = np.minimum(H_NEW * np.exp(beta1_arr * ages), H_EOL)

    # ── 상태 평가 (관측 + 측정)
    obs_keys  = list(OBS_STATES.keys())
    meas_keys = list(MEAS_STATES.keys())
    obs_idx   = rng.choice(len(obs_keys),  size=n, p=[0.22, 0.40, 0.27, 0.11])
    meas_idx  = rng.choice(len(meas_keys), size=n, p=[0.20, 0.45, 0.25, 0.10])

    obs_labels  = np.array([obs_keys[i]  for i in obs_idx],  dtype=object)
    meas_labels = np.array([meas_keys[i] for i in meas_idx], dtype=object)
    obs_factors  = np.array([OBS_STATES[k]["factor"]  for k in obs_labels])
    meas_factors = np.array([MEAS_STATES[k]["factor"] for k in meas_labels])
    obs_collars  = np.array([OBS_STATES[k]["collar"]  for k in obs_labels])
    meas_collars = np.array([MEAS_STATES[k]["collar"] for k in meas_labels])

    # ── 건강지수계수 (MMI, EQ.7~9)
    hs_factors = np.array([mmi_combine(obs_factors[i], meas_factors[i]) for i in range(n)])
    hs_caps    = np.minimum(
        np.array([OBS_STATES[k]["cap"] for k in obs_labels]),
        np.array([MEAS_STATES[k]["cap"] for k in meas_labels])
    )
    hs_collars = np.maximum(obs_collars, meas_collars)

    curr_hs = np.clip(init_hs * hs_factors, hs_collars, hs_caps)
    curr_hs = np.minimum(curr_hs, 10.0)

    # ── HI 밴드 (CNAIM Table 5)
    hi_bands = np.select(
        [curr_hs < 3.0, curr_hs < 5.0, curr_hs < 6.5, curr_hs < 8.5],
        ["HI1", "HI2", "HI3", "HI4"], default="HI5"
    )

    # ── 현재 PoF (EQ.3)
    pof_current = compute_pof(curr_hs, K, C, HSL)

    # ── CoF 계산 (Section 7)
    num_customers = np.clip(
        rng.lognormal(np.log(REF_CUSTOMERS[asset_type]), 0.60, size=n).astype(int),
        1, 600
    )
    cust_factors = np.minimum(num_customers / REF_CUSTOMERS[asset_type], 5.0)
    near_water   = rng.choice([True, False], size=n, p=[0.12, 0.88])
    size_cats    = rng.choice(["소", "중", "대"], size=n, p=[0.30, 0.50, 0.20])

    loc_env_f  = np.where(near_water, 1.30, 1.00)
    size_env_f = np.select([size_cats=="소", size_cats=="중"], [0.80, 1.00], default=1.30)

    cof_F  = np.full(n, cof_row["F"])
    cof_S  = np.full(n, cof_row["S"])
    cof_E  = cof_row["E"] * loc_env_f * size_env_f
    cof_NP = cof_row["NP"] * cust_factors
    cof_total = cof_F + cof_S + cof_E + cof_NP

    # ── 현재 위험도 (Risk = PoF × CoF)
    risk_current = pof_current * cof_total

    # ── β2 / 노화감소계수 (EQ.10, 11 + Table 216)
    age_safe = np.maximum(ages, 1)
    beta2_arr = np.where(
        curr_hs > H_NEW,
        np.minimum(np.log(np.maximum(curr_hs / H_NEW, 1.0001)) / age_safe,
                   2.0 * beta1_arr),
        beta1_arr
    )
    arf_arr = np.array([ageing_reduction_factor(h) for h in curr_hs])

    # ── 투자가치 NPV 계산 (벡터화)
    T = NEL  # 분석기간 = 설계수명

    t_vec      = np.arange(T, dtype=float)
    disc_vec   = 1.0 / (1.0 + discount_rate) ** t_vec  # (T,)

    # 교체 전 위험도 스트림: FHS(t) = CHS × exp((β2/ARF) × t)
    rates_pre  = (beta2_arr / arf_arr)[:, np.newaxis]      # (n, 1)
    fhs_pre    = curr_hs[:, np.newaxis] * np.exp(rates_pre * t_vec[np.newaxis, :])
    fhs_pre    = np.minimum(fhs_pre, 15.0)                   # cap at 15 (CNAIM EQ.12)
    pof_pre    = compute_pof(fhs_pre, K, C, HSL)             # (n, T)
    risk_pre   = pof_pre * cof_total[:, np.newaxis]           # (n, T)
    npv_pre    = np.sum(risk_pre * disc_vec[np.newaxis, :], axis=1)  # (n,)

    # 교체 후 위험도 스트림: 새 설비 0세 → NEL세 (동일 위치/운전 조건)
    beta1_new  = np.log(H_EOL / H_NEW) / expected_lives        # (n,)  각 설비 기대수명 기준
    fhs_post   = H_NEW * np.exp(beta1_new[:, np.newaxis] * t_vec[np.newaxis, :])
    fhs_post   = np.minimum(fhs_post, H_EOL)
    pof_post   = compute_pof(fhs_post, K, C, HSL)              # (n, T)
    risk_post  = pof_post * cof_total[:, np.newaxis]            # (n, T)
    npv_post   = np.sum(risk_post * disc_vec[np.newaxis, :], axis=1)  # (n,)

    invest_val = npv_pre - npv_post   # 투자가치 (양수 = 교체 유리)

    # ── 의무교체 여부
    mand_age   = int(inv_row["mand_age"])
    mand_cond  = str(inv_row["mand_cond"])
    mandatory  = ((ages >= mand_age) & (obs_labels == mand_cond)).astype(int)

    # ── 자산 ID 생성 (나중에 전체 통합 시 재번호 부여)
    type_code = {"주상변압기":"TR1","지상변압기":"TR2","가공개폐기":"SW1",
                 "지중개폐기":"SW2","배전선로":"DL1"}[asset_type]
    ids = [f"{type_code}-{i+1:05d}" for i in range(n)]

    # ── DataFrame 구성
    df = pd.DataFrame({
        "자산ID":          ids,
        "자산유형":        asset_type,
        "나이(년)":        ages,
        "설계수명(년)":    NEL,
        "기대수명(년)":    np.round(expected_lives, 1),
        "해안거리(km)":    coast_cats,
        "고도(m)":         altitude_cats,
        "부식등급":        corrosion_cats,
        "위치계수":        loc_factors,
        f"{duty_label}":   duty_input,
        "운전계수":        duty_factors,
        "관측상태":        obs_labels,
        "관측계수":        obs_factors,
        "측정상태":        meas_labels,
        "측정계수":        meas_factors,
        "건강지수계수":    np.round(hs_factors, 4),
        "초기건강지수":    np.round(init_hs, 4),
        "현재건강지수":    np.round(curr_hs, 4),
        "HI밴드":          hi_bands,
        "현재PoF":         np.round(pof_current, 7),
        "고객수":          num_customers,
        "고객수계수":      np.round(cust_factors, 3),
        "환경위치계수":    loc_env_f,
        "환경규모계수":    size_env_f,
        "CoF_재무(천원)":  np.round(cof_F, 0),
        "CoF_안전(천원)":  np.round(cof_S, 0),
        "CoF_환경(천원)":  np.round(cof_E, 0),
        "CoF_계통(천원)":  np.round(cof_NP, 0),
        "CoF합계(천원)":   np.round(cof_total, 0),
        "현재위험도(천원/년)": np.round(risk_current, 3),
        "β1초기노화율":    np.round(beta1_arr, 6),
        "β2예측노화율":    np.round(beta2_arr, 6),
        "노화감소계수":    np.round(arf_arr, 3),
        "교체전_위험도NPV(천원)": np.round(npv_pre, 2),
        "교체후_위험도NPV(천원)": np.round(npv_post, 2),
        "투자가치_NPV(천원)":     np.round(invest_val, 2),
        "의무교체여부":    mandatory,
        "교체비용(만원)":  int(inv_row["replace_cost"]),
        "공사시간(h)":     int(inv_row["maint_hours"]),
        "투입인력(인일)":  int(inv_row["labor_days"]),
    })

    return df


def generate_all(settings, df_pof, df_cof, df_inv) -> dict[str, pd.DataFrame]:
    rng = np.random.default_rng(int(settings["랜덤 시드"]))
    n   = int(settings["자산군별 생성 수량"])

    all_dfs = {}
    for asset_type in df_pof.index:
        print(f"  [{asset_type}] {n:,}개 생성 중...", end="", flush=True)
        t0 = datetime.now()
        discount_rate = float(settings["할인율 (%)"] / 100.0)
        df = generate_single_asset_type(
            asset_type, n,
            df_pof.loc[asset_type],
            df_cof.loc[asset_type],
            df_inv.loc[asset_type],
            rng,
            discount_rate=discount_rate
        )
        elapsed = (datetime.now() - t0).total_seconds()
        print(f" 완료 ({elapsed:.1f}s)")
        all_dfs[asset_type] = df

    return all_dfs


# ══════════════════════════════════════════════════════════════════
# 5. 요약 통계 생성
# ══════════════════════════════════════════════════════════════════

def build_summary(all_dfs: dict) -> pd.DataFrame:
    rows = []
    for atype, df in all_dfs.items():
        mand = df["의무교체여부"].sum()
        rows.append({
            "자산유형":           atype,
            "생성수량(기)":        len(df),
            "평균나이(년)":        df["나이(년)"].mean().round(1),
            "평균건강지수":        df["현재건강지수"].mean().round(3),
            "평균PoF":            df["현재PoF"].mean().round(6),
            "평균CoF(천원)":       df["CoF합계(천원)"].mean().round(0),
            "평균위험도(천원/년)": df["현재위험도(천원/년)"].mean().round(2),
            "평균투자가치(천원)":  df["투자가치_NPV(천원)"].mean().round(0),
            "의무교체수":          mand,
            "의무교체비용(만원)":  (df[df["의무교체여부"]==1]["교체비용(만원)"].sum()),
            "HI1수": (df["HI밴드"]=="HI1").sum(),
            "HI2수": (df["HI밴드"]=="HI2").sum(),
            "HI3수": (df["HI밴드"]=="HI3").sum(),
            "HI4수": (df["HI밴드"]=="HI4").sum(),
            "HI5수": (df["HI밴드"]=="HI5").sum(),
        })
    return pd.DataFrame(rows)


# ══════════════════════════════════════════════════════════════════
# 6. 출력 Excel 작성
# ══════════════════════════════════════════════════════════════════

HEADER_COLOR = "1F4E79"
ALT_COLOR    = "DEEAF1"
MAND_COLOR   = "FFC7CE"
HI5_COLOR    = "FF0000"
HI4_COLOR    = "FF9900"

def style_header_row(ws, row_idx: int, n_cols: int):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row_idx, c)
        cell.font  = Font(bold=True, color="FFFFFF", size=9)
        cell.fill  = PatternFill("solid", fgColor=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)

def auto_width(ws, df: pd.DataFrame, max_w: int = 25):
    for i, col in enumerate(df.columns, 1):
        max_len = max(len(str(col)), df[col].astype(str).str.len().max())
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, max_w)


def write_output_excel(all_dfs: dict, summary_df: pd.DataFrame,
                        settings: dict, df_pof, output_path: Path):
    print("\n출력 Excel 작성 중...")
    wb = openpyxl.Workbook()

    # ── Sheet 0: 분석 개요 ─────────────────────────────────────
    ws0 = wb.active
    ws0.title = "0_분석개요"
    ws0.column_dimensions["A"].width = 30
    ws0.column_dimensions["B"].width = 25
    ws0["A1"] = "AIP 배전설비 위험도 분석 결과"
    ws0["A1"].font = Font(bold=True, size=14, color=HEADER_COLOR)
    ws0["A3"] = "분석 일시";  ws0["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws0["A4"] = "총 분석 설비 수";  ws0["B4"] = sum(len(d) for d in all_dfs.values())
    ws0["A5"] = "할인율 (%)";  ws0["B5"] = settings["할인율 (%)"]
    ws0["A6"] = "자산군 수";   ws0["B6"] = len(all_dfs)
    ws0["A7"] = "CNAIM 버전"; ws0["B7"] = "v2.1 (April 2021)"
    ws0["A9"] = "자산군별 파라미터 (CNAIM Table 20, 21)"
    ws0["A9"].font = Font(bold=True)
    for c, h in enumerate(["자산군", "K값(%)", "C값", "HSL", "NEL(년)"], 1):
        ws0.cell(10, c).value = h
        ws0.cell(10, c).font = Font(bold=True, color="FFFFFF")
        ws0.cell(10, c).fill = PatternFill("solid", fgColor=HEADER_COLOR)
        ws0.cell(10, c).alignment = Alignment(horizontal="center")
    for r, (idx, row) in enumerate(df_pof.iterrows(), 11):
        ws0.cell(r, 1, idx).alignment = Alignment(horizontal="center")
        ws0.cell(r, 2, row["K_pct"]).alignment = Alignment(horizontal="center")
        ws0.cell(r, 3, row["C"]).alignment = Alignment(horizontal="center")
        ws0.cell(r, 4, int(row["HSL"])).alignment = Alignment(horizontal="center")
        ws0.cell(r, 5, int(row["NEL"])).alignment = Alignment(horizontal="center")

    # ── Sheet 1: 요약 통계 ─────────────────────────────────────
    ws1 = wb.create_sheet("1_요약통계")
    ws1.row_dimensions[1].height = 30
    for c, col in enumerate(summary_df.columns, 1):
        cell = ws1.cell(1, c, col)
        cell.font      = Font(bold=True, color="FFFFFF", size=9)
        cell.fill      = PatternFill("solid", fgColor=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws1.column_dimensions[get_column_letter(c)].width = 18

    for r, row in enumerate(summary_df.itertuples(index=False), 2):
        for c, val in enumerate(row, 1):
            cell = ws1.cell(r, c, val)
            cell.alignment = Alignment(horizontal="center")
            if r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ALT_COLOR)

    # ── Sheet 2~6: 자산군별 상세 데이터 ─────────────────────────
    sheet_names = {
        "주상변압기": "2_주상변압기",
        "지상변압기": "3_지상변압기",
        "가공개폐기": "4_가공개폐기",
        "지중개폐기": "5_지중개폐기",
        "배전선로":   "6_배전선로",
    }

    for asset_type, df in all_dfs.items():
        sname = sheet_names.get(asset_type, asset_type)
        ws = wb.create_sheet(sname)
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 30

        # 헤더 작성
        for c, col_name in enumerate(df.columns, 1):
            cell = ws.cell(1, c, col_name)
            cell.font      = Font(bold=True, color="FFFFFF", size=8)
            cell.fill      = PatternFill("solid", fgColor=HEADER_COLOR)
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            ws.column_dimensions[get_column_letter(c)].width = 14

        # 데이터 작성 (배치 처리로 속도 향상)
        print(f"  [{asset_type}] Excel 작성 중...", end="", flush=True)
        rows_to_write = df.values.tolist()
        for r_idx, row_data in enumerate(rows_to_write, 2):
            for c_idx, val in enumerate(row_data, 1):
                cell = ws.cell(r_idx, c_idx)
                # NaN/inf 처리
                if isinstance(val, float) and (np.isnan(val) or np.isinf(val)):
                    cell.value = None
                else:
                    cell.value = val
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 의무교체 행 강조 (빨간 배경)
            mand_col_idx = df.columns.get_loc("의무교체여부") + 1
            if row_data[mand_col_idx - 1] == 1:
                for c_idx in range(1, len(df.columns) + 1):
                    ws.cell(r_idx, c_idx).fill = PatternFill("solid", fgColor=MAND_COLOR)

        print(" 완료")

    # ── Sheet 7: 투자가치 전체 순위 (상위 500개) ──────────────
    ws7 = wb.create_sheet("7_투자가치순위")
    all_concat = pd.concat(all_dfs.values(), ignore_index=True)
    top_n = all_concat.nlargest(500, "투자가치_NPV(천원)")
    top_n.insert(0, "순위", range(1, len(top_n) + 1))

    cols_rank = ["순위", "자산ID", "자산유형", "나이(년)", "현재건강지수", "HI밴드",
                 "현재PoF", "CoF합계(천원)", "현재위험도(천원/년)",
                 "교체전_위험도NPV(천원)", "교체후_위험도NPV(천원)",
                 "투자가치_NPV(천원)", "의무교체여부",
                 "교체비용(만원)", "공사시간(h)", "투입인력(인일)"]
    top_n_out = top_n[[c for c in cols_rank if c in top_n.columns]]

    ws7.row_dimensions[1].height = 30
    for c, col_name in enumerate(top_n_out.columns, 1):
        cell = ws7.cell(1, c, col_name)
        cell.font  = Font(bold=True, color="FFFFFF", size=9)
        cell.fill  = PatternFill("solid", fgColor=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws7.column_dimensions[get_column_letter(c)].width = 18

    for r_idx, row in enumerate(top_n_out.itertuples(index=False), 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws7.cell(r_idx, c_idx, val)
            cell.alignment = Alignment(horizontal="center")
            if r_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ALT_COLOR)

    # ── Sheet 8: HI밴드별 위험도 통계 ─────────────────────────
    ws8 = wb.create_sheet("8_HI밴드별통계")
    hi_summary = (all_concat
                  .groupby(["자산유형", "HI밴드"])
                  .agg(
                      수량=("자산ID", "count"),
                      평균건강지수=("현재건강지수", "mean"),
                      평균PoF=("현재PoF", "mean"),
                      평균위험도=("현재위험도(천원/년)", "mean"),
                      평균투자가치=("투자가치_NPV(천원)", "mean"),
                  )
                  .round(4)
                  .reset_index())

    for c, h in enumerate(hi_summary.columns, 1):
        ws8.column_dimensions[get_column_letter(c)].width = 16
        cell = ws8.cell(1, c, h)
        cell.font  = Font(bold=True, color="FFFFFF")
        cell.fill  = PatternFill("solid", fgColor=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center")
    for r_idx, row in enumerate(hi_summary.itertuples(index=False), 2):
        for c_idx, val in enumerate(row, 1):
            ws8.cell(r_idx, c_idx, val).alignment = Alignment(horizontal="center")

    wb.save(output_path)
    print(f"\n[저장 완료] → {output_path}")


# ══════════════════════════════════════════════════════════════════
# 7. 메인 실행
# ══════════════════════════════════════════════════════════════════

def main():
    print("=" * 62)
    print("  AIP 배전설비 위험도 분석 시스템 v2.0")
    print("  CNAIM v2.1 기반 합성 데이터 생성 및 투자가치 산출")
    print("=" * 62)

    # ── 입력 템플릿 생성 (없으면)
    if not INPUT_PATH.exists():
        print("\n[입력 파일 없음] 템플릿 자동 생성 중...")
        create_input_template()
        print(f"  → {INPUT_PATH} 을 열어 파라미터를 확인/수정 후 재실행하세요.")
        return

    # ── 파라미터 읽기
    print(f"\n[파라미터 읽기] {INPUT_PATH}")
    settings, df_pof, df_cof, df_inv = load_parameters()
    discount_rate = settings["할인율 (%)"]
    n_per_type    = int(settings["자산군별 생성 수량"])
    print(f"  할인율: {discount_rate}%,  자산군별 생성수: {n_per_type:,}개")

    # ── 합성 데이터 생성 + 계산
    print(f"\n[데이터 생성 및 계산]")
    t_start = datetime.now()
    all_dfs  = generate_all(settings, df_pof, df_cof, df_inv)
    elapsed  = (datetime.now() - t_start).total_seconds()
    total_n  = sum(len(d) for d in all_dfs.values())
    print(f"  완료: 총 {total_n:,}개 ({elapsed:.1f}초)")

    # ── 요약 통계
    summary_df = build_summary(all_dfs)
    print("\n[요약 통계]")
    print(summary_df[["자산유형","생성수량(기)","평균건강지수","평균PoF",
                       "평균위험도(천원/년)","평균투자가치(천원)","의무교체수"]].to_string(index=False))

    # ── 출력 Excel 저장
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = DATA_DIR / f"output_results_{ts}.xlsx"
    write_output_excel(all_dfs, summary_df, settings, df_pof, output_path)

    print("\n[완료] 분석 종료")
    print("=" * 62)


if __name__ == "__main__":
    main()
