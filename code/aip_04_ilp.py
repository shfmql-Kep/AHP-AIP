"""
AIP Step 4 -- 투자 포트폴리오 최적화 (6방법 x 3예산 시나리오)
================================================================
입력:
  data/output_results_*.xlsx  (Step 2 결과, 최신 파일 자동 선택)
  code/survey_data.py         (Step 3 AHP 가중치 재산출)

출력:
  data/optimization_results_YYYYMMDD_HHMMSS.xlsx

[비교 구조]
  Baseline (기존 관행: 자산유형별 예산 분리):
    B1: 자산유형별 Risk Greedy  -> 합산
    B2: 자산유형별 NPV Greedy   -> 합산
    B3: 자산유형별 Risk-ILP     -> 합산
    B4: 자산유형별 NPV-ILP      -> 합산

  Proposed (제안: 전체 통합 MCDM 최적화):
    P1: 통합 AHP+ILP
    P2: 통합 AHP+GA

[민감도 분석]
  예산 시나리오: 3%, 5%, 7% (전체 교체비용 대비 선택적 교체 예산)

[저압설비 제외]
  LV배전반_UGB 제외 (고압 설비만 대상)
"""

import sys
import numpy as np
import pandas as pd
from pathlib import Path
from datetime import datetime
import warnings
warnings.filterwarnings("ignore")

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"
CODE_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(CODE_DIR))

from scipy.optimize import milp, LinearConstraint, Bounds

# ======================================================================
# 0. 설정 상수
# ======================================================================

# 저압설비 제외 목록
LV_ASSETS = {'LV배전반_UGB'}

# 민감도 분석 예산 시나리오 (전체 교체비용 대비 선택적 교체 예산 비율)
BUDGET_SCENARIOS = [0.03, 0.05, 0.07]

# MTTR (자산유형별 평균 수리 복구 시간, 시간 단위)
MTTR_HOURS = {
    '주상변압기':     3.0,
    '지상변압기':     5.0,
    '지중변압기':     8.0,
    '가공개폐기':     2.0,
    '지중개폐기_RMU': 6.0,
    '특고압차단기':   6.0,
    '가공배전선로':   3.0,
    '지중케이블':     8.0,
    '목주':           4.0,
    '콘크리트주':     4.0,
    '철주':           6.0,
}
MTTR_DEFAULT = 4.0

# GA 파라미터
GA_POP_SIZE    = 300
GA_GENERATIONS = 150
GA_CROSSOVER   = 0.8
GA_MUTATION    = 0.02
GA_TOURNAMENT  = 5
GA_SEED        = 42

# 방법 순서 (출력/보고서 일관성)
METHODS_ORDER = [
    'B1 유형별Risk-Greedy',
    'B2 유형별NPV-Greedy',
    'B3 유형별Risk-ILP',
    'B4 유형별NPV-ILP',
    'P1 통합AHP-ILP',
    'P2 통합AHP-GA',
]


# ======================================================================
# 1. 데이터 로드
# ======================================================================

def load_latest_results() -> pd.DataFrame:
    """가장 최근 output_results_*.xlsx 로드 + 저압설비 제외"""
    import openpyxl as _oxl
    files = sorted(DATA_DIR.glob("output_results_*.xlsx"), reverse=True)
    if not files:
        raise FileNotFoundError(f"output_results_*.xlsx 없음: {DATA_DIR}")
    path = files[0]
    print(f"  분석 결과: {path.name}")
    wb = _oxl.load_workbook(path, read_only=True)
    sheet_name = next((s for s in wb.sheetnames if '전체결과' in s), wb.sheetnames[0])
    wb.close()
    df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
    required = ["자산유형", "현재PoF", "CoF_재무_천원", "CoF_안전_천원",
                "CoF_환경_천원", "CoF_계통_천원", "현재위험도_천원_년",
                "투자가치NPV_천원", "교체비용_천원", "의무교체여부"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise KeyError(f"필수 컬럼 누락: {missing}")
    before = len(df)
    df = df[~df['자산유형'].isin(LV_ASSETS)].copy().reset_index(drop=True)
    print(f"  저압설비 제외: {before - len(df)}기 제거 -> 잔존 {len(df):,}기")
    return df


def get_ahp_weights() -> dict:
    """AHP 가중치 산출 (Classical AHP, 20명 전문가)"""
    from aip_03_ahp import run_pipeline
    from survey_data import SURVEY_20
    print("  AHP 가중치 산출 중 (Classical AHP, 20명 전문가)...")
    results = run_pipeline(SURVEY_20)
    gw = results['classical_ahp']['global_weights']
    print(f"  가중치: { {k: round(v, 4) for k, v in gw.items()} }")
    return gw


# ======================================================================
# 2. 점수 계산 및 혼합가치 산출
# ======================================================================

def _minmax(series: pd.Series) -> pd.Series:
    s = series.clip(lower=0)
    rng = s.max() - s.min()
    if rng < 1e-10:
        return pd.Series(np.zeros(len(s)), index=s.index)
    return (s - s.min()) / rng


def compute_scores(df: pd.DataFrame) -> pd.DataFrame:
    """7개 하위기준 점수 Min-Max 정규화"""
    out = df.copy()
    out['_mttr'] = out['자산유형'].map(MTTR_HOURS).fillna(MTTR_DEFAULT)
    cost = out['교체비용_천원'].replace(0, np.nan)

    out['score_NPV']   = _minmax(out['투자가치NPV_천원'].clip(lower=0))
    bcr = (out['투자가치NPV_천원'] / cost.fillna(1)).clip(lower=0)
    out['score_BCR']   = _minmax(bcr)
    out['score_SAIDI'] = _minmax(out['현재PoF'] * out['_mttr'] * out['CoF_계통_천원'])
    out['score_ENF']   = _minmax(out['현재PoF'] * out['CoF_계통_천원'])
    out['score_재무R']  = _minmax(out['현재PoF'] * out['CoF_재무_천원'])
    out['score_안전R']  = _minmax(out['현재PoF'] * out['CoF_안전_천원'])
    out['score_환경R']  = _minmax(out['현재PoF'] * out['CoF_환경_천원'])
    out.drop(columns=['_mttr'], inplace=True)
    return out


def compute_mixed_value(df: pd.DataFrame, weights: dict) -> pd.Series:
    """혼합가치 V_j = sum w_i * score_{j,i}"""
    score_cols = {
        'C1-1_NPV':       'score_NPV',
        'C1-2_BCR':       'score_BCR',
        'C2-1_SAIDI저감': 'score_SAIDI',
        'C2-2_ENF저감':   'score_ENF',
        'C3-1_재무Risk':  'score_재무R',
        'C3-2_안전Risk':  'score_안전R',
        'C3-3_환경Risk':  'score_환경R',
    }
    V = pd.Series(np.zeros(len(df)), index=df.index)
    for wkey, scol in score_cols.items():
        w = weights.get(wkey, 0.0)
        if scol in df.columns:
            V += w * df[scol]
    return V


# ======================================================================
# 3. 예산 분배 (비용 비례 동적 계산)
# ======================================================================

def calc_type_budgets(df: pd.DataFrame, opt_budget: float,
                      mandatory_mask: pd.Series) -> dict:
    """
    자산유형별 선택적 교체 예산 배분.
    배분 기준: 비의무 자산의 교체비용 비율 (비용 비례).
    """
    opt_df = df[~mandatory_mask & (df['교체비용_천원'] > 0)]
    total_opt_cost = opt_df['교체비용_천원'].sum()
    budgets = {}
    for atype in opt_df['자산유형'].unique():
        type_cost = opt_df.loc[opt_df['자산유형'] == atype, '교체비용_천원'].sum()
        ratio = type_cost / total_opt_cost if total_opt_cost > 0 else 0.0
        budgets[atype] = opt_budget * ratio
    return budgets


# ======================================================================
# 4. Baseline: 자산유형별 독립 최적화
# ======================================================================

def solve_per_type(df: pd.DataFrame,
                   priority_col: str,
                   opt_budget: float,
                   mandatory_mask: pd.Series,
                   use_ilp: bool = False,
                   value_col: str = None) -> pd.Series:
    """
    자산유형별 독립 최적화 (기존 관행 시뮬레이션).
    의무교체 자산은 별도 고정 포함.
    선택적 교체 예산을 비용 비례로 자산유형별 배분.
    """
    selected = mandatory_mask.copy()
    type_budgets = calc_type_budgets(df, opt_budget, mandatory_mask)

    for atype, type_budget in type_budgets.items():
        type_mask = (df['자산유형'] == atype) & ~mandatory_mask & (df['교체비용_천원'] > 0)
        if not type_mask.any() or type_budget <= 0:
            continue

        sub_df   = df[type_mask]
        sub_cost = sub_df['교체비용_천원'].values.astype(float)
        col      = value_col if value_col else priority_col
        sub_val  = sub_df[col].values.astype(float)
        n        = len(sub_df)

        if use_ilp:
            c_lp   = -sub_val
            constr = LinearConstraint(sub_cost[np.newaxis, :], lb=-np.inf, ub=type_budget)
            result = milp(c_lp, constraints=constr, integrality=np.ones(n),
                          bounds=Bounds(0.0, 1.0), options={'disp': False})
            if result.success:
                chosen = result.x.round().astype(int)
                for i, idx in enumerate(sub_df.index):
                    if chosen[i] == 1:
                        selected.loc[idx] = True
                continue
            # ILP 실패 시 Greedy fallback

        # Greedy (use_ilp=False 또는 ILP fallback)
        order = np.argsort(-sub_val)
        spent = 0.0
        for i in order:
            if spent + sub_cost[i] <= type_budget:
                selected.loc[sub_df.index[i]] = True
                spent += sub_cost[i]

    return selected


# ======================================================================
# 5. Proposed: 전체 통합 ILP / GA
# ======================================================================

def solve_ilp(df: pd.DataFrame, V: pd.Series,
              budget: float, mandatory_mask: pd.Series,
              method_label: str = "AHP+ILP",
              value_col: str = None) -> pd.Series:
    """전체 통합 0-1 ILP: maximize sum V_j * x_j"""
    mand_idx   = df.index[mandatory_mask]
    opt_mask   = ~mandatory_mask & (df['교체비용_천원'] > 0)
    opt_idx    = df.index[opt_mask]
    n          = opt_mask.sum()
    if n == 0:
        return mandatory_mask.copy()

    mand_cost  = df.loc[mand_idx, '교체비용_천원'].sum()
    rem_budget = budget - mand_cost
    if rem_budget <= 0:
        print(f"    [{method_label}] 의무교체 비용이 예산 초과 - 의무교체만 반환")
        return mandatory_mask.copy()

    val  = df.loc[opt_idx, value_col].values if value_col else V.loc[opt_idx].values
    cost = df.loc[opt_idx, '교체비용_천원'].values.astype(float)
    constr = LinearConstraint(cost[np.newaxis, :], lb=-np.inf, ub=rem_budget)
    result = milp(-val.astype(float), constraints=constr,
                  integrality=np.ones(n), bounds=Bounds(0.0, 1.0),
                  options={'disp': False, 'time_limit': 60.0})

    selected = mandatory_mask.copy()
    if result.success:
        chosen = result.x.round().astype(int)
        for i, idx in enumerate(opt_idx):
            if chosen[i] == 1:
                selected.loc[idx] = True
    else:
        print(f"    [{method_label}] ILP 미수렴 ({result.message}) - Greedy 대체")
        ratio = val / np.maximum(cost, 1.0)
        order = np.argsort(-ratio)
        spent = 0.0
        for i in order:
            if spent + cost[i] <= rem_budget:
                selected.loc[opt_idx[i]] = True
                spent += cost[i]
    return selected


def solve_ga(df: pd.DataFrame, V: pd.Series,
             budget: float, mandatory_mask: pd.Series) -> pd.Series:
    """전체 통합 GA: 대규모 확장성 발견적 해법"""
    rng        = np.random.default_rng(GA_SEED)
    mand_idx   = df.index[mandatory_mask]
    opt_mask   = ~mandatory_mask & (df['교체비용_천원'] > 0)
    opt_idx    = df.index[opt_mask]
    n          = opt_mask.sum()
    mand_cost  = df.loc[mand_idx, '교체비용_천원'].sum()
    rem_budget = budget - mand_cost
    val  = V.loc[opt_idx].values.astype(float)
    cost = df.loc[opt_idx, '교체비용_천원'].values.astype(float)

    if rem_budget <= 0 or n == 0:
        return mandatory_mask.copy()

    pop = np.zeros((GA_POP_SIZE, n), dtype=np.int8)
    for k in range(GA_POP_SIZE):
        order = rng.permutation(n)
        spent = 0.0
        for i in order:
            if spent + cost[i] <= rem_budget:
                pop[k, i] = 1
                spent += cost[i]

    def fitness(chroms):
        values  = chroms @ val
        costs   = chroms @ cost
        penalty = np.maximum(costs - rem_budget, 0) * 1e6
        return values - penalty

    def repair(chrom):
        if chrom @ cost > rem_budget:
            chrom = chrom.copy()
            ratio = np.where(chrom == 1, val / np.maximum(cost, 1.0), np.inf)
            for i in np.argsort(ratio):
                if chrom[i] == 1:
                    chrom[i] = 0
                    if chrom @ cost <= rem_budget:
                        break
        return chrom

    best_fitness = -np.inf
    best_chrom   = pop[0].copy()

    for gen in range(GA_GENERATIONS):
        fit = fitness(pop)
        elite_idx = np.argmax(fit)
        if fit[elite_idx] > best_fitness:
            best_fitness = fit[elite_idx]
            best_chrom   = pop[elite_idx].copy()
        new_pop = [best_chrom.copy()]
        while len(new_pop) < GA_POP_SIZE:
            def tournament():
                idx = rng.integers(0, GA_POP_SIZE, GA_TOURNAMENT)
                return pop[idx[np.argmax(fit[idx])]].copy()
            p1, p2 = tournament(), tournament()
            if rng.random() < GA_CROSSOVER:
                pt = rng.integers(1, n)
                c1 = np.concatenate([p1[:pt], p2[pt:]])
                c2 = np.concatenate([p2[:pt], p1[pt:]])
            else:
                c1, c2 = p1.copy(), p2.copy()
            for ch in (c1, c2):
                flip = rng.random(n) < GA_MUTATION
                ch[flip] ^= 1
            new_pop.append(repair(c1))
            if len(new_pop) < GA_POP_SIZE:
                new_pop.append(repair(c2))
        pop = np.array(new_pop[:GA_POP_SIZE], dtype=np.int8)
        if (gen + 1) % 30 == 0:
            print(f"    GA 세대 {gen+1}/{GA_GENERATIONS}: 최적 혼합가치={best_fitness:.4f}")

    selected = mandatory_mask.copy()
    for i, idx in enumerate(opt_idx):
        if best_chrom[i] == 1:
            selected.loc[idx] = True
    return selected


# ======================================================================
# 6. KPI 계산
# ======================================================================

def compute_kpi(df: pd.DataFrame, selected: pd.Series,
                V: pd.Series, label: str) -> dict:
    sub      = df[selected]
    tot_npv  = sub['투자가치NPV_천원'].sum()
    tot_cost = sub['교체비용_천원'].sum()
    tot_risk = sub['현재위험도_천원_년'].sum()
    tot_mv   = V[selected].sum()
    bcr      = tot_npv / max(tot_cost, 1.0)
    mttr     = sub['자산유형'].map(MTTR_HOURS).fillna(MTTR_DEFAULT)
    saidi    = (sub['현재PoF'] * mttr * sub['CoF_계통_천원']).sum()
    return {
        '방법':                label,
        '선택수(기)':          int(selected.sum()),
        '의무교체(기)':        int(sub['의무교체여부'].sum()),
        '총NPV(천원)':         round(tot_npv),
        '총비용(천원)':        round(tot_cost),
        'BCR':                  round(bcr, 3),
        '총Risk저감(천원/년)': round(tot_risk),
        'SAIDI기여합':         round(saidi, 2),
        '혼합가치합':          round(tot_mv, 4),
    }


def compute_type_kpi(df: pd.DataFrame, selected: pd.Series,
                     V: pd.Series) -> pd.DataFrame:
    """자산유형별 선택 분해"""
    rows = []
    for atype in sorted(df['자산유형'].unique()):
        mask  = selected & (df['자산유형'] == atype)
        total = int((df['자산유형'] == atype).sum())
        sub   = df[mask]
        rows.append({
            '자산유형':    atype,
            '전체수(기)':  total,
            '선택수(기)':  int(mask.sum()),
            '선택률(%)':   round(mask.sum() / max(total, 1) * 100, 1),
            '총NPV(천원)': round(sub['투자가치NPV_천원'].sum()),
            '총비용(천원)':round(sub['교체비용_천원'].sum()),
            '혼합가치합':  round(V[mask].sum(), 4),
        })
    return pd.DataFrame(rows)


# ======================================================================
# 7. 보고서 저장
# ======================================================================

def save_report(scenario_results: list, df: pd.DataFrame,
                V: pd.Series, out_path: Path):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb   = openpyxl.Workbook()
    CHDR = "1F497D"; CGOOD = "E2EFDA"; CWARN = "FFEEBA"
    CBST = "C6EFCE"; CALT  = "F5F5F5"
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'),  bottom=Side(style='thin'))

    def hcell(ws, r, c, val, bg=CHDR, fc="FFFFFF", bold=True):
        cell = ws.cell(row=r, column=c, value=val)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=bold, color=fc, size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin
        return cell

    def dcell(ws, r, c, val, fmt=None, bg=None, bold=False):
        cell = ws.cell(row=r, column=c, value=val)
        if fmt:  cell.number_format = fmt
        if bg:   cell.fill = PatternFill("solid", fgColor=bg)
        if bold: cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin
        return cell

    KPI_DISPLAY = [
        ('선택수(기)',          '선택 설비 수 (기)',       '#,##0'),
        ('총NPV(천원)',         '총 투자가치 NPV (천원)',  '#,##0'),
        ('총비용(천원)',        '총 투자비용 (천원)',      '#,##0'),
        ('BCR',                'BCR',                    '0.000'),
        ('총Risk저감(천원/년)', 'Risk 저감량 (천원/년)',   '#,##0'),
        ('SAIDI기여합',        'SAIDI 개선 기대량',       '#,##0'),
        ('혼합가치합',         '혼합가치 합계 (V)',       '0.0000'),
    ]

    # -- 시트 1: 민감도 시나리오 통합 요약 --------------------------------
    ws = wb.active
    ws.title = "민감도_시나리오비교"
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 22
    for ci in range(3, 3 + len(METHODS_ORDER)):
        ws.column_dimensions[get_column_letter(ci)].width = 18

    row = 1
    for sc in scenario_results:
        ratio    = sc['ratio']
        kpi_rows = sc['kpi_rows']
        kpi_df   = pd.DataFrame(kpi_rows)

        # 시나리오 헤더
        hcell(ws, row, 1, f"{ratio*100:.0f}%", bg="2F5496")
        hcell(ws, row, 2, "KPI 항목", bg="2F5496")
        for ci, m in enumerate(METHODS_ORDER, start=3):
            hcell(ws, row, ci, m, bg="2F5496")
        row += 1

        # 구분 행 (Baseline / 관행개선 / Proposed)
        hcell(ws, row, 1, "", bg="D9D9D9")
        hcell(ws, row, 2, "", bg="D9D9D9")
        for ci, label, bg in [
            (3, "< Baseline (관행) >",  "FFF2CC"),
            (4, "< Baseline (관행) >",  "FFF2CC"),
            (5, "< 관행 개선 >",         "DDEBF7"),
            (6, "< 관행 개선 >",         "DDEBF7"),
            (7, "< Proposed >",         "E2EFDA"),
            (8, "< Proposed >",         "E2EFDA"),
        ]:
            hcell(ws, row, ci, label, bg=bg, fc="000000", bold=False)
        row += 1

        for key, label, fmt in KPI_DISPLAY:
            bgrow = CALT if row % 2 == 0 else "FFFFFF"
            hcell(ws, row, 1, "", bg="D9D9D9")
            hcell(ws, row, 2, label, bg="4472C4")
            vals = []
            for m in METHODS_ORDER:
                mrows = [r for r in kpi_rows if r['방법'] == m]
                vals.append(mrows[0][key] if mrows else 0)
            try:
                best_idx = np.argmin(vals) if key == '총비용(천원)' else np.argmax(vals)
            except Exception:
                best_idx = None
            for ci, v in enumerate(vals, start=3):
                is_best = (ci - 3 == best_idx)
                dcell(ws, row, ci, v, fmt=fmt,
                      bg=CBST if is_best else bgrow, bold=is_best)
            row += 1
        row += 1  # 시나리오 간 공백

    # -- 시트 2~4: 시나리오별 상세 KPI ------------------------------------
    for sc in scenario_results:
        ratio    = sc['ratio']
        kpi_rows = sc['kpi_rows']
        ws_s = wb.create_sheet(f"KPI_{ratio*100:.0f}pct")
        ws_s.column_dimensions['A'].width = 24
        for ci in range(2, 2 + len(METHODS_ORDER)):
            ws_s.column_dimensions[get_column_letter(ci)].width = 18

        hcell(ws_s, 1, 1, "KPI / 방법")
        for ci, m in enumerate(METHODS_ORDER, start=2):
            hcell(ws_s, 1, ci, m)

        for ri, (key, label, fmt) in enumerate(KPI_DISPLAY, start=2):
            bg = CALT if ri % 2 == 0 else "FFFFFF"
            hcell(ws_s, ri, 1, label, bg="4472C4")
            vals = []
            for m in METHODS_ORDER:
                mrows = [r for r in kpi_rows if r['방법'] == m]
                vals.append(mrows[0][key] if mrows else 0)
            try:
                best_idx = np.argmin(vals) if key == '총비용(천원)' else np.argmax(vals)
            except Exception:
                best_idx = None
            for ci, v in enumerate(vals, start=2):
                is_best = (ci - 2 == best_idx)
                dcell(ws_s, ri, ci, v, fmt=fmt,
                      bg=CBST if is_best else bg, bold=is_best)

        # 향상률 행
        ri = len(KPI_DISPLAY) + 3
        hcell(ws_s, ri, 1, "B4 대비 P1 향상률 (%)", bg="4472C4")
        b4_rows = [r for r in kpi_rows if r['방법'] == 'B4 유형별NPV-ILP']
        p1_rows = [r for r in kpi_rows if r['방법'] == 'P1 통합AHP-ILP']
        if b4_rows and p1_rows:
            for ci, (key, _, fmt) in enumerate(KPI_DISPLAY, start=2):
                vB = b4_rows[0][key]
                vP = p1_rows[0][key]
                if abs(vB) > 1e-9 and key != '총비용(천원)':
                    pct = (vP - vB) / abs(vB) * 100
                    dcell(ws_s, ri, ci, round(pct, 2), fmt='0.00"%"',
                          bg=CGOOD if pct >= 0 else CWARN)
                else:
                    dcell(ws_s, ri, ci, "-")

    # -- 시트 5: 자산유형별 선택 분포 (5% 시나리오) -----------------------
    sc5 = next((s for s in scenario_results if abs(s['ratio'] - 0.05) < 1e-6),
               scenario_results[-1])
    ws4 = wb.create_sheet("자산유형별_선택분포")
    ws4.column_dimensions['A'].width = 20
    ws4.column_dimensions['B'].width = 10
    for ci in range(3, 3 + len(METHODS_ORDER)):
        ws4.column_dimensions[get_column_letter(ci)].width = 16

    hcell(ws4, 1, 1, "자산유형")
    hcell(ws4, 1, 2, "전체(기)")
    for ci, m in enumerate(METHODS_ORDER, start=3):
        hcell(ws4, 1, ci, m)

    for ri, atype in enumerate(sorted(df['자산유형'].unique()), start=2):
        bg    = CALT if ri % 2 == 0 else "FFFFFF"
        total = int((df['자산유형'] == atype).sum())
        dcell(ws4, ri, 1, atype, bg=bg)
        dcell(ws4, ri, 2, total, fmt='#,##0', bg=bg)
        for ci, m in enumerate(METHODS_ORDER, start=3):
            sel = sc5['selections'].get(m)
            cnt = int((df[sel]['자산유형'] == atype).sum()) if sel is not None else 0
            dcell(ws4, ri, ci, cnt, fmt='#,##0', bg=bg)

    # -- 시트 6: P1 선택목록 상세 (5% 시나리오) ---------------------------
    ws5 = wb.create_sheet("P1_AHP_ILP_선택목록")
    for col, w in zip(['A','B','C','D','E','F','G','H','I'],
                      [16, 10, 10, 12, 14, 10, 14, 10, 10]):
        ws5.column_dimensions[col].width = w
    hdrs = ["자산유형", "HI밴드", "나이(년)", "교체비용(천원)",
            "투자가치NPV(천원)", "BCR", "현재위험도(천원/년)", "혼합가치V", "의무교체"]
    for ci, h in enumerate(hdrs, start=1):
        hcell(ws5, 1, ci, h)

    sel_p1 = sc5['selections'].get('P1 통합AHP-ILP', pd.Series(dtype=bool))
    sub_p1 = df[sel_p1].copy()
    sub_p1['BCR']   = (sub_p1['투자가치NPV_천원'] /
                       sub_p1['교체비용_천원'].replace(0, np.nan)).fillna(0)
    sub_p1['혼합V'] = V[sel_p1].values
    sub_p1 = sub_p1.sort_values('혼합V', ascending=False)

    for ri, (_, row) in enumerate(sub_p1.iterrows(), start=2):
        bg = "FFC7CE" if row['의무교체여부'] == 1 else (CALT if ri % 2 == 0 else "FFFFFF")
        dcell(ws5, ri, 1, row['자산유형'],             bg=bg)
        dcell(ws5, ri, 2, row['HI밴드'],               bg=bg)
        dcell(ws5, ri, 3, int(row.get('나이_년', 0)),  bg=bg)
        dcell(ws5, ri, 4, int(row['교체비용_천원']),   fmt='#,##0', bg=bg)
        dcell(ws5, ri, 5, int(row['투자가치NPV_천원']),fmt='#,##0', bg=bg)
        dcell(ws5, ri, 6, round(row['BCR'], 3),        fmt='0.000', bg=bg)
        dcell(ws5, ri, 7, int(row['현재위험도_천원_년']), fmt='#,##0', bg=bg)
        dcell(ws5, ri, 8, round(row['혼합V'], 4),      fmt='0.0000', bg=bg)
        dcell(ws5, ri, 9, "O" if row['의무교체여부'] == 1 else "", bg=bg)

    wb.save(out_path)
    print(f"\n[보고서 저장] {out_path}")


# ======================================================================
# 8. 진입점
# ======================================================================

if __name__ == "__main__":
    print("=" * 70)
    print("  AIP Step 4 -- 투자 포트폴리오 최적화 (6방법 x 3예산 시나리오)")
    print("=" * 70)

    print("\n[데이터 로드]")
    df = load_latest_results()
    print(f"  전체 자산: {len(df):,}기  |  자산유형: {df['자산유형'].nunique()}종")

    print("\n[AHP 가중치]")
    weights = get_ahp_weights()

    print("\n[점수 계산]")
    df = compute_scores(df)
    V  = compute_mixed_value(df, weights)
    df['혼합가치_V'] = V
    print(f"  혼합가치 범위: [{V.min():.4f}, {V.max():.4f}]")

    mandatory_mask = df['의무교체여부'].astype(bool)
    total_cost     = df['교체비용_천원'].sum()
    mand_cost      = df.loc[mandatory_mask, '교체비용_천원'].sum()

    print(f"\n[기본 정보]")
    print(f"  전체 교체비용   : {total_cost:,.0f} 천원")
    print(f"  의무교체 대상   : {mandatory_mask.sum():,}기 ({mand_cost:,.0f} 천원, 별도 집행)")

    # 자산유형별 비용 비례 배분 비율 출력
    type_cost_df = (df[~mandatory_mask & (df['교체비용_천원'] > 0)]
                    .groupby('자산유형')['교체비용_천원'].sum())
    total_opt_cost = type_cost_df.sum()
    print(f"\n[예산 배분 기준 (비용 비례)]")
    for atype, cost in type_cost_df.sort_values(ascending=False).items():
        print(f"  {atype:16s}: {cost/total_opt_cost*100:.1f}%")

    scenario_results = []

    for ratio in BUDGET_SCENARIOS:
        opt_budget = total_cost * ratio
        budget     = mand_cost + opt_budget

        print(f"\n{'='*60}")
        print(f"  [시나리오: 선택 예산 {ratio*100:.0f}%]  "
              f"선택 예산={opt_budget:,.0f} 천원  |  총={budget:,.0f} 천원")
        print(f"{'='*60}")

        kpi_rows   = []
        selections = {}

        # B1: 자산유형별 Risk Greedy
        print(f"\n[B1] 자산유형별 Risk Greedy")
        sel = solve_per_type(df, '현재위험도_천원_년', opt_budget, mandatory_mask,
                             use_ilp=False)
        selections['B1 유형별Risk-Greedy'] = sel
        kpi_rows.append(compute_kpi(df, sel, V, 'B1 유형별Risk-Greedy'))
        print(f"  선택: {sel.sum():,}기  NPV: {df.loc[sel,'투자가치NPV_천원'].sum():,.0f} 천원")

        # B2: 자산유형별 NPV Greedy
        print(f"\n[B2] 자산유형별 NPV Greedy")
        sel = solve_per_type(df, '투자가치NPV_천원', opt_budget, mandatory_mask,
                             use_ilp=False)
        selections['B2 유형별NPV-Greedy'] = sel
        kpi_rows.append(compute_kpi(df, sel, V, 'B2 유형별NPV-Greedy'))
        print(f"  선택: {sel.sum():,}기  NPV: {df.loc[sel,'투자가치NPV_천원'].sum():,.0f} 천원")

        # B3: 자산유형별 Risk-ILP
        print(f"\n[B3] 자산유형별 Risk-ILP")
        sel = solve_per_type(df, '현재위험도_천원_년', opt_budget, mandatory_mask,
                             use_ilp=True, value_col='현재위험도_천원_년')
        selections['B3 유형별Risk-ILP'] = sel
        kpi_rows.append(compute_kpi(df, sel, V, 'B3 유형별Risk-ILP'))
        print(f"  선택: {sel.sum():,}기  NPV: {df.loc[sel,'투자가치NPV_천원'].sum():,.0f} 천원")

        # B4: 자산유형별 NPV-ILP
        print(f"\n[B4] 자산유형별 NPV-ILP")
        sel = solve_per_type(df, '투자가치NPV_천원', opt_budget, mandatory_mask,
                             use_ilp=True, value_col='투자가치NPV_천원')
        selections['B4 유형별NPV-ILP'] = sel
        kpi_rows.append(compute_kpi(df, sel, V, 'B4 유형별NPV-ILP'))
        print(f"  선택: {sel.sum():,}기  NPV: {df.loc[sel,'투자가치NPV_천원'].sum():,.0f} 천원")

        # P1: 통합 AHP+ILP
        print(f"\n[P1] 통합 AHP+ILP")
        sel = solve_ilp(df, V, budget, mandatory_mask, "AHP+ILP")
        selections['P1 통합AHP-ILP'] = sel
        kpi_rows.append(compute_kpi(df, sel, V, 'P1 통합AHP-ILP'))
        print(f"  선택: {sel.sum():,}기  NPV: {df.loc[sel,'투자가치NPV_천원'].sum():,.0f} 천원")

        # P2: 통합 AHP+GA
        print(f"\n[P2] 통합 AHP+GA  (집단={GA_POP_SIZE}, 세대={GA_GENERATIONS})")
        sel = solve_ga(df, V, budget, mandatory_mask)
        selections['P2 통합AHP-GA'] = sel
        kpi_rows.append(compute_kpi(df, sel, V, 'P2 통합AHP-GA'))
        print(f"  선택: {sel.sum():,}기  NPV: {df.loc[sel,'투자가치NPV_천원'].sum():,.0f} 천원")

        scenario_results.append({
            'ratio':      ratio,
            'budget':     budget,
            'opt_budget': opt_budget,
            'kpi_rows':   kpi_rows,
            'selections': selections,
        })

        # KPI 출력
        print(f"\n[KPI 비교 - {ratio*100:.0f}% 시나리오]")
        kpi_df = pd.DataFrame(kpi_rows).set_index('방법').T
        print(kpi_df.to_string())

    # 보고서 저장
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = DATA_DIR / f"optimization_results_{ts}.xlsx"
    save_report(scenario_results, df, V, out_path)

    print("\n" + "=" * 70)
    print("  완료")
    print("=" * 70)
