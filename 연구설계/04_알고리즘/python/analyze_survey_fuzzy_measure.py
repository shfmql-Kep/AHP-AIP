from __future__ import annotations

import itertools
import math
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


FUZZY_ITEMS = ["투자가치", "투자효율", "SAIDI저감", "고장예측저감", "안전영향", "환경영향"]
CRITERIA_ITEMS = ["경제성", "신뢰도", "안전·환경"]


def geometric_mean(values: list[float]) -> float:
    return math.exp(sum(math.log(value) for value in values) / len(values))


def weights_from_pairwise(items: list[str], pairs: list[tuple[str, str, float]]) -> list[float]:
    # 쌍대비교 행렬을 만들고 행 기하평균법으로 가중치를 산정한다.
    n = len(items)
    index = {item: i for i, item in enumerate(items)}
    matrix = [[1.0 for _ in range(n)] for _ in range(n)]
    for left, right, value in pairs:
        i = index[left]
        j = index[right]
        matrix[i][j] = value
        matrix[j][i] = 1.0 / value

    row_gm = []
    for i in range(n):
        product = 1.0
        for j in range(n):
            product *= matrix[i][j]
        row_gm.append(product ** (1.0 / n))
    total = sum(row_gm)
    return [value / total for value in row_gm]


def solve_lambda(weights: list[float]) -> float:
    # Sugeno λ-퍼지측도 조건 Π(1+λw_i)=1+λ 의 비영해를 구한다.
    total = sum(weights)
    if abs(total - 1.0) < 1e-10:
        return 0.0

    def f(lam: float) -> float:
        product = 1.0
        for weight in weights:
            product *= 1.0 + lam * weight
        return product - (1.0 + lam)

    if total > 1.0:
        # λ는 음수이며, 모든 1+λw_i가 양수여야 한다.
        lower_bound = max(-1.0 / max(weights) + 1e-12, -1e12)
        upper_bound = -1e-12
    else:
        lower_bound = 1e-12
        upper_bound = 1.0
        while f(lower_bound) * f(upper_bound) > 0:
            upper_bound *= 2.0
            if upper_bound > 1e12:
                raise RuntimeError("λ 양수 해의 상한을 찾지 못했습니다.")

    # 음수 해에서 경계가 -1에 붙는 경우를 안정적으로 처리한다.
    lo = lower_bound
    hi = upper_bound
    if f(lo) * f(hi) > 0:
        steps = 20000
        grid = [lo + (hi - lo) * k / steps for k in range(steps + 1)]
        prev = grid[0]
        prev_f = f(prev)
        found = False
        for point in grid[1:]:
            value = f(point)
            if prev_f * value <= 0:
                lo, hi = prev, point
                found = True
                break
            prev, prev_f = point, value
        if not found:
            raise RuntimeError(f"λ 해를 포함하는 구간을 찾지 못했습니다. sum={total}, f(lo)={f(lower_bound)}, f(hi)={f(upper_bound)}")

    for _ in range(300):
        mid = (lo + hi) / 2.0
        if f(lo) * f(mid) <= 0:
            hi = mid
        else:
            lo = mid
    return (lo + hi) / 2.0


def fuzzy_measure(subset: tuple[int, ...], weights: list[float], lam: float) -> float:
    if not subset:
        return 0.0
    if abs(lam) < 1e-12:
        return sum(weights[i] for i in subset)
    product = 1.0
    for index in subset:
        product *= 1.0 + lam * weights[index]
    return (product - 1.0) / lam


def shapley_values(weights: list[float], lam: float) -> list[float]:
    n = len(weights)
    values = []
    for i in range(n):
        shapley = 0.0
        others = [j for j in range(n) if j != i]
        for size in range(n):
            for subset in itertools.combinations(others, size):
                coefficient = math.factorial(size) * math.factorial(n - size - 1) / math.factorial(n)
                before = fuzzy_measure(tuple(subset), weights, lam)
                after = fuzzy_measure(tuple(sorted(subset + (i,))), weights, lam)
                shapley += coefficient * (after - before)
        values.append(shapley)
    return values


def choquet_integral(scores: list[float], weights: list[float], lam: float) -> float:
    # 점수를 오름차순으로 정렬한 뒤 Choquet 적분을 계산한다.
    order = sorted(range(len(scores)), key=lambda index: scores[index])
    sorted_scores = [scores[index] for index in order]
    result = 0.0
    prev = 0.0
    for rank, score in enumerate(sorted_scores):
        active_subset = tuple(order[rank:])
        result += (score - prev) * fuzzy_measure(active_subset, weights, lam)
        prev = score
    return result


def read_raw_rows(workbook_path: Path) -> list[tuple[int, str, str, str, float]]:
    workbook = load_workbook(workbook_path, data_only=True)
    raw_sheet_name = next((name for name in workbook.sheetnames if name.endswith("RAW")), None)
    if raw_sheet_name is None:
        raise KeyError("RAW 응답 시트를 찾지 못했습니다.")

    sheet = workbook[raw_sheet_name]
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        respondent, group, section, item, value = row
        if respondent is None:
            continue
        rows.append((int(respondent), str(group), str(section), str(item), float(value)))
    return rows


def aggregate_ahp(rows: list[tuple[int, str, str, str, float]]) -> tuple[dict[str, float], dict[str, float]]:
    pair_values: dict[tuple[str, str], list[float]] = defaultdict(list)
    for _, _, section, item, value in rows:
        if section.startswith("AHP"):
            pair_values[(section, item)].append(value)

    criterion_pairs = []
    for item, left, right in [
        ("경제/신뢰", "경제성", "신뢰도"),
        ("경제/안전환경", "경제성", "안전·환경"),
        ("신뢰/안전환경", "신뢰도", "안전·환경"),
    ]:
        criterion_pairs.append((left, right, geometric_mean(pair_values[("AHP기준", item)])))
    criterion_weights_list = weights_from_pairwise(CRITERIA_ITEMS, criterion_pairs)
    criterion_weights = dict(zip(CRITERIA_ITEMS, criterion_weights_list))

    sub_weights: dict[str, float] = {}
    for item, left, right, parent in [
        ("NPV/BCR", "투자가치", "투자효율", "경제성"),
        ("SAIDI/고장", "SAIDI저감", "고장예측저감", "신뢰도"),
        ("안전/환경", "안전영향", "환경영향", "안전·환경"),
    ]:
        local_weights = weights_from_pairwise([left, right], [(left, right, geometric_mean(pair_values[("AHP하위", item)]))])
        sub_weights[left] = criterion_weights[parent] * local_weights[0]
        sub_weights[right] = criterion_weights[parent] * local_weights[1]

    return criterion_weights, sub_weights


def aggregate_fuzzy_scores(rows: list[tuple[int, str, str, str, float]]) -> dict[str, list[float]]:
    scores = {item: [] for item in FUZZY_ITEMS}
    for _, _, section, item, value in rows:
        if section == "퍼지측도" and item in scores:
            scores[item].append(value)
    return scores


def normalize_for_target_sum(base_weights: list[float], target_sum: float) -> list[float]:
    factor = target_sum / sum(base_weights)
    return [weight * factor for weight in base_weights]


def main() -> None:
    survey_dir = Path("Survey")
    response_files = [path for path in survey_dir.glob("*.xlsx") if "응답완료" in path.name]
    if not response_files:
        raise FileNotFoundError("Survey 폴더에서 응답완료 xlsx 파일을 찾지 못했습니다.")
    workbook_path = sorted(response_files, key=lambda path: path.stat().st_mtime)[-1]

    rows = read_raw_rows(workbook_path)
    respondents = sorted({row[0] for row in rows})
    criterion_weights, ahp_weights = aggregate_ahp(rows)
    fuzzy_scores = aggregate_fuzzy_scores(rows)

    average_scores = {item: sum(values) / len(values) for item, values in fuzzy_scores.items()}
    wa = [average_scores[item] / 6.0 for item in FUZZY_ITEMS]
    lam = solve_lambda(wa)
    shapley = shapley_values(wa, lam)

    print(f"분석 파일: {workbook_path}")
    print(f"응답자 수: {len(respondents)}명")
    print(f"RAW 행 수: {len(rows)}행")
    print()

    print("[AHP 전역 하위지표 가중치]")
    for item in FUZZY_ITEMS:
        print(f"{item}\t{ahp_weights[item]:.6f}")
    print(f"합계\t{sum(ahp_weights.values()):.6f}")
    print()

    print("[퍼지측도 점수와 Wa=score/6]")
    print("항목\t평균점수\tWa\t최소\t최대")
    for item, weight in zip(FUZZY_ITEMS, wa):
        values = fuzzy_scores[item]
        print(f"{item}\t{average_scores[item]:.4f}\t{weight:.6f}\t{min(values):.0f}\t{max(values):.0f}")
    print(f"ΣWa\t\t{sum(wa):.6f}")
    print(f"λ\t\t{lam:.8f}")
    print()

    print("[Shapley 보정 가중치]")
    print("항목\tAHP_Wr\tShapley_phi\t차이")
    for item, phi in zip(FUZZY_ITEMS, shapley):
        print(f"{item}\t{ahp_weights[item]:.6f}\t{phi:.6f}\t{phi - ahp_weights[item]:+.6f}")
    print(f"Shapley 합계\t\t{sum(shapley):.6f}")
    print()

    print("[목표 ΣWa별 λ 민감도: 단순 스케일 조정]")
    print("target_sum\tλ\tShapley_min\tShapley_max\tmax-min")
    for target_sum in [1.0, 1.2, 1.5, 2.0, 2.5, 3.0, sum(wa)]:
        adjusted = normalize_for_target_sum(wa, target_sum)
        adjusted_lambda = solve_lambda(adjusted)
        adjusted_shapley = shapley_values(adjusted, adjusted_lambda)
        print(f"{target_sum:.3f}\t{adjusted_lambda:.6f}\t{min(adjusted_shapley):.6f}\t{max(adjusted_shapley):.6f}\t{max(adjusted_shapley)-min(adjusted_shapley):.6f}")
    print()

    print("[Choquet 이중보정 예시: 평균 정규점수를 점수로 가정]")
    normalized_scores = [average_scores[item] / 6.0 for item in FUZZY_ITEMS]
    wsm_ahp = sum(ahp_weights[item] * score for item, score in zip(FUZZY_ITEMS, normalized_scores))
    wsm_shapley = sum(phi * score for phi, score in zip(shapley, normalized_scores))
    choquet_only = choquet_integral(normalized_scores, wa, lam)
    choquet_after_shapley_like = choquet_integral([phi * score for phi, score in zip(shapley, normalized_scores)], wa, lam)
    print(f"AHP 가중합\t{wsm_ahp:.6f}")
    print(f"Shapley 가중합\t{wsm_shapley:.6f}")
    print(f"Choquet 단독\t{choquet_only:.6f}")
    print(f"Shapley×score 후 Choquet\t{choquet_after_shapley_like:.6f}")
    print()

    print("[AHP와 Fuzzy를 결합하는 대안 가중치]")
    ahp_times_wa = {item: ahp_weights[item] * (average_scores[item] / 6.0) for item in FUZZY_ITEMS}
    ahp_times_wa_total = sum(ahp_times_wa.values())
    ahp_times_shapley = {item: ahp_weights[item] * phi for item, phi in zip(FUZZY_ITEMS, shapley)}
    ahp_times_shapley_total = sum(ahp_times_shapley.values())
    print("항목\tAHP×Wa 정규화\tAHP×Shapley 정규화")
    for item in FUZZY_ITEMS:
        print(f"{item}\t{ahp_times_wa[item] / ahp_times_wa_total:.6f}\t{ahp_times_shapley[item] / ahp_times_shapley_total:.6f}")
    print()

    print("[부모 기준 내부 2지표별 λ: 전역 6지표 대신 단계별 적용]")
    for parent, left, right in [
        ("경제성", "투자가치", "투자효율"),
        ("신뢰도", "SAIDI저감", "고장예측저감"),
        ("안전·환경", "안전영향", "환경영향"),
    ]:
        left_wa = average_scores[left] / 6.0
        right_wa = average_scores[right] / 6.0
        pair_lambda = solve_lambda([left_wa, right_wa])
        pair_shapley = shapley_values([left_wa, right_wa], pair_lambda)
        print(
            f"{parent}\tΣWa={left_wa + right_wa:.6f}\tλ={pair_lambda:.6f}\t"
            f"{left}={pair_shapley[0]:.6f}\t{right}={pair_shapley[1]:.6f}"
        )


if __name__ == "__main__":
    main()
