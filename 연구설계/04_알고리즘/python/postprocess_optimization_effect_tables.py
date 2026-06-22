"""단일지표 최적화 결과를 논문용 투자효과 표로 재구성한다.

지표 정의
---------
1. Risk 저감량:
   교체시점부터 설계수명 동안의 기존 설비 누적 Risk와 신규 설비 누적 Risk의 차이.
   원천값은 `data/pof_5yr_output.xlsx`의 `risk_reduction_YYYY_kkrw`이며,
   MATLAB 결과의 `selected_assets.risk_reduction_kkrw`로 전달된다.
2. 투자가치:
   Risk 저감량 - 투자비용.
3. 투자효율:
   Risk 저감량 / 투자비용.

출력 대상
---------
- 설비 전체 합산 연도별/합산 표
- 개별 설비군별 연도별/합산 표
- Risk greedy를 베이스라인으로 한 기법별 증가율 표
"""

from __future__ import annotations

from pathlib import Path
from typing import Iterable

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent.parent
OUTPUT_XLSX = BASE_DIR / "outputs" / "single_metric_optimization_matlab.xlsx"

YEARS = [2026, 2027, 2028, 2029, 2030]
BASELINE_METHOD = "risk_greedy"
METHOD_ORDER = [
    "risk_greedy",
    "investment_value_greedy",
    "investment_value_ilp",
    "investment_value_nsga2",
]
METHOD_LABELS = {
    "risk_greedy": "베이스라인(Risk 그리디)",
    "investment_value_greedy": "투자가치 그리디",
    "investment_value_ilp": "투자가치 ILP",
    "investment_value_nsga2": "투자가치 NSGA-II",
}
METRICS = [
    ("investment_count", "투자대수"),
    ("investment_cost_kkrw", "투자비용"),
    ("risk_reduction_kkrw", "Risk 저감량"),
    ("investment_value_kkrw", "투자가치"),
    ("saidi_min", "SAIDI"),
    ("investment_efficiency", "투자효율"),
]
BROKEN_TOTAL_LABEL = "\u00c7\u0150\u00bb\u0119"


def method_sort_key(method: str) -> int:
    """정해진 방법론 순서에 맞는 정렬 키를 반환한다."""
    try:
        return METHOD_ORDER.index(str(method))
    except ValueError:
        return len(METHOD_ORDER)


def safe_divide(numerator: float, denominator: float) -> float:
    """0 나누기와 결측값을 방지한다."""
    if denominator == 0 or pd.isna(denominator):
        return 0.0
    return float(numerator) / float(denominator)


def pct_change(value: float, baseline: float) -> float:
    """베이스라인 대비 변화율을 계산한다.

    투자가치가 음수일 수 있으므로 분모는 베이스라인의 절댓값을 사용한다.
    """
    return safe_divide(float(value) - float(baseline), abs(float(baseline))) * 100.0


def format_number(value: float, metric: str | None = None) -> str:
    """표시용 숫자 문자열을 만든다."""
    if pd.isna(value):
        return "-"
    value = float(value)
    if metric == "investment_count":
        return f"{value:,.0f}"
    if metric == "investment_efficiency":
        return f"{value:.4f}"
    if metric == "saidi_min":
        return f"{value:.6f}"
    if abs(value) >= 1000:
        return f"{value:,.0f}"
    return f"{value:,.3f}"


def normalize_selected_assets(selected: pd.DataFrame) -> pd.DataFrame:
    """MATLAB 선택 상세표에 새 지표 컬럼을 보강한다."""
    out = selected.copy()

    if "risk_reduction_kkrw" not in out.columns:
        # 과거 출력과의 호환용: 기존 investment_value가 총 Risk 저감 편익이던 경우.
        out["risk_reduction_kkrw"] = out["investment_value_kkrw"].astype(float)
        out["investment_value_kkrw"] = (
            out["risk_reduction_kkrw"].astype(float)
            - out["replacement_cost_kkrw"].astype(float)
        )

    out["risk_reduction_kkrw"] = out["risk_reduction_kkrw"].astype(float)
    out["replacement_cost_kkrw"] = out["replacement_cost_kkrw"].astype(float)
    out["investment_value_kkrw"] = out["investment_value_kkrw"].astype(float)
    out["investment_efficiency"] = (
        out["risk_reduction_kkrw"]
        / out["replacement_cost_kkrw"].replace(0, np.nan)
    ).fillna(0.0)

    if "bcr" in out.columns:
        out = out.drop(columns=["bcr"])

    return out


def summarize_effect(df: pd.DataFrame, group_cols: list[str]) -> pd.DataFrame:
    """투자효과를 주어진 그룹 기준으로 집계한다."""
    grouped = (
        df.groupby(group_cols, as_index=False)
        .agg(
            investment_count=("asset_id", "count"),
            investment_cost_kkrw=("replacement_cost_kkrw", "sum"),
            risk_reduction_kkrw=("risk_reduction_kkrw", "sum"),
            investment_value_kkrw=("investment_value_kkrw", "sum"),
            saidi_min=("saidi_at_replacement_min", "sum"),
        )
    )
    grouped["investment_efficiency"] = (
        grouped["risk_reduction_kkrw"]
        / grouped["investment_cost_kkrw"].replace(0, np.nan)
    ).fillna(0.0)
    return grouped


def append_total_rows(summary: pd.DataFrame, group_cols_without_year: list[str]) -> pd.DataFrame:
    """연도별 표에 합산 행을 추가한다."""
    total = (
        summary.groupby(group_cols_without_year, as_index=False)
        .agg(
            investment_count=("investment_count", "sum"),
            investment_cost_kkrw=("investment_cost_kkrw", "sum"),
            risk_reduction_kkrw=("risk_reduction_kkrw", "sum"),
            investment_value_kkrw=("investment_value_kkrw", "sum"),
            saidi_min=("saidi_min", "sum"),
        )
    )
    total["investment_efficiency"] = (
        total["risk_reduction_kkrw"]
        / total["investment_cost_kkrw"].replace(0, np.nan)
    ).fillna(0.0)
    total["year"] = "합산"

    ordered_cols = list(summary.columns)
    total = total[ordered_cols]
    return pd.concat([summary, total], ignore_index=True)


def sort_annual_table(df: pd.DataFrame, extra_cols: Iterable[str] = ()) -> pd.DataFrame:
    """방법론·설비군·연도 순으로 표를 정렬한다."""
    out = df.copy()
    out["_method_order"] = out["method"].map(method_sort_key)
    out["_year_order"] = out["year"].apply(lambda x: 9999 if str(x) == "합산" else int(x))
    sort_cols = ["_method_order", *extra_cols, "_year_order"]
    out = out.sort_values(sort_cols).drop(columns=["_method_order", "_year_order"])
    return out


def make_effect_tables(selected: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """전체·설비군별 투자효과 원시 표를 만든다."""
    base = selected.rename(columns={"replacement_year": "year"})

    total_annual = summarize_effect(base, ["method", "year"])
    total_annual = append_total_rows(total_annual, ["method"])
    total_annual = sort_annual_table(total_annual)

    total_summary = total_annual[total_annual["year"].astype(str) == "합산"].drop(columns=["year"])
    total_summary = total_summary.sort_values(
        by="method",
        key=lambda s: s.map(method_sort_key),
    )

    type_annual = summarize_effect(base, ["method", "asset_type", "year"])
    type_annual = append_total_rows(type_annual, ["method", "asset_type"])
    type_annual = sort_annual_table(type_annual, extra_cols=["asset_type"])

    type_summary = type_annual[type_annual["year"].astype(str) == "합산"].drop(columns=["year"])
    type_summary = type_summary.sort_values(
        by=["method", "asset_type"],
        key=lambda s: s.map(method_sort_key) if s.name == "method" else s,
    )

    return {
        "effect_total_annual": total_annual,
        "effect_total_summary": total_summary,
        "effect_type_annual": type_annual,
        "effect_type_summary": type_summary,
    }


def make_korean_annual(df: pd.DataFrame, include_asset_type: bool) -> pd.DataFrame:
    """사용자가 제시한 양식에 맞춘 한국어 연도별 표를 만든다."""
    out = df.copy()
    out["기법"] = out["method"].map(METHOD_LABELS).fillna(out["method"])
    out["구분"] = out["year"].astype(str).str.replace(".0", "", regex=False)
    rename_map = {
        "investment_count": "투자대수",
        "investment_cost_kkrw": "투자비용",
        "risk_reduction_kkrw": "Risk 저감량",
        "investment_value_kkrw": "투자가치",
        "saidi_min": "SAIDI",
        "investment_efficiency": "투자효율",
    }
    out = out.rename(columns=rename_map)
    columns = ["기법"]
    if include_asset_type:
        columns.append("asset_type")
    columns += ["구분", "투자대수", "투자비용", "Risk 저감량", "투자가치", "SAIDI", "투자효율"]
    return out[columns]


def make_baseline_comparison(summary: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """설비 전체 기준의 베이스라인 대비 변화율 표를 만든다."""
    baseline = summary[summary["method"] == BASELINE_METHOD].iloc[0]
    raw_rows: list[dict[str, float | str]] = []
    display_rows: list[dict[str, str]] = []

    for metric_col, metric_label in METRICS:
        baseline_value = float(baseline[metric_col])
        raw_row: dict[str, float | str] = {
            "metric": metric_col,
            "metric_label": metric_label,
            "baseline_value": baseline_value,
        }
        display_row: dict[str, str] = {
            "구분": metric_label,
            METHOD_LABELS[BASELINE_METHOD]: format_number(baseline_value, metric_col),
        }

        for method in METHOD_ORDER:
            method_rows = summary[summary["method"] == method]
            if method_rows.empty:
                continue
            value = float(method_rows.iloc[0][metric_col])
            change = pct_change(value, baseline_value)
            raw_row[f"{method}_value"] = value
            raw_row[f"{method}_change_vs_baseline_pct"] = change
            if method != BASELINE_METHOD:
                display_row[METHOD_LABELS[method]] = (
                    f"{format_number(value, metric_col)} ({change:+.2f}%)"
                )

        raw_rows.append(raw_row)
        display_rows.append(display_row)

    return pd.DataFrame(raw_rows), pd.DataFrame(display_rows)


def make_type_baseline_comparison(type_summary: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """설비군별 베이스라인 대비 변화율 표를 만든다."""
    raw_rows: list[dict[str, float | str]] = []
    display_rows: list[dict[str, str]] = []

    for asset_type, sub in type_summary.groupby("asset_type", sort=True):
        baseline_rows = sub[sub["method"] == BASELINE_METHOD]
        if baseline_rows.empty:
            continue
        baseline = baseline_rows.iloc[0]

        for metric_col, metric_label in METRICS:
            baseline_value = float(baseline[metric_col])
            display_row: dict[str, str] = {
                "asset_type": str(asset_type),
                "구분": metric_label,
                METHOD_LABELS[BASELINE_METHOD]: format_number(baseline_value, metric_col),
            }

            for method in METHOD_ORDER:
                method_rows = sub[sub["method"] == method]
                if method_rows.empty:
                    continue
                value = float(method_rows.iloc[0][metric_col])
                change = pct_change(value, baseline_value)
                raw_rows.append(
                    {
                        "asset_type": asset_type,
                        "metric": metric_col,
                        "metric_label": metric_label,
                        "baseline_method": BASELINE_METHOD,
                        "baseline_value": baseline_value,
                        "method": method,
                        "method_value": value,
                        "change_vs_baseline_pct": change,
                    }
                )
                if method != BASELINE_METHOD:
                    display_row[METHOD_LABELS[method]] = (
                        f"{format_number(value, metric_col)} ({change:+.2f}%)"
                    )

            display_rows.append(display_row)

    return pd.DataFrame(raw_rows), pd.DataFrame(display_rows)


def write_sheets(sheets: dict[str, pd.DataFrame]) -> None:
    """기존 최적화 workbook에 결과 시트를 추가·교체한다."""
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)


def format_workbook(sheet_names: list[str]) -> None:
    """결과 시트의 가독성을 높인다."""
    wb = load_workbook(OUTPUT_XLSX)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    alt_fill = PatternFill("solid", fgColor="EAF2F8")

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_idx in range(2, ws.max_row + 1):
            if row_idx % 2 == 0:
                for cell in ws[row_idx]:
                    cell.fill = alt_fill

        for col_idx in range(1, ws.max_column + 1):
            header = str(ws.cell(1, col_idx).value or "")
            max_len = len(header)
            for row_idx in range(2, min(ws.max_row, 80) + 1):
                value = ws.cell(row_idx, col_idx).value
                if value is not None:
                    max_len = max(max_len, len(str(value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 34)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, str) and cell.value == BROKEN_TOTAL_LABEL:
                    cell.value = "합산"
                if isinstance(cell.value, float):
                    cell.number_format = "#,##0.000000"
                elif isinstance(cell.value, int):
                    cell.number_format = "#,##0"

    wb.save(OUTPUT_XLSX)


def main() -> None:
    """투자효과 표를 생성한다."""
    if not OUTPUT_XLSX.exists():
        raise FileNotFoundError(f"최적화 결과 파일이 없습니다: {OUTPUT_XLSX}")

    selected = pd.read_excel(OUTPUT_XLSX, sheet_name="selected_assets")
    selected = normalize_selected_assets(selected)

    effect_tables = make_effect_tables(selected)
    baseline_compare_total, baseline_compare_display = make_baseline_comparison(
        effect_tables["effect_total_summary"]
    )
    baseline_compare_type, baseline_compare_type_display = make_type_baseline_comparison(
        effect_tables["effect_type_summary"]
    )

    table_total_annual_kr = make_korean_annual(
        effect_tables["effect_total_annual"], include_asset_type=False
    )
    table_type_annual_kr = make_korean_annual(
        effect_tables["effect_type_annual"], include_asset_type=True
    )

    sheets = {
        "selected_assets": selected,
        "effect_total_annual": effect_tables["effect_total_annual"],
        "effect_total_summary": effect_tables["effect_total_summary"],
        "effect_type_annual": effect_tables["effect_type_annual"],
        "effect_type_summary": effect_tables["effect_type_summary"],
        "baseline_compare_total": baseline_compare_total,
        "baseline_compare_display": baseline_compare_display,
        "baseline_compare_type": baseline_compare_type,
        "baseline_compare_type_display": baseline_compare_type_display,
        "table_total_annual_kr": table_total_annual_kr,
        "table_type_annual_kr": table_type_annual_kr,
    }

    write_sheets(sheets)
    format_workbook(list(sheets))

    print(f"updated: {OUTPUT_XLSX}")
    print("added sheets:")
    for sheet_name in sheets:
        print(f"- {sheet_name}: {len(sheets[sheet_name]):,} rows")


if __name__ == "__main__":
    main()
